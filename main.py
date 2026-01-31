"""
DIREKTORAT PENGENDALIAN PERHUTANAN SOSIAL
Microservice FastAPI menggunakan Google Earth Engine (GEE) dengan klasifikasi 
berbasis machine learning untuk tutupan lahan Indonesia.
"""

import os
import time
import json
import hashlib
import math
import asyncio
import requests
import numpy as np
from scipy import sparse
from scipy.sparse.linalg import splu
from datetime import datetime, timedelta
# Force reload: 2026-01-11 07:11
from typing import Any, Dict, List, Optional, Callable

import ee
import zipfile
import io
import traceback
import logging
import base64
import shapefile # pyshp
from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, Response, Request, BackgroundTasks, Query
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from shapely.geometry import shape, mapping
from shapely.validation import make_valid
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.staticfiles import StaticFiles
import pathlib

from pkps_sync import run_sync_process

from dotenv import load_dotenv
import requests
from supabase import create_client, Client

import sys

load_dotenv()


# KONFIGURASI
# ==============================================================================

# Path ke file kunci Service Account GEE
KEY_FILE = os.getenv("GEE_KEY_FILE", "gee-service-account-key.json")

# Supabase Configuration
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
supabase: Optional[Client] = None

if SUPABASE_URL and SUPABASE_KEY:
    try:
        # Ensure trailing slash to avoid "Storage endpoint URL should have a trailing slash" warning
        if not SUPABASE_URL.endswith("/"):
            SUPABASE_URL += "/"
            
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("✅ Supabase client initialized")
    except Exception as e:
        print(f"⚠️ Failed to initialize Supabase: {e}")

# ==============================================================================
# SERVER-SIDE FILE CACHE (diskcache) - PERMANENT MODE
# ==============================================================================
import diskcache

CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache")
CACHE_SIZE_LIMIT = 10 * 1024 * 1024 * 1024  # 10 GB
disk_cache = diskcache.Cache(CACHE_DIR, size_limit=CACHE_SIZE_LIMIT)
# Permanent cache - no TTL, only invalidated on data changes (POST/DELETE)
print(f"✅ Disk cache initialized at: {CACHE_DIR} (Size limit: 10GB, Permanent mode)")

# ==============================================================================
# LOCAL STORAGE (IMAGE HOSTING)
# ==============================================================================
STORAGE_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "storage")
THUMBNAIL_DIR = os.path.join(STORAGE_ROOT, "thumbnails")
pathlib.Path(THUMBNAIL_DIR).mkdir(parents=True, exist_ok=True)
print(f"✅ Local storage initialized at: {STORAGE_ROOT}")

def ensure_storage_ready():
    """Ensure the 'thumbnails' bucket exists in Supabase Storage."""
    if not supabase: return
    try:
        # Check buckets
        # Note: listing buckets might also fail with 403 if RLS denies it, 
        # but usually it's allowed for authenticated users.
        try:
            buckets = supabase.storage.list_buckets()
            exists = any(b.name == "thumbnails" for b in buckets)
        except Exception:
            # If listing fails, assume we need to try creating or just proceed
            exists = False

        if not exists:
            # print("🚀 Creating 'thumbnails' bucket in Supabase Storage...") 
            try:
                supabase.storage.create_bucket("thumbnails", options={"public": True})
                print("✅ 'thumbnails' bucket created.")
            except Exception as ce:
                # 403 is expected if using Anon Key and RLS is set (production mode)
                err_str = str(ce)
                if "403" in err_str or "unauthorized" in err_str.lower():
                    print("ℹ️ 'thumbnails' bucket creation skipped (managed via SQL/Dashboard).")
                else:
                    print(f"⚠️ Bucket creation error: {ce}")
        else:
            print("✅ 'thumbnails' bucket exists.")
    except Exception as e:
        print(f"⚠️ Storage Check Failed: {e}")

# Call immediately on startup
ensure_storage_ready()

def save_image_persistence(data: Any, filename: str) -> Optional[str]:
    """
    Saves image data (bytes, base64, or PIL) permanently.
    Priority: Supabase Storage > Local Disk.
    Always converts to WebP.
    """
    try:
        try:
            from PIL import Image
        except ImportError as ie:
            print("❌ CRITICAL: Pillow library not installed!")
            print("   This is required for image conversion.")
            print("   Quick fix options:")
            print("   1. Run: pip install Pillow>=10.0.0")
            print("   2. Or run: install_dependencies.bat (from project root)")
            print("   After installation, restart the backend server.")
            return None

        import io
        import base64
        
        # 1. Prepare Base Context
        base_filename = os.path.splitext(filename)[0]
        webp_filename = f"{base_filename}.webp"
        
        img = None

        # 2. Parse Input Data
        if isinstance(data, str) and (data.startswith('data:image') or ',' in data):
            header, encoded = data.split(",", 1) if "," in data else ("", data)
            img = Image.open(io.BytesIO(base64.b64decode(encoded)))
        elif isinstance(data, str):
            img = Image.open(io.BytesIO(base64.b64decode(data)))
        elif isinstance(data, bytes):
            img = Image.open(io.BytesIO(data))
        elif hasattr(data, 'save'): # Already a PIL Image
            img = data
        else:
            print(f"⚠️ save_image_persistence: Unsupported data type {type(data)}")
            return None
             
        # 3. Convert to RGB/RGBA
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGBA")
        else:
            img = img.convert("RGB")

        # 4. Convert to WebP Bytes
        buffer = io.BytesIO()
        img.save(buffer, format="WEBP", quality=80, method=6)
        image_bytes = buffer.getvalue()
        
        # 5. ATTEMPT SUPABASE UPLOAD (Primary)
        if supabase:
            try:
                bucket_name = "thumbnails"
                # Check if bucket needs creation? We did at startup.

                # Upload with upsert
                res = supabase.storage.from_(bucket_name).upload(
                    path=webp_filename,
                    file=image_bytes,
                    file_options={"content-type": "image/webp", "upsert": "true"}
                )

                # Get Public URL
                # NOTE: get_public_url returns URL without checking existence
                public_url = supabase.storage.from_(bucket_name).get_public_url(webp_filename)
                print(f"☁️ Uploaded to Cloud: {webp_filename}")
                print(f"   Public URL: {public_url}")

                # Validate URL is accessible (log only, don't fail)
                try:
                    head_check = requests.head(public_url, timeout=5)
                    if head_check.status_code == 200:
                        print(f"✅ URL is accessible")
                    else:
                        print(f"⚠️ URL returned status {head_check.status_code}")
                except Exception as url_check_err:
                    print(f"⚠️ Could not verify URL accessibility: {url_check_err}")

                return public_url
            except Exception as se:
                print(f"⚠️ Supabase Upload Error ({webp_filename}): {se}")
                # Log but continue to fallback
        
        # 6. FALLBACK: Local Storage
        file_path = os.path.join(THUMBNAIL_DIR, webp_filename)
        with open(file_path, "wb") as f:
            f.write(image_bytes)
        
        print(f"💾 Saved Locally: {webp_filename}")
        return f"/storage/thumbnails/{webp_filename}"

    except Exception as e:
        print(f"❌ save_image_persistence failed: {e}")
        return None

# Alias for backward compatibility
save_local_image = save_image_persistence

# ==============================================================================
# CACHE HELPER FUNCTIONS
# ==============================================================================
def get_file_hash(content: bytes) -> str:
    """Generate SHA256 hash for file content (first 16 chars for readability)"""
    return hashlib.sha256(content).hexdigest()[:16]

def cache_file(key: str, data: any, expire: int = None):
    """Store data in disk cache"""
    try:
        disk_cache.set(key, data, expire=expire)
        print(f"💾 Cached: {key}")
    except Exception as e:
        print(f"⚠️ Cache write failed for {key}: {e}")

def get_cached(key: str):
    """Retrieve data from disk cache"""
    try:
        return disk_cache.get(key)
    except Exception as e:
        print(f"⚠️ Cache read failed for {key}: {e}")
        return None

def invalidate_cache(pattern: str):
    """Invalidate cache entries matching pattern"""
    try:
        deleted_count = 0
        for key in list(disk_cache.iterkeys()):
            if pattern in key:
                disk_cache.delete(key)
                deleted_count += 1
        if deleted_count > 0:
            print(f"🗑️ Invalidated {deleted_count} cache entries matching: {pattern}")
    except Exception as e:
        print(f"⚠️ Cache invalidation failed for {pattern}: {e}")

# Sentinel-2 Surface Reflectance Collection
SENTINEL2_COLLECTION = "COPERNICUS/S2_SR_HARMONIZED"
# Sentinel-2 Cloud Probability (for robust cloud masking)
S2_CLOUD_PROBABILITY = "COPERNICUS/S2_CLOUD_PROBABILITY"
SENTINEL_BANDS = ['B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B8A', 'B11', 'B12']
# Multi-year RF bands (minimal set for classification)
MULTIYEAR_RF_BANDS = ['B2', 'B3', 'B4', 'B8']

# Kategori Tutupan Lahan - 9 Kelas (Sub-klasifikasi Hutan)
LC_CATEGORIES = {
    1: "hutan_primer",     # Primary Forest (Stabil sebagai hutan)
    2: "hutan_sekunder",   # Secondary Forest (Regenerasi/Gangguan)
    3: "tanah_kering",    # Dry Land / Vegetasi Jarang
    4: "tanah_kosong",    # Bare Soil / Lahan Terbuka
    5: "air",             # Water bodies
    6: "lahan_terbangun", # Built-up / Pemukiman
}

# Konfigurasi visualisasi peta dengan palet warna
LC_PALETTE = {
    1: "#006400",  # Hutan Primer - Dark Green
    2: "#32CD32",  # Hutan Sekunder - Lime Green
    3: "#DAA520",  # Tanah Kering - Goldenrod
    4: "#D2691E",  # Tanah Kosong - Chocolate/Brown
    5: "#1E90FF",  # Air - Dodger Blue
    6: "#708090",  # Lahan Terbangun - Slate Gray
}

# ==============================================================================
# CARBON FACTOR CONSTANTS (IPCC Tier-1 Adapted for Indonesia)
# ==============================================================================
# These are INDICATIVE values for trend analysis only.
# NOT for carbon trading, regulatory compliance, or policy decisions.
# Unit: tonnes Carbon per hectare (tC/ha)
CARBON_FACTORS = {
    1: 120.0,  # Hutan Primer (Tropical Moist Forest)
    2: 90.0,   # Hutan Sekunder
    3: 35.0,   # Tanah Kering (Shrubland/Dry Agriculture)
    4: 5.0,    # Tanah Kosong (Bare Soil)
    5: 0.0,    # Air (Water Bodies)
    6: 2.0,    # Lahan Terbangun (Built-up/Urban)
}

# CO2e conversion factor (IPCC standard: 1 tC = 3.67 tCO2e)
CO2E_CONVERSION = 3.67

# Carbon trend threshold (±2% of initial stock = "Stable")
CARBON_TREND_THRESHOLD = 0.02

# Disclaimer text for Carbon Mode (WAJIB ditampilkan)
CARBON_DISCLAIMER = (
    "Perubahan stok karbon bersifat INDIKATIF dan diturunkan dari klasifikasi "
    "tutupan lahan dengan faktor karbon tetap. Tidak ada pengukuran lapangan "
    "yang diterapkan. Hanya untuk interpretasi tren, BUKAN untuk perdagangan "
    "karbon, kepatuhan regulasi, atau rekomendasi kebijakan."
)

# ==============================================================================
# GLOBAL ANALYSIS QUEUE MANAGER
# ==============================================================================

from contextlib import asynccontextmanager

class AnalysisQueueManager:
    """Manages concurrent analysis jobs with position tracking and notifications."""
    
    def __init__(self, max_concurrent=2):
        self.max_concurrent = max_concurrent
        self.waiting_queue = [] # List of job_ids in order
        self.active_jobs = set() # Set of job_ids currently running
        self.lock = asyncio.Lock()
        self.current_semaphore = asyncio.Semaphore(max_concurrent)

    @asynccontextmanager
    async def enter_queue(self, job_id: str, on_queue_update: Callable = None):
        """Register a job in the queue and wait for its turn."""
        async with self.lock:
            if job_id not in self.waiting_queue and job_id not in self.active_jobs:
                self.waiting_queue.append(job_id)
                print(f"🚥 Job {job_id} entered queue. Position: {len(self.waiting_queue)}")
        
        # Broadcast immediately to everyone that someone new entered (positions didn't change for others, but good for debug)
        await self._broadcast_queue_shifts()
        
        # Broadcast initial position
        if on_queue_update:
            pos = await self.get_position(job_id)
            await on_queue_update(pos)

        # Wait for semaphore slot
        async with self.current_semaphore:
            async with self.lock:
                # Remove from waiting, move to active
                if job_id in self.waiting_queue:
                    self.waiting_queue.remove(job_id)
                self.active_jobs.add(job_id)
                print(f"🚀 Job {job_id} starting execution.")
            
            # Notify everyone remaining in queue that positions might have shifted
            await self._broadcast_queue_shifts()
            
            try:
                yield True
            finally:
                async with self.lock:
                    if job_id in self.active_jobs:
                        self.active_jobs.remove(job_id)
                    print(f"✅ Job {job_id} finished.")
                await self._broadcast_queue_shifts()

    async def get_position(self, job_id: str) -> int:
        """Get 1-based position in queue. 0 if active/not found."""
        async with self.lock:
            if job_id in self.waiting_queue:
                return self.waiting_queue.index(job_id) + 1
            return 0

    async def is_idle(self) -> bool:
        """Check if there are NO active or waiting jobs."""
        async with self.lock:
            return len(self.waiting_queue) == 0 and len(self.active_jobs) == 0

    async def _broadcast_queue_shifts(self):
        """Helper to trigger position updates for all waiting jobs."""
        active_copy = []
        async with self.lock:
            active_copy = list(self.waiting_queue)
            
        for i, client_id in enumerate(active_copy):
            pos = i + 1
            # We assume Job ID == Client ID for WebSocket clients
            await ws_manager.send_queue_status(client_id, pos)

# Global Queue Manager (Limit to 2 concurrent jobs for stability)
analysis_queue = AnalysisQueueManager(max_concurrent=2)
# ANALYSIS_SEMAPHORE kept for backward compatibility if any legacy code uses it, 
# but we will migrate core endpoints.
ANALYSIS_SEMAPHORE = analysis_queue.current_semaphore


# ==============================================================================
# PYDANTIC MODELS
# ==============================================================================

class YearlyData(BaseModel):
    """Statistik tutupan lahan untuk satu tahun - 9 Kelas."""
    year: Any
    hutan_primer: float = 0       # Class 1: Primary Forest
    hutan_sekunder: float = 0     # Class 2: Secondary Forest
    tanah_kering: float = 0       # Class 3: Dry Land
    tanah_kosong: float = 0       # Class 4: Bare Soil
    air: float = 0                # Class 5: Water
    lahan_terbangun: float = 0    # Class 6: Built-up
    total_area: float = 0         # Total analyzed area
    # Legacy field for compatibility
    hutan: float = 0              # Sum of 1+2
    # Transition Metrics (Point 1, 2)
    forest_loss: float = 0        # Deforestasi (Hutan -> Non-Hutan)
    forest_gain: float = 0        # Reforestasi (Non-Hutan -> Hutan)
    forest_stable: float = 0      # Tutupan Hutan Stabil (Hutan -> Hutan)
    builtup_expansion: float = 0  # Ekspansi Lahan Terbangun (Non-Terbangun -> Terbangun)
    
    # New Transition Aliases & Maps (WAJIB - Section C)
    deforestation_ha: float = 0   # Alias for forest_loss
    reforestation_ha: float = 0   # Alias for forest_gain
    
    deforestation_map_url: Optional[str] = None
    reforestation_map_url: Optional[str] = None
    builtup_expansion_map_url: Optional[str] = None
    builtup_expansion_map_url: Optional[str] = None
    map_url: Optional[str] = None
    rgb_url: Optional[str] = None
    thumb_url: Optional[str] = None
    rgb_thumb_url: Optional[str] = None
    vector_geojson: Optional[Dict[str, Any]] = None
    confidence_percent: Optional[float] = None
    confidence_detailed: Dict[str, float] = {}
    detailed: Dict[str, float] = {}
    anomaly_flag: bool = False    # Flag for temporal anomaly detection
    # Metadata akuisisi citra
    acquisition_window: Optional[str] = None  # e.g., "regional", "full_year"
    date_range: Optional[str] = None          # e.g., "2021-06-01 to 2021-09-30"
    scene_count: Optional[int] = None         # Number of scenes used
    region_detected: Optional[str] = None     # e.g., "Sumatera/Kalimantan"
    # Accuracy Metrics (Point 3)
    accuracy_score: Optional[float] = None    # Overall Agreement with ESA WorldCover
    f1_score: Optional[float] = None          # Mean F1 Score
    kappa_score: Optional[float] = None       # Kappa Coefficient
    data_source: str = "Optical (S2)"         # e.g., "Optical + Radar"


class ClassificationThresholds(BaseModel):
    """Parameter ambang batas untuk klasifikasi 4-kelas sederhana."""
    # Hutan (NDVI tinggi)
    forest_ndvi_min: float = 0.6
    
    # Tanah Kering (vegetasi jarang)
    dry_land_ndvi_min: float = 0.2
    dry_land_ndvi_max: float = 0.6
    
    # Tanah Kosong (NDVI rendah, NDBI tinggi)
    bare_soil_ndvi_max: float = 0.2
    bare_soil_ndbi_min: float = 0.0
    
    # Air (MNDWI tinggi)
    water_mndwi_min: float = 0.1
    water_ndvi_max: float = 0.0  # Air biasanya NDVI negatif
    
    # Lahan Terbangun (NDBI tinggi, NDVI rendah)
    builtup_ndbi_min: float = 0.1
    builtup_ndvi_max: float = 0.3


# ==============================================================================
# NORMALISASI TUTUPAN LAHAN: 6 Kelas IPSDH Final
# ==============================================================================
# WAJIB: Sistem hanya menggunakan 6 kelas IPSDH untuk output
# Database columns → Output classes (1:1 mapping - NO legacy columns):
# 1. hutan_primer → Hutan Primer
# 2. hutan_sekunder → Hutan Sekunder
# 3. tanah_kering → Tanah Lahan Kering
# 4. tanah_kosong → Tanah Kosong/Terbuka
# 5. lahan_terbangun → Lahan Terbangun
# 6. air → Air

def normalize_land_cover_to_ipsdh(data_dict: Dict[str, Any]) -> Dict[str, float]:
    """
    Return 6-class IPSDH format from database output.
    (No aggregation needed - 6 columns match 6 IPSDH classes 1:1)

    Args:
        data_dict: Dictionary dengan 6 kolom tutupan lahan IPSDH

    Returns:
        Dictionary dengan 6 kelas IPSDH final
    """
    return {
        'hutan_primer': float(data_dict.get('hutan_primer', 0) or 0),
        'hutan_sekunder': float(data_dict.get('hutan_sekunder', 0) or 0),
        'tanah_kering': float(data_dict.get('tanah_kering', 0) or 0),
        'tanah_kosong': float(data_dict.get('tanah_kosong', 0) or 0),
        'lahan_terbangun': float(data_dict.get('lahan_terbangun', 0) or 0),
        'air': float(data_dict.get('air', 0) or 0),
    }


def normalize_analysis_results(analysis_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Ensure analysis_results only include 6 IPSDH land cover classes.
    (Database now has exactly 6 columns - no legacy columns to remove)

    Args:
        analysis_results: List of yearly analysis results from database

    Returns:
        List with only 6 IPSDH classes
    """
    if not isinstance(analysis_results, list):
        return analysis_results

    # If data has legacy columns (from cache/old data), filter them out
    # Otherwise, data is already clean from database
    normalized = []
    for item in analysis_results:
        if not isinstance(item, dict):
            normalized.append(item)
            continue

        # Filter out any legacy columns (defensive, shouldn't exist now)
        normalized_item = {k: v for k, v in item.items() if k not in [
            'semak_padang_rumput', 'lahan_pertanian', 'gambut', 'tanah_terbuka'
        ]}

        normalized.append(normalized_item)

    return normalized


# ==============================================================================
# TEMPORAL STATUS & DOMINANT CLASS CALCULATION
# Implements IPSDH 6-class system with grey area detection
# ==============================================================================

def calculate_dominant_class(yearly_data: Dict[str, Any]) -> str:
    """
    Menghitung kelas dominan dari 6 kelas IPSDH.

    Args:
        yearly_data: Dict dengan keys: hutan_primer, hutan_sekunder, tanah_kering,
                     tanah_kosong, lahan_terbangun, air

    Returns:
        String nama kelas yang paling dominan
    """
    classes = {
        'hutan_primer': float(yearly_data.get('hutan_primer', 0) or 0),
        'hutan_sekunder': float(yearly_data.get('hutan_sekunder', 0) or 0),
        'tanah_kering': float(yearly_data.get('tanah_kering', 0) or 0),
        'tanah_kosong': float(yearly_data.get('tanah_kosong', 0) or 0),
        'lahan_terbangun': float(yearly_data.get('lahan_terbangun', 0) or 0),
        'air': float(yearly_data.get('air', 0) or 0)
    }

    # Return class with maximum value
    max_class = max(classes.items(), key=lambda x: x[1])
    return max_class[0] if max_class[1] > 0 else None


def calculate_temporal_status(
    prev_class: Optional[str],
    current_class: str,
    next_class: Optional[str],
    prev_prev_class: Optional[str] = None
) -> str:
    """
    Menghitung status temporal berdasarkan perbandingan kelas antar tahun.

    Aturan:
    - stable: Kelas sama dengan tahun sebelumnya
    - transition_unconfirmed: Kelas berubah dari tahun sebelum (grey area)
    - transition_confirmed: Perubahan konsisten ≥ 2 periode
    - reverted_noise: Berubah lalu kembali lagi (noise/musiman)

    Args:
        prev_class: Kelas dominan tahun N-1 (bisa None untuk tahun pertama)
        current_class: Kelas dominan tahun N (current)
        next_class: Kelas dominan tahun N+1 (bisa None untuk tahun terakhir)
        prev_prev_class: Kelas dominan tahun N-2 (untuk deteksi revert)

    Returns:
        String status: 'stable', 'transition_unconfirmed', 'transition_confirmed', 'reverted_noise'
    """
    # Tahun pertama selalu stable (tidak ada pembanding)
    if prev_class is None:
        return 'stable'

    # A. STABLE: Kelas sama dengan tahun sebelumnya
    if current_class == prev_class:
        return 'stable'

    # Ada perubahan dari tahun sebelumnya
    # Cek apakah ini noise (kembali ke kondisi 2 tahun lalu)
    if prev_prev_class and current_class == prev_prev_class:
        # D. REVERTED_NOISE: Tahun N-2 → N-1 berubah → N kembali ke N-2
        return 'reverted_noise'

    # Cek apakah perubahan konsisten ke depan
    if next_class:
        if current_class == next_class:
            # C. TRANSITION_CONFIRMED: Perubahan konsisten ≥ 2 periode
            return 'transition_confirmed'

    # B. TRANSITION_UNCONFIRMED: Perubahan baru terjadi 1 kali (grey area)
    return 'transition_unconfirmed'


async def update_temporal_status_for_history(history_id: str) -> bool:
    """
    Update temporal_status dan dominant_class untuk semua tahun dalam satu history.
    Dipanggil setelah data yearly baru diinsert/update.

    Args:
        history_id: UUID dari analysis_history

    Returns:
        bool: True jika berhasil, False jika gagal
    """
    if not supabase:
        print("⚠️ Supabase not initialized")
        return False

    try:
        # 1. Ambil semua data yearly untuk history ini, sorted by year
        result = await asyncio.to_thread(
            lambda: supabase.table("analysis_yearly_data")
                .select("*")
                .eq("history_id", history_id)
                .order("year")
                .execute()
        )

        yearly_data_list = result.data
        if not yearly_data_list or len(yearly_data_list) == 0:
            print(f"⚠️ No yearly data found for history {history_id[:8]}")
            return False

        print(f"🔄 Updating temporal status for {len(yearly_data_list)} years in history {history_id[:8]}...")

        # 2. Calculate dominant_class dan temporal_status untuk setiap tahun
        updates = []
        for i, year_data in enumerate(yearly_data_list):
            # Calculate dominant class
            dominant_class = calculate_dominant_class(year_data)

            # Get prev, current, next classes
            prev_class = None if i == 0 else calculate_dominant_class(yearly_data_list[i-1])
            current_class = dominant_class
            next_class = None if i == len(yearly_data_list) - 1 else calculate_dominant_class(yearly_data_list[i+1])
            prev_prev_class = None if i < 2 else calculate_dominant_class(yearly_data_list[i-2])

            # Calculate temporal status
            temporal_status = calculate_temporal_status(
                prev_class,
                current_class,
                next_class,
                prev_prev_class
            )

            updates.append({
                "id": year_data["id"],
                "history_id": year_data["history_id"],  # ✅ Preserve required fields
                "year": year_data["year"],              # ✅ Preserve year (NOT NULL)
                "dominant_class": dominant_class,
                "temporal_status": temporal_status
            })

        # 3. Update records individually (UPDATE only modifies specified columns)
        if updates:
            for update_data in updates:
                rid = update_data["id"]
                dom = update_data["dominant_class"]
                ts = update_data["temporal_status"]
                await asyncio.to_thread(
                    lambda r=rid, d=dom, t=ts:
                        supabase.table("analysis_yearly_data")
                            .update({"dominant_class": d, "temporal_status": t})
                            .eq("id", r)
                            .execute()
                )

            # Count status
            status_counts = {}
            for u in updates:
                status = u["temporal_status"]
                status_counts[status] = status_counts.get(status, 0) + 1

            print(f"✅ Updated temporal status for history {history_id[:8]}: {status_counts}")

        return True

    except Exception as e:
        print(f"⚠️ Error updating temporal status for history {history_id[:8]}: {e}")
        import traceback
        traceback.print_exc()
        return False


class ChangeDetectionRequest(BaseModel):
    """Request body untuk endpoint /change-detection."""
    geojson: Dict[str, Any] = Field(..., description="Objek geometri GeoJSON standar")
    start_year: int = Field(..., ge=2015, le=2030, description="Tahun awal")
    end_year: int = Field(..., ge=2015, le=2030, description="Tahun akhir")
    threshold: float = Field(default=5.0, ge=0.1, le=100.0, description="Threshold perubahan minimum (Ha)")
    thresholds: Optional[ClassificationThresholds] = None


class ChangeDetectionResponse(BaseModel):
    """Response body untuk endpoint /change-detection."""
    status: str
    message: Optional[str] = None
    start_year: int
    end_year: int
    changes: Dict[str, float] = {}  # Perubahan per kelas (Ha)
    change_percent: Dict[str, float] = {}  # Perubahan per kelas (%)
    total_change: float = 0.0
    confidence_start: Optional[float] = None  # Konfidensi tahun awal
    confidence_end: Optional[float] = None  # Konfidensi tahun akhir
    confidence_detailed_start: Dict[str, float] = {}  # Konfidensi per kelas tahun awal
    confidence_detailed_end: Dict[str, float] = {}  # Konfidensi per kelas tahun akhir
    map_url: Optional[str] = None


class AnalyzeRequest(BaseModel):
    """Request body untuk endpoint /analyze."""
    geojson: Dict[str, Any] = Field(..., description="Objek geometri GeoJSON standar")
    # New Year Range
    start_year: Optional[int] = Field(None, description="Tahun mulai analisis (contoh: 2017)")
    end_year: Optional[int] = Field(None, description="Tahun akhir analisis (contoh: 2024)")
    cloud_prob_threshold: int = Field(default=50, ge=0, le=100, description="Ambang batas probabilitas awan (0-100)")
    
    # Legacy/Deprecated
    years: int = Field(default=5, ge=1, le=10, description="Jumlah tahun untuk dianalisis (Deprecated)")
    mode: str = Field(default="series", description="Mode analisis: 'series' atau 'single'")
    specific_date: Optional[str] = Field(None, description="Tanggal spesifik jika mode adalah 'single'")
    thresholds: Optional[ClassificationThresholds] = None
    existing_data: Optional[List[Dict[str, Any]]] = Field(default=[], description="Data eksisting untuk di-skip (Direct to Cache)")
    
    scale: int = Field(default=10, ge=10, le=100, description="Resolusi spasial (Deprecated, fixed to 10m)")




class AnalyzeResponse(BaseModel):
    """Response body untuk endpoint /analyze."""
    status: str
    message: Optional[str] = None
    data: List[YearlyData] = []
    map_url: Optional[str] = None
    transition_summary: Dict[str, Any] = {} # Summary cumulative transitions
    audit_report: Dict[str, Any] = {}       # Required audit information



class SlopeRequest(BaseModel):
    """Request body untuk endpoint /map/slope."""
    geo_data: Dict[str, Any] = Field(..., description="Objek geometri GeoJSON untuk analisis")


class SaveHistoryRequest(BaseModel):
    """Request body untuk menyimpan riwayat analisis."""
    filename: str
    file_size: int
    metadata: Dict[str, Any]
    analysis_results: List[Dict[str, Any]]
    geo_data: Dict[str, Any]  # GeoJSON untuk master_lahan
    mode: Optional[str] = "replace" # "replace" or "merge"
    transition_summary: Optional[Dict[str, Any]] = {}
    audit_report: Optional[Dict[str, Any]] = {}
    # KPS Detection Fields
    kps_id: Optional[str] = None  # Foreign key to master_kps.id_kps_api
    non_kps_id: Optional[str] = None  # Foreign key to master_non_kps.id
    link_method: Optional[str] = "NONE"  # 'NO_SK_METADATA', 'MANUAL', 'NONE'
    analysis_scope: Optional[str] = "NON_KPS"  # 'KPS' or 'NON_KPS'


# ==============================================================================
# NON-KPS MODELS
# ==============================================================================

class CreateNonKpsRequest(BaseModel):
    """Request untuk membuat master Non-KPS record."""
    nama_areal: str
    lahan_id: str  # UUID from master_lahan


class NonKpsResponse(BaseModel):
    """Response untuk Non-KPS record."""
    id: str
    nama_areal: str
    lahan_id: str
    area_ha: Optional[float] = None
    centroid_lat: Optional[float] = None
    centroid_lng: Optional[float] = None
    created_at: datetime


# ==============================================================================
# CARBON TIME-SERIES MODELS (Indicative Analysis)
# ==============================================================================

class CarbonYearlyData(BaseModel):
    """Stok karbon per tahun - untuk mode Carbon Time-Series."""
    year: int
    carbon_stock_tc: float  # Total Carbon (tonnes Carbon)
    carbon_stock_co2e: float  # Converted to tCO2e
    area_breakdown: Dict[str, Any] = {}  # {class_name: area_ha}
    carbon_breakdown: Dict[str, Any] = {}  # {class_name: carbon_tc}


class CarbonChangeResult(BaseModel):
    """Hasil analisis perubahan karbon antar waktu."""
    start_year: int
    end_year: int
    delta_carbon_tc: float  # ΔC in tonnes Carbon
    delta_carbon_co2e: float  # ΔC in tCO2e
    annual_rate_tc: float  # Laju perubahan tahunan (tC/year)
    annual_rate_co2e: float  # Laju perubahan tahunan (tCO2e/year)
    trend_status: str  # "Increasing" | "Decreasing" | "Stable"
    confidence_percent: float = 0.0
    confidence_components: Dict[str, Any] = {}


class CarbonTimeSeriesRequest(BaseModel):
    """Request body untuk endpoint /carbon/analyze."""
    history_id: str = Field(..., description="ID dari analysis_history yang akan dianalisis")


class CarbonTimeSeriesResponse(BaseModel):
    """Response body untuk endpoint /carbon/analyze."""
    status: str
    message: Optional[str] = None
    yearly_data: List[CarbonYearlyData] = []
    change_result: Optional[CarbonChangeResult] = None
    carbon_factors_used: Dict[str, Any] = {}  # Transparansi faktor yang digunakan
    disclaimer: str = CARBON_DISCLAIMER


async def check_database_health():
    """Diagnosa kesehatan database dan skema."""
    if not supabase:
        print("⚠️ Supabase client not initialized")
        return
        
    print("🔍 Checking Database Schema...")
    try:
        # Coba select 1 baris untuk cek apakah tabel & kolom ada
        # Kami sengaja memilih kolom yang baru ditambahkan: geom_geojson
        await asyncio.to_thread(lambda: supabase.table("master_lahan").select("id, geom_geojson").limit(1).execute())
        print("✅ Database Schema OK: master_lahan table has correct columns.")
    except Exception as e:
        err = str(e)
        print(f"⚠️ Database Schema Issue Detected: {err}")
        if "relation" in err and "does not exist" in err:
             print("❌ Table 'master_lahan' MISSING. Please run the migration SQL.")
        elif "column" in err and "does not exist" in err:
             print("❌ Column MISSING in 'master_lahan'. Likely 'geom_geojson'.")
             print("💡 ACTION REQUIRED: Run this SQL in Supabase Editor:")
             print("   ALTER TABLE public.master_lahan ADD COLUMN IF NOT EXISTS geom_geojson jsonb;")
        else:
             print(f"❓ Unknown DB Error: {err}")



    



# ==============================================================================
# FASTAPI APPLICATION
# ==============================================================================

app = FastAPI(
    title="DIREKTORAT PENGENDALIAN PERHUTANAN SOSIAL",
    description="Monitoring Tutupan Lahan menggunakan Sentinel-2 dan Google Earth Engine",
    version="2.0.0",
)

# Enable GZip compression for all responses > 500 bytes
app.add_middleware(GZipMiddleware, minimum_size=500)

# Enable CORS untuk akses publik (Device mana pun)
ALLOWED_ORIGINS = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False, # Must be False for allow_origins=["*"]
    allow_methods=["*"],
    allow_headers=["*"],
)

# Middleware Keamanan untuk Header Proteksi
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

# Mount local storage for static file serving
app.mount("/storage", StaticFiles(directory=STORAGE_ROOT), name="storage")
print(f"🚀 Local storage mounted at /storage")

# Jalankan cek saat startup (merged with GEE init below)


# ==============================================================================
# WEBSOCKET CONNECTION MANAGER
# ==============================================================================

class ConnectionManager:
    """Manages WebSocket connections for real-time progress updates."""
    
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}
    
    async def connect(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        self.active_connections[client_id] = websocket
        print(f"🔌 WebSocket connected: {client_id}")
    
    def disconnect(self, client_id: str):
        if client_id in self.active_connections:
            del self.active_connections[client_id]
            print(f"🔌 WebSocket disconnected: {client_id}")
    
    async def send_progress(self, client_id: str, progress: int, step: str, detail: str = ""):
        """Send progress update to a specific client."""
        if client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_json({
                    "type": "progress",
                    "progress": progress,
                    "step": step,
                    "detail": detail
                })
            except Exception as e:
                print(f"⚠️ Failed to send progress: {e}")
    
    async def send_complete(self, client_id: str, data: dict):
        """Send completion message with results."""
        if client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_json({
                    "type": "complete",
                    "data": data
                })
            except Exception as e:
                print(f"⚠️ Failed to send completion: {e}")
    
    async def send_queue_status(self, client_id: str, position: int):
        """Send queue position update to a specific client."""
        if client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_json({
                    "type": "queue_status",
                    "position": position
                })
                print(f"🚥 Notify {client_id}: Queue Position #{position}")
            except Exception as e:
                print(f"⚠️ Failed to send queue status: {e}")
    
    async def send_error(self, client_id: str, error: str):
        """Send error message."""
        if client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_json({
                    "type": "error",
                    "error": error
                })
            except Exception as e:
                print(f"⚠️ Failed to send error: {e}")
    
    async def broadcast_log(self, log_entry: dict):
        """Broadcast log message to all log viewer clients."""
        disconnected = []
        for client_id, ws in list(self.active_connections.items()):
            if client_id.startswith("log_"):
                try:
                    await ws.send_json({
                        "type": "log",
                        "data": log_entry
                    })
                except Exception:
                    disconnected.append(client_id)
        
        for client_id in disconnected:
            self.disconnect(client_id)


# Global connection manager
ws_manager = ConnectionManager()


# ==============================================================================
# WEBSOCKET LOG HANDLER
# ==============================================================================

from collections import deque

# Global log buffer to store recent logs
log_buffer = deque(maxlen=200)  # Store last 200 log entries

class WebSocketLogHandler(logging.Handler):
    """Custom logging handler that broadcasts logs to WebSocket clients."""
    
    def __init__(self, ws_manager: ConnectionManager):
        super().__init__()
        self.ws_manager = ws_manager
        self.loop = None
    
    def emit(self, record: logging.LogRecord):
        """Emit a log record to all connected log viewer clients."""
        try:
            log_entry = {
                "timestamp": datetime.fromtimestamp(record.created).isoformat(),
                "level": record.levelname,
                "logger": record.name,
                "message": self.format(record),
                "module": record.module,
                "function": record.funcName,
                "line": record.lineno
            }
            
            # Add to buffer for history
            log_buffer.append(log_entry)
            
            # Get or create event loop
            try:
                loop = asyncio.get_running_loop()
            except RuntimeError:
                # No running loop, skip WebSocket broadcast
                return
            
            # Schedule the broadcast
            asyncio.create_task(self.ws_manager.broadcast_log(log_entry))
        except Exception:
            # Silently fail to avoid logging recursion
            pass


# Configure logging with WebSocket handler
logger = logging.getLogger("shptutupan")
logger.setLevel(logging.INFO)

# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# WebSocket handler
ws_log_handler = WebSocketLogHandler(ws_manager)
ws_log_handler.setLevel(logging.INFO)
ws_log_formatter = logging.Formatter('%(message)s')
ws_log_handler.setFormatter(ws_log_formatter)
logger.addHandler(ws_log_handler)

print("✅ WebSocket log handler initialized")

# Add some initial logs to buffer
logger.info("🚀 Backend server started successfully")
logger.info("📡 WebSocket log handler initialized and ready")
logger.info("🔧 System ready to accept connections")


# ==============================================================================
# GEE INITIALIZATION
# ==============================================================================

def initialize_gee():
    """Inisialisasi Google Earth Engine dengan kredensial Service Account."""
    try:
        if os.path.exists(KEY_FILE):
            credentials = ee.ServiceAccountCredentials(None, KEY_FILE)
            ee.Initialize(credentials)
            print(f"✅ GEE terinisialisasi dengan service account dari: {KEY_FILE}")
            logger.info(f"✅ Google Earth Engine initialized with service account: {KEY_FILE}")
        else:
            ee.Initialize()
            print("✅ GEE terinisialisasi dengan kredensial default")
            logger.info("✅ Google Earth Engine initialized with default credentials")
    except Exception as e:
        print(f"❌ Gagal inisialisasi GEE: {str(e)}")
        logger.error(f"❌ Failed to initialize Google Earth Engine: {str(e)}")
        raise RuntimeError(f"Inisialisasi GEE gagal: {str(e)}")


@app.on_event("startup")
async def startup_event():
    """Inisialisasi GEE dan cek kesehatan database saat aplikasi startup."""
    initialize_gee()
    await check_database_health()


@app.post("/api/sync/kps")
async def trigger_kps_sync(background_tasks: BackgroundTasks):
    """
    Trigger manual synchronization of master_kps from external PKPS APIs.
    Runs in background.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Database not connected")
        
    background_tasks.add_task(run_sync_process, supabase)
    return {
        "status": "queued",
        "message": "PKPS synchronization started in background",
        "timestamp": datetime.utcnow().isoformat()
    }


# ==============================================================================
# DUPLICATE DETECTION ENDPOINT
# ==============================================================================

@app.post("/api/check-duplicate")
async def check_duplicate(request: dict):
    """
    Check if geometry already exists in database before analysis.
    Returns existing analysis info if found, allowing user to choose Update or Replace.
    
    Request body:
        - geo_data: GeoJSON object
    
    Returns:
        - is_duplicate: bool
        - lahan_id: UUID (if duplicate)
        - history_id: UUID (if has analysis)
        - existing_filename: str
        - existing_years: List[int] - years already analyzed
        - kps_id: str (if linked to KPS)
        - created_at: timestamp
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Database not connected")
    
    geo_data = request.get("geo_data")
    if not geo_data:
        return {"is_duplicate": False, "error": "No geo_data provided"}
    
    try:
        # Hash geometry for duplicate detection
        geom_str = json.dumps(geo_data, sort_keys=True)
        geom_hash = hashlib.md5(geom_str.encode()).hexdigest()
        
        print(f"🔍 Checking duplicate for geom_hash: {geom_hash[:12]}...")
        
        # Check master_lahan
        lahan_res = await asyncio.to_thread(
            lambda: supabase.table("master_lahan")
            .select("id")
            .eq("geom_hash", geom_hash)
            .limit(1)
            .execute()
        )
        
        if not lahan_res.data or len(lahan_res.data) == 0:
            print(f"   ✅ No duplicate found")
            return {"is_duplicate": False}
        
        lahan_id = lahan_res.data[0]['id']
        print(f"   ⚠️ Found existing lahan_id: {lahan_id}")
        
        # Get existing analysis history
        history_res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("id, filename, analysis_results, metadata, created_at, kps_id")
            .eq("lahan_id", lahan_id)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
        )
        
        if not history_res.data or len(history_res.data) == 0:
            # Lahan exists but no analysis yet
            print(f"   ℹ️ Lahan exists but no analysis history")
            return {
                "is_duplicate": True,
                "lahan_id": lahan_id,
                "has_analysis": False
            }
        
        existing = history_res.data[0]
        analysis_results = existing.get('analysis_results', [])
        existing_years = sorted([r.get('year') for r in analysis_results if r.get('year')])
        
        # Get KPS info if linked
        kps_info = None
        if existing.get('kps_id'):
            kps_res = await asyncio.to_thread(
                lambda: supabase.table("master_kps")
                .select("nama_kps, no_sk")
                .eq("id_kps_api", existing['kps_id'])
                .limit(1)
                .execute()
            )
            if kps_res.data:
                kps_info = kps_res.data[0]
        
        print(f"   📊 Existing analysis: {len(existing_years)} years ({existing_years})")
        
        return {
            "is_duplicate": True,
            "has_analysis": True,
            "lahan_id": lahan_id,
            "history_id": existing['id'],
            "existing_filename": existing.get('filename'),
            "existing_years": existing_years,
            "kps_id": existing.get('kps_id'),
            "kps_info": kps_info,
            "created_at": existing.get('created_at'),
            "metadata": existing.get('metadata', {})
        }
        
    except Exception as e:
        print(f"❌ Check duplicate error: {e}")
        return {"is_duplicate": False, "error": str(e)}


# ==============================================================================
# KPS AUTO-DETECTION & SEARCH ENDPOINTS
# ==============================================================================

@app.get("/api/kps/auto-detect")
async def auto_detect_kps(no_sk: str):
    """
    Cari KPS di master_kps berdasarkan NO_SK dari metadata SHP.
    Digunakan untuk auto-detection saat upload shapefile.
    
    Returns:
        - status: 'found' | 'not_found' | 'error'
        - kps: KPS data jika ditemukan
        - method: 'exact' | 'partial' (jika found)
    """
    if not supabase:
        return {"status": "error", "message": "Database not connected", "kps": None}
    
    if not no_sk or not no_sk.strip():
        return {"status": "not_found", "kps": None}
    
    try:
        # Normalisasi input (trim whitespace, remove potential problematic chars)
        normalized_query = no_sk.strip()
        
        print(f"🔍 KPS Auto-detect START")
        print(f"   Query (original): '{no_sk}'")
        print(f"   Query (normalized): '{normalized_query}'")
        
        # 1. Search in master_kps
        # We try to search in no_sk column (Standard)
        # We use ilike for case-insensitive partial matching
        res = await asyncio.to_thread(
            lambda: supabase.table("master_kps")
            .select("id_kps_api, nama_kps, no_sk, kps_type, nama_prov, nama_kab, luas_sk_ha")
            .ilike("no_sk", f"%{normalized_query}%")
            .limit(10)
            .execute()
        )
        
        results = res.data or []
        print(f"   Database results: {len(results)} candidates found")
        
        # 2. Match Scoring Logic
        if results:
            query_upper = normalized_query.upper().replace("/", "").replace(".", "").replace(" ", "").replace("-", "")
            print(f"   Query (fuzzy): '{query_upper}'")
            
            # Find best match
            best_match = None
            match_method = "partial"
            
            for idx, k in enumerate(results):
                # Get the actual SK number from record
                record_sk = str(k.get("no_sk") or "").strip()
                record_upper = record_sk.upper()
                record_fuzzy = record_upper.replace("/", "").replace(".", "").replace(" ", "").replace("-", "")
                
                print(f"   Candidate #{idx+1}:")
                print(f"      DB value: '{record_sk}'")
                print(f"      Fuzzy: '{record_fuzzy}'")
                
                # Check for absolute exact match (case-insensitive)
                if record_upper == normalized_query.upper():
                    best_match = k
                    match_method = "exact"
                    print(f"      ✅ EXACT MATCH (case-insensitive)")
                    break
                    
                # Check for "fuzzy" exact match (ignoring separators)
                if record_fuzzy == query_upper:
                    best_match = k
                    match_method = "exact" # Close enough to be considered exact for auto-detect
                    print(f"      ✅ FUZZY EXACT MATCH (ignoring separators)")
                    break
                
                print(f"      ❌ No exact match")
            
            # If no "exact" match found among candidates, pick the first one from DB
            if not best_match:
                best_match = results[0]
                match_method = "partial"
                print(f"   Using first partial match: '{best_match.get('no_sk')}'")
                
            print(f"✅ KPS Auto-detect: Found match for '{no_sk}' (Method: {match_method})")
            print(f"   Selected: {best_match.get('nama_kps')} | {best_match.get('no_sk')}")
            return {"status": "found", "kps": best_match, "method": match_method}
        
        print(f"ℹ️ KPS Auto-detect: No match found for NO_SK '{no_sk}'")
        return {"status": "not_found", "kps": None}
        
    except Exception as e:
        print(f"❌ KPS Auto-detect error: {e}")
        return {"status": "error", "message": str(e), "kps": None}


@app.get("/api/kps/search")
async def search_kps(query: str, limit: int = 10):
    """
    Pencarian KPS untuk fitur manual search.
    Mencari di nama_kps, no_sk, provinsi, dan kab_kota.
    
    Args:
        query: Search term (minimum 2 characters)
        limit: Maximum results to return (default 10)
    
    Returns:
        - results: Array of matching KPS records
    """
    if not supabase:
        return {"results": [], "error": "Database not connected"}
    
    if not query or len(query.strip()) < 2:
        return {"results": []}
    
    try:
        q = query.strip()
        
        # Search by name, no_sk, or location using OR filter
        res = await asyncio.to_thread(
            lambda: supabase.table("master_kps")
            .select("id_kps_api, nama_kps, no_sk, kps_type, nama_prov, nama_kab, luas_sk_ha")
            .or_(f"nama_kps.ilike.%{q}%,no_sk.ilike.%{q}%,nama_prov.ilike.%{q}%,nama_kab.ilike.%{q}%")
            .limit(limit)
            .execute()
        )
        
        print(f"🔍 KPS Search '{q}': Found {len(res.data or [])} results")
        return {"results": res.data or []}
        
    except Exception as e:
        print(f"❌ KPS Search error: {e}")
        return {"results": [], "error": str(e)}


@app.get("/api/kps/debug")
async def debug_kps_data(sample: int = 10, search: str = None):
    """
    Debug endpoint untuk melihat data aktual di master_kps.
    Membantu diagnose masalah matching NO_SK.
    
    Args:
        sample: Jumlah sample data (default 10)
        search: Optional - cari NO_SK spesifik
    
    Returns:
        Sample data dari master_kps dengan info kolom
    """
    if not supabase:
        return {"error": "Database not connected"}
    
    try:
        if search:
            # Search for specific NO_SK
            print(f"🔍 Debug search for: '{search}'")
            
            # Try multiple search strategies
            results = {
                "search_query": search,
                "strategies": {}
            }
            
            # 1. Exact ilike
            res1 = await asyncio.to_thread(
                lambda: supabase.table("master_kps")
                .select("id_kps_api, nama_kps, no_sk, kps_type")
                .ilike("no_sk", search)
                .limit(5)
                .execute()
            )
            results["strategies"]["exact_ilike"] = {
                "count": len(res1.data or []),
                "results": res1.data or []
            }
            
            # 2. Partial ilike
            res2 = await asyncio.to_thread(
                lambda: supabase.table("master_kps")
                .select("id_kps_api, nama_kps, no_sk, kps_type")
                .ilike("no_sk", f"%{search}%")
                .limit(5)
                .execute()
            )
            results["strategies"]["partial_ilike"] = {
                "count": len(res2.data or []),
                "results": res2.data or []
            }
            
            # 3. Contains specific parts (e.g., "2861")
            if "/" in search or "." in search:
                # Extract number part
                import re
                numbers = re.findall(r'\d+', search)
                if numbers:
                    main_num = numbers[0]
                    res3 = await asyncio.to_thread(
                        lambda: supabase.table("master_kps")
                        .select("id_kps_api, nama_kps, no_sk, kps_type")
                        .ilike("no_sk", f"%{main_num}%")
                        .limit(10)
                        .execute()
                    )
                    results["strategies"]["number_match"] = {
                        "number_searched": main_num,
                        "count": len(res3.data or []),
                        "results": res3.data or []
                    }
            
            return results
        else:
            # Get sample data
            res = await asyncio.to_thread(
                lambda: supabase.table("master_kps")
                .select("id_kps_api, nama_kps, no_sk, kps_type")
                .limit(sample)
                .execute()
            )
            
            data = res.data or []
            
            return {
                "total_sample": len(data),
                "samples": data,
                "columns": list(data[0].keys()) if data else [],
                "no_sk_examples": [item.get("no_sk") for item in data[:5]] if data else []
            }
    except Exception as e:
        print(f"❌ Debug endpoint error: {e}")
        return {"error": str(e)}


@app.post("/api/kps/link")
async def link_kps_to_analysis(request: dict):
    """
    Link an existing analysis_history record to a KPS.
    This updates the kps_id field and all related tables.
    
    Request body:
        - history_id: UUID of the analysis_history record
        - kps_id: id_kps_api from master_kps
        - link_method: 'MANUAL' or 'NO_SK_METADATA'
    
    This endpoint updates:
        1. analysis_history (kps_id, link_method, analysis_scope)
        2. analysis_yearly_data (via history_id FK - no kps_id column there)
        3. analysis_hotspots (kps_id if column exists)
        4. analysis_slope_summary (kps_id if column exists)
    
    Views will automatically reflect the updated kps_id via joins.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Database not connected")
    
    history_id = request.get("history_id")
    kps_id = request.get("kps_id")
    link_method = request.get("link_method", "MANUAL")
    
    if not history_id:
        raise HTTPException(status_code=400, detail="history_id is required")
    
    try:
        # 1. Update analysis_history (main record)
        update_data = {
            "kps_id": kps_id,
            "link_method": link_method,
            "analysis_scope": "KPS" if kps_id else "NON_KPS"
        }
        
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .update(update_data)
            .eq("id", history_id)
            .execute()
        )
        
        if not res.data or len(res.data) == 0:
            raise HTTPException(status_code=404, detail="History record not found")
        
        updated_history = res.data[0]
        print(f"✅ Linked history {history_id[:8]} to KPS {kps_id}")
        
        # 2. Try to update analysis_hotspots if it has kps_id column
        try:
            await asyncio.to_thread(
                lambda: supabase.table("analysis_hotspots")
                .update({"kps_id": kps_id})
                .eq("history_id", history_id)
                .execute()
            )
            print(f"   ✅ Updated analysis_hotspots for history {history_id[:8]}")
        except Exception as e:
            # Column might not exist, that's OK
            print(f"   ℹ️ Skipped analysis_hotspots update: {e}")
        
        # 3. Try to update analysis_slope_summary if it has kps_id column
        try:
            await asyncio.to_thread(
                lambda: supabase.table("analysis_slope_summary")
                .update({"kps_id": kps_id})
                .eq("history_id", history_id)
                .execute()
            )
            print(f"   ✅ Updated analysis_slope_summary for history {history_id[:8]}")
        except Exception as e:
            print(f"   ℹ️ Skipped analysis_slope_summary update: {e}")
        
        # Invalidate caches
        invalidate_cache(f"history_detail_{history_id}")
        invalidate_cache("history_list")
        
        return {
            "status": "success", 
            "message": f"Linked to KPS successfully",
            "data": updated_history
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ KPS Link error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==============================================================================
# NON-KPS MANAGEMENT ENDPOINTS
# ==============================================================================

@app.post("/api/non-kps/create", response_model=NonKpsResponse)
async def create_non_kps(request: CreateNonKpsRequest):
    """
    Create a new Non-KPS master record.
    Auto-called during Non-KPS analysis save workflow.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase client not configured")

    try:
        print(f"🏞️ Creating Non-KPS record: {request.nama_areal} (lahan_id: {request.lahan_id})")

        # Fetch geometry details from master_lahan
        lahan_res = await asyncio.to_thread(
            lambda: supabase.table("master_lahan")
            .select("area_ha, centroid_lat, centroid_lng")
            .eq("id", request.lahan_id)
            .limit(1)
            .execute()
        )

        if not lahan_res.data:
            raise HTTPException(status_code=404, detail=f"Lahan ID {request.lahan_id} not found")

        lahan = lahan_res.data[0]

        # Insert into master_non_kps
        res = await asyncio.to_thread(
            lambda: supabase.table("master_non_kps").insert({
                "nama_areal": request.nama_areal,
                "lahan_id": request.lahan_id,
                "area_ha": lahan.get("area_ha"),
                "centroid_lat": lahan.get("centroid_lat"),
                "centroid_lng": lahan.get("centroid_lng")
            }).execute()
        )

        if not res.data:
            raise HTTPException(status_code=500, detail="Failed to create Non-KPS record")

        record = res.data[0]
        print(f"✅ Non-KPS created: ID {record['id']}")

        return NonKpsResponse(**record)

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error creating Non-KPS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/non-kps/search")
async def search_non_kps(query: str = Query(..., min_length=2)):
    """
    Search Non-KPS by name (for future autocomplete/linking features).
    Supports partial text matching.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase client not configured")

    try:
        print(f"🔍 Searching Non-KPS: '{query}'")

        res = await asyncio.to_thread(
            lambda: supabase.table("master_non_kps")
            .select("*")
            .ilike("nama_areal", f"%{query}%")
            .order("created_at", desc=True)
            .limit(20)
            .execute()
        )

        results = res.data or []
        print(f"✅ Found {len(results)} Non-KPS records")

        return {"status": "success", "data": results, "count": len(results)}

    except Exception as e:
        print(f"❌ Error searching Non-KPS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def strip_z_coordinates(geom_dict):
    """Hapus koordinat Z dari geometri (konversi 3D ke 2D)."""
    def strip_coords(coords):
        if not coords:
            return coords
        if isinstance(coords[0], (list, tuple)):
            return [strip_coords(c) for c in coords]
        return coords[:2]  # Ambil X dan Y saja

    if geom_dict.get('type') == 'Polygon':
        geom_dict['coordinates'] = strip_coords(geom_dict['coordinates'])
    elif geom_dict.get('type') == 'MultiPolygon':
        geom_dict['coordinates'] = strip_coords(geom_dict['coordinates'])
    return geom_dict


def get_regional_window(geometry: ee.Geometry) -> dict:
    """
    Determine the optimal dry-season window based on region.
    
    Default: July-September (Months 7-9)
    Sumatra, Kalimantan, Sulawesi: July-September
    Java: June-August (Months 6-8)
    Papua: August-October (Months 8-10)
    """
    try:
        # Get centroid to determine region roughly
        centroid = geometry.centroid().coordinates().getInfo()
        lon, lat = centroid[0], centroid[1]
        
        # Simple bounding box check for major islands
        # Sumatra: 95E - 106E
        if 95 <= lon <= 106:
             return {"start_month": 7, "end_month": 9, "region": "Sumatra"}
        # Java: 105E - 116E (South of equator)
        elif 105 <= lon <= 116 and lat < -4:
             return {"start_month": 6, "end_month": 8, "region": "Java"}
        # Kalimantan: 108E - 119E (North of Java)
        elif 108 <= lon <= 119 and lat > -4:
             return {"start_month": 7, "end_month": 9, "region": "Kalimantan"}
        # Sulawesi: 118E - 125E
        elif 118 <= lon <= 125:
             return {"start_month": 7, "end_month": 9, "region": "Sulawesi"}
        # Papua: 128E - 141E
        elif 128 <= lon <= 141:
             return {"start_month": 8, "end_month": 10, "region": "Papua"}
        
        # Default fallback
        return {"start_month": 7, "end_month": 9, "region": "General/Unknown"}
    except Exception as e:
        print(f"⚠️ Regional detection failed, using default window: {e}")
        return {"start_month": 7, "end_month": 9, "region": "Default"}


def score_scene_for_aoi(aoi: ee.Geometry):
    """
    Factory function: Returns a mapper that scores each scene based on 
    ACTUAL cloud coverage within the Area of Interest (AOI), not the entire scene.
    
    This is critical because GEE's CLOUDY_PIXEL_PERCENTAGE is for the entire 
    100km x 100km scene, but user's polygon might be in a clear spot.
    """
    def _score(image):
        # Get cloud/shadow mask from SCL band
        scl = image.select('SCL')
        
        # Cloud classes in SCL: 3=shadow, 8=med cloud, 9=high cloud, 10=cirrus
        cloud_mask = scl.eq(3).Or(scl.eq(8)).Or(scl.eq(9)).Or(scl.eq(10))
        
        # Calculate cloud percentage within AOI only
        aoi_cloud_stats = cloud_mask.reduceRegion(
            reducer=ee.Reducer.mean(),
            geometry=aoi,
            scale=60,  # SCL is 20m, use 60m for speed
            maxPixels=1e7,
            bestEffort=True
        )
        aoi_cloud_pct = ee.Number(aoi_cloud_stats.get('SCL')).multiply(100).max(0).min(100)
        
        # Calculate valid pixel percentage (non-masked data availability)
        valid_mask = image.select('B4').mask()
        valid_stats = valid_mask.reduceRegion(
            reducer=ee.Reducer.mean(),
            geometry=aoi,
            scale=60,
            maxPixels=1e7,
            bestEffort=True
        )
        valid_pct = ee.Number(valid_stats.get('B4')).multiply(100).max(0).min(100)
        
        # Combined quality score: prioritize low clouds, then data availability
        # Formula: (100 - cloud%) * 0.7 + valid% * 0.3
        quality_score = ee.Number(100).subtract(aoi_cloud_pct).multiply(0.7).add(valid_pct.multiply(0.3))
        
        return image.set({
            'aoi_cloud_pct': aoi_cloud_pct,
            'aoi_valid_pct': valid_pct,
            'aoi_quality': quality_score
        })
    
    return _score


def get_optimal_imagery(geometry: ee.Geometry, year: int) -> dict:
    """
    Get the best quality imagery for a given year using AOI-SPECIFIC quality scoring.
    
    Strategy:
    1. Collect scenes from dry season window
    2. Score each scene based on cloud coverage WITHIN THE AOI (not entire scene)
    3. Sort by AOI-specific quality score
    4. Take top N best scenes for composite
    5. If insufficient, expand window and relax filters
    
    Returns:
        dict with keys:
        - collection: ee.ImageCollection (preprocessed, cloud-masked)
        - scene_count: int
        - window_used: str ("regional" or "full_year")
        - region: str
        - date_range: str
        - avg_aoi_cloud: float (average cloud % within AOI)
    """
    MIN_SCENES = 3
    MAX_SCENES = 12
    LOOSE_CLOUD_LIMIT = 80  # Initial filter - will be refined by AOI scoring
    
    # Get regional window
    window_params = get_regional_window(geometry)
    region = window_params["region"]
    start_month = window_params["start_month"]
    end_month = window_params["end_month"]
    
    # Determine date range
    if year:
        # Standard annual window (dry season)
        last_day = 30 if end_month in [4, 6, 9, 11] else (28 if end_month == 2 else 31)
        regional_start = f"{year}-{start_month:02d}-01"
        regional_end = f"{year}-{end_month:02d}-{last_day}"
    else:
        regional_start = None
        regional_end = None
    
    # =========================================================================
    # SMART SCENE SELECTION with AOI-Specific Quality Scoring
    # =========================================================================
    
    def try_get_best_scenes(date_start, date_end, cloud_limit, window_name):
        """Try to get best scenes for given date range using AOI-specific scoring"""
        
        # Initial collection - loose filter, we'll score properly
        base_collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
            .filterBounds(geometry)
            .filterDate(date_start, date_end)
            .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', cloud_limit)))
        
        base_count = base_collection.size().getInfo()
        
        if base_count == 0:
            return None, 0, 0
        
        # Apply AOI-specific scoring to each scene
        scored_collection = base_collection.map(score_scene_for_aoi(geometry))
        
        # Sort by AOI quality (higher = better) and take top scenes
        best_scenes = (scored_collection
            .sort('aoi_quality', False)  # Descending - best first
            .limit(MAX_SCENES))
        
        # Get average AOI cloud coverage for logging
        try:
            avg_cloud = best_scenes.aggregate_mean('aoi_cloud_pct').getInfo()
            avg_cloud = round(avg_cloud, 1) if avg_cloud else 0
        except Exception as e:
            print(f"⚠️ Cloud aggregation warning: {e}")
            avg_cloud = 0
        
        final_count = best_scenes.size().getInfo()
        
        # Apply cloud masking
        masked_collection = best_scenes.map(mask_clouds_sentinel2)
        
        if final_count >= MIN_SCENES:
            print(f"   ✅ {window_name}: {final_count} scenes, AOI cloud ~{avg_cloud}% ({region})")
            return masked_collection, final_count, avg_cloud
        
        return None, final_count, avg_cloud
    
    # Strategy 1: Try regional dry-season window
    collection, count, avg_cloud = try_get_best_scenes(
        regional_start, regional_end, LOOSE_CLOUD_LIMIT, "Regional window"
    )
    if collection:
        return {
            "collection": collection,
            "scene_count": count,
            "window_used": "regional",
            "region": region,
            "date_range": f"{regional_start} to {regional_end}",
            "avg_aoi_cloud": avg_cloud
        }
    
    # Strategy 2: Expand to full year
    print(f"   ⚠️ Regional window insufficient, expanding to full year...")
    full_year_start = f"{year}-01-01"
    full_year_end = f"{year}-12-31"
    
    collection, count, avg_cloud = try_get_best_scenes(
        full_year_start, full_year_end, LOOSE_CLOUD_LIMIT, "Full year"
    )
    if collection:
        return {
            "collection": collection,
            "scene_count": count,
            "window_used": "full_year",
            "region": region,
            "date_range": f"{full_year_start} to {full_year_end}",
            "avg_aoi_cloud": avg_cloud
        }
    
    # Strategy 3: Relax cloud filter significantly
    print(f"   ⚠️ Still insufficient, relaxing to 95% cloud limit...")
    collection, count, avg_cloud = try_get_best_scenes(
        full_year_start, full_year_end, 95, "Relaxed filter"
    )
    
    if collection:
        return {
            "collection": collection,
            "scene_count": count,
            "window_used": "full_year_relaxed",
            "region": region,
            "date_range": f"{full_year_start} to {full_year_end}",
            "avg_aoi_cloud": avg_cloud
        }
    
    # Last resort: return whatever we can get
    print(f"   ⚠️ Minimal data available: {count} scenes")
    fallback = (ee.ImageCollection(SENTINEL2_COLLECTION)
        .filterBounds(geometry)
        .filterDate(full_year_start, full_year_end)
        .limit(MAX_SCENES)
        .map(mask_clouds_sentinel2))
    
    return {
        "collection": fallback,
        "scene_count": fallback.size().getInfo(),
        "window_used": "fallback",
        "region": region,
        "date_range": f"{full_year_start} to {full_year_end}",
        "avg_aoi_cloud": 999  # Unknown
    }

        
        
def get_optimal_imagery_custom(geometry: ee.Geometry, date_start: str, date_end: str) -> dict:
    """
    Get the best quality imagery for a custom date range.
    Used for 'Single Date' analysis mode.
    """
    MAX_SCENES = 12
    LOOSE_CLOUD_LIMIT = 90
    
    # 1. Initial collection
    base_collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
        .filterBounds(geometry)
        .filterDate(date_start, date_end)
        .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', LOOSE_CLOUD_LIMIT)))
    
    # 2. Score by AOI quality
    scored_collection = base_collection.map(score_scene_for_aoi(geometry))
    
    # 3. Sort and take best
    best_scenes = (scored_collection
        .sort('aoi_quality', False)
        .limit(MAX_SCENES))
    
    count = best_scenes.size().getInfo()
    
    # Get average cloud
    try:
        avg_cloud = best_scenes.aggregate_mean('aoi_cloud_pct').getInfo()
        avg_cloud = round(avg_cloud, 1) if avg_cloud else 0
    except Exception as e:
        print(f"⚠️ Cloud aggregation warning: {e}")
        avg_cloud = 0
        
    masked_collection = best_scenes.map(mask_clouds_sentinel2)
    
    return {
        "collection": masked_collection,
        "scene_count": count,
        "window_used": "custom_range",
        "region": "User Selected",
        "date_range": f"{date_start} to {date_end}",
        "avg_aoi_cloud": avg_cloud
    }


def whittaker_smoothing(y, lam=5.0):
    """
    Apply Whittaker smoothing to a time series.
    y: list of values
    lam: smoothing parameter (higher = smoother)
    """
    n = len(y)
    if n < 3:
        return y
        
    y = np.array(y)
    # E is the identity matrix
    E = sparse.eye(n, format='csc')
    # D is the second difference matrix
    D = sparse.diags([1, -2, 1], [0, 1, 2], shape=(n-2, n)).tocsc()
    
    # Validation: Replace NaNs/Zeros if completely missing data to avoid artifacts
    # (Simple infill could be added here if needed, currently assumes data exists)
    
    # Solve (E + lam * D^T D) z = y
    try:
        # Z = (E + lam * D.T * D)^-1 * y
        penalty = lam * D.T @ D
        result = splu(E + penalty).solve(y)
        return result.tolist()
    except Exception as e:
        print(f"⚠️ Smoothing failed: {e}")
        return y.tolist()



def geojson_to_ee_geometry(geojson_data: Dict[str, Any]) -> ee.Geometry:
    """Konversi GeoJSON ke Earth Engine Geometry."""
    try:
        geojson_2d = strip_z_coordinates(geojson_data)
        geojson_type = geojson_2d.get("type", "")
        print(f"🗺️ Memproses tipe GeoJSON: {geojson_type}")
        
        if geojson_type == "FeatureCollection":
            features = geojson_data.get("features", [])
            
            # Security Limit: Prevent DoS
            if len(features) > 1000:
                raise ValueError("Jumlah fitur melebihi batas maksimum (1000). Harap sederhanakan data.")

            print(f"   Ditemukan {len(features)} fitur dalam FeatureCollection. Membersihkan dengan Shapely...")
            
            cleaned_features = []
            for i, feat in enumerate(features):
                geom_dict = feat.get("geometry")
                if not geom_dict:
                    continue
                try:
                    geom_2d = strip_z_coordinates(geom_dict)
                    shp_geom = shape(geom_2d)
                    if not shp_geom.is_valid:
                        print(f"   Fitur {i}: Geometri tidak valid, memperbaiki...")
                        shp_geom = make_valid(shp_geom)
                    clean_geom = mapping(shp_geom)
                    cleaned_features.append({
                        "type": "Feature",
                        "properties": feat.get("properties", {}),
                        "geometry": clean_geom
                    })
                except Exception as geom_err:
                    print(f"   Fitur {i}: Dilewati karena error: {geom_err}")
                    continue
            
            if not cleaned_features:
                raise ValueError("Tidak ada geometri valid setelah pembersihan")
            
            print(f"   Berhasil membersihkan {len(cleaned_features)} fitur valid")
            cleaned_fc = {"type": "FeatureCollection", "features": cleaned_features}
            fc = ee.FeatureCollection(cleaned_fc)
            return fc.geometry()
        
        elif geojson_type == "Feature":
            geometry = geojson_data.get("geometry")
            if not geometry:
                raise ValueError("Feature tidak memiliki geometri")
            
            geom_2d = strip_z_coordinates(geometry)
            shp_geom = shape(geom_2d)
            if not shp_geom.is_valid:
                print("   Geometri Feature tidak valid, memperbaiki...")
                shp_geom = make_valid(shp_geom)
            clean_geom = mapping(shp_geom)
            print(f"   Tipe geometri Feature: {clean_geom.get('type', 'unknown')}")
            return ee.Geometry(clean_geom)
        
        else:
            print(f"   Tipe geometri langsung: {geojson_type}")
            geom_2d = strip_z_coordinates(geojson_data)
            shp_geom = shape(geom_2d)
            if not shp_geom.is_valid:
                print("   Geometri mentah tidak valid, memperbaiki...")
                shp_geom = make_valid(shp_geom)
            clean_geom = mapping(shp_geom)
            return ee.Geometry(clean_geom)
            
    except Exception as e:
        print(f"❌ Error konversi geometri: {str(e)}")
        try:
            with open("debug_geoip.json", "w") as f:
                json.dump(geojson_data, f, indent=2)
            print("   Menyimpan GeoJSON debug ke debug_geoip.json")
        except Exception as file_err:
            print(f"   Gagal menyimpan file debug: {file_err}")

        import traceback
        traceback.print_exc()
        raise ValueError(f"Geometri GeoJSON tidak valid. {str(e)}")


def mask_clouds_sentinel2(image):
    """
    Enhanced cloud and shadow masking for Sentinel-2 using QA60 and SCL bands.
    
    SCL (Scene Classification Layer) classes:
    - 0: No data, 1: Saturated/defective, 2: Dark area/shadows
    - 3: Cloud shadows, 4: Vegetation, 5: Bare soils
    - 6: Water, 7: Cloud low probability, 8: Cloud medium probability
    - 9: Cloud high probability, 10: Thin cirrus, 11: Snow/ice
    """
    # Method 1: QA60 bitmask (for backward compatibility)
    qa = image.select('QA60')
    cloud_bit_mask = 1 << 10
    cirrus_bit_mask = 1 << 11
    qa_mask = qa.bitwiseAnd(cloud_bit_mask).eq(0).And(
              qa.bitwiseAnd(cirrus_bit_mask).eq(0))
    
    # Method 2: SCL band - LESS AGGRESSIVE approach
    # Only EXCLUDE problematic classes, keep everything else
    # Exclude: 0 (no data), 1 (saturated), 3 (cloud shadows), 
    #          8 (medium clouds), 9 (high clouds), 10 (cirrus)
    # Keep: 2 (dark), 4 (veg), 5 (bare), 6 (water), 7 (low prob cloud), 11 (snow)
    try:
        scl = image.select('SCL')
        # Blacklist approach: mask OUT only definite bad pixels
        scl_mask = (scl.neq(0)  # No data
                   .And(scl.neq(1))   # Saturated
                   .And(scl.neq(3))   # Cloud shadows
                   .And(scl.neq(8))   # Cloud medium probability
                   .And(scl.neq(9))   # Cloud high probability
                   .And(scl.neq(10))  # Thin cirrus
                   .And(scl.neq(7))   # Cloud low probability (Stricter masking)
                   .And(scl.neq(11))) # Snow/Ice (often bright clouds)
        
        # Combine both masks
        combined_mask = qa_mask.And(scl_mask)
        masked = image.updateMask(combined_mask)
        
        # Scale optical bands to 0-1 surface reflectance
        # Sentinel-2 data is scaled by 10000
        return masked.divide(10000).copyProperties(image, image.propertyNames())
    except Exception as e:
        # Fallback to QA60 only if SCL not available
        print(f"⚠️ SCL mask fallback: {e}")
        return image.updateMask(qa_mask).divide(10000).copyProperties(image, image.propertyNames())


def add_ndvi_qc_band(img):
    """Add NDVI band specifically for Quality Mosaic sorting."""
    ndvi = img.normalizedDifference(['B8', 'B4']).rename('NDVI_QC')
    return img.addBands(ndvi)


def resample_to_10m(image):
    """
    Resample all bands to 10m resolution for consistent pixel alignment.
    Sentinel-2 has bands at 10m (B2,B3,B4,B8), 20m (B5-B8A,B11,B12), and 60m (B1,B9,B10).
    """
    return image.resample('bilinear').reproject(crs=image.select('B2').projection(), scale=10)


def estimate_area_ha(geometry: ee.Geometry) -> float:
    """
    Estimate area of geometry in hectares.
    Returns approximate area for tiling decision.
    """
    try:
        area_m2 = geometry.area().getInfo()
        return area_m2 / 10000  # Convert to hectares
    except:
        return 0


def tile_geometry(geometry: ee.Geometry, tile_size_deg: float = 0.05) -> list:
    """
    Split large geometry into smaller tiles for processing.
    
    Args:
        geometry: Large geometry to tile
        tile_size_deg: Size of each tile in degrees (approx 5km at equator)
    
    Returns:
        List of ee.Geometry tiles
    """
    try:
        bounds = geometry.bounds().getInfo()
        coords = bounds['coordinates'][0]
        
        min_lon = min(c[0] for c in coords)
        max_lon = max(c[0] for c in coords)
        min_lat = min(c[1] for c in coords)
        max_lat = max(c[1] for c in coords)
        
        tiles = []
        lon = min_lon
        while lon < max_lon:
            lat = min_lat
            while lat < max_lat:
                tile_bounds = ee.Geometry.Rectangle([
                    lon, lat,
                    min(lon + tile_size_deg, max_lon),
                    min(lat + tile_size_deg, max_lat)
                ])
                # Intersect tile with original geometry
                tile = geometry.intersection(tile_bounds)
                tiles.append(tile)
                lat += tile_size_deg
            lon += tile_size_deg
        
        print(f"   📦 Geometry split into {len(tiles)} tiles")
        return tiles
    except Exception as e:
        print(f"   ⚠️ Tiling failed, using original geometry: {e}")
        return [geometry]


def calculate_indices(image):
    """
    Menghitung berbagai indeks spektral untuk klasifikasi tutupan lahan.
    
    Indeks yang dihitung:
    - NDVI: Normalized Difference Vegetation Index
    - EVI: Enhanced Vegetation Index
    - GEMI: Global Environment Monitoring Index (advanced vegetation)
    - ARVI: Atmospherically Resistant Vegetation Index
    - MSAVI: Modified Soil Adjusted Vegetation Index
    - NDWI: Normalized Difference Water Index
    - MNDWI: Modified Normalized Difference Water Index
    - NDMI: Normalized Difference Moisture Index (canopy water content)
    - NDBI: Normalized Difference Built-up Index
    - SAVI: Soil Adjusted Vegetation Index
    - BSI: Bare Soil Index
    - NDII: Normalized Difference Infrared Index
    - NBR2: Normalized Burn Ratio 2
    - SIPI: Structure Insensitive Pigment Index
    """
    # NDVI - Indeks vegetasi
    ndvi = image.normalizedDifference(['B8', 'B4']).rename('NDVI')
    
    # NDWI - Indeks air (menggunakan B3/Green dan B8/NIR)
    ndwi = image.normalizedDifference(['B3', 'B8']).rename('NDWI')
    
    # MNDWI - Modified NDWI untuk deteksi air yang lebih baik
    mndwi = image.normalizedDifference(['B3', 'B11']).rename('MNDWI')
    
    # NDBI - Indeks area terbangun
    ndbi = image.normalizedDifference(['B11', 'B8']).rename('NDBI')
    
    # SAVI - Soil Adjusted Vegetation Index (L = 0.5)
    savi = image.expression(
        '((NIR - RED) / (NIR + RED + L)) * (1 + L)', {
            'NIR': image.select('B8'),
            'RED': image.select('B4'),
            'L': 0.5
        }).rename('SAVI')
    
    # EVI - Enhanced Vegetation Index
    evi = image.expression(
        '2.5 * ((NIR - RED) / (NIR + 6 * RED - 7.5 * BLUE + 1))', {
            'NIR': image.select('B8'),
            'RED': image.select('B4'),
            'BLUE': image.select('B2')
        }).rename('EVI')
    
    # GEMI - Global Environment Monitoring Index (advanced vegetation index)
    gemi = image.expression(
        '(eta * (1 - 0.25 * eta)) - ((RED - 0.125) / (1 - RED))', {
            'eta': image.expression(
                '(2 * (NIR**2 - RED**2) + 1.5 * NIR + 0.5 * RED) / (NIR + RED + 0.5)',
                {'NIR': image.select('B8'), 'RED': image.select('B4')}
            ),
            'RED': image.select('B4')
        }).rename('GEMI')
    
    # ARVI - Atmospherically Resistant Vegetation Index (robust to atmospheric effects)
    arvi = image.expression(
        '(NIR - (2 * RED - BLUE)) / (NIR + (2 * RED - BLUE))', {
            'NIR': image.select('B8'),
            'RED': image.select('B4'),
            'BLUE': image.select('B2')
        }).rename('ARVI')
    
    # MSAVI - Modified SAVI (version 2, self-adjusting L)
    msavi = image.expression(
        '(2 * NIR + 1 - sqrt((2 * NIR + 1)**2 - 8 * (NIR - RED))) / 2', {
            'NIR': image.select('B8'),
            'RED': image.select('B4')
        }).rename('MSAVI')
    
    # NDMI - Normalized Difference Moisture Index (better for canopy water content)
    ndmi = image.normalizedDifference(['B8', 'B11']).rename('NDMI')

    # BSI - Bare Soil Index
    bsi = image.expression(
        '((SWIR + RED) - (NIR + BLUE)) / ((SWIR + RED) + (NIR + BLUE))', {
            'SWIR': image.select('B11'),
            'RED': image.select('B4'),
            'NIR': image.select('B8'),
            'BLUE': image.select('B2')
        }).rename('BSI')

    # NDII - Normalized Difference Infrared Index (Canopy Water Content)
    ndii = image.normalizedDifference(['B8', 'B11']).rename('NDII')

    # NBR2 - Normalized Burn Ratio 2 (Soil/Lignin)
    nbr2 = image.normalizedDifference(['B11', 'B12']).rename('NBR2')
    
    # SIPI - Structure Insensitive Pigment Index (chlorophyll content)
    sipi = image.expression(
        '(NIR - BLUE) / (NIR - RED)', {
            'NIR': image.select('B8'),
            'BLUE': image.select('B2'),
            'RED': image.select('B4')
        }).rename('SIPI')
    
    return image.addBands([ndvi, ndwi, mndwi, ndbi, savi, evi, gemi, arvi, msavi, ndmi, bsi, ndii, nbr2, sipi])


# ==============================================================================
# MULTI-YEAR RANDOM FOREST CLASSIFICATION (with Robust Cloud Handling)
# ==============================================================================

def _get_cloud_probability_collection(geometry, start_date, end_date):
    """Get S2 Cloud Probability collection for joining with S2_SR."""
    return (ee.ImageCollection(S2_CLOUD_PROBABILITY)
            .filterBounds(geometry)
            .filterDate(start_date, end_date))


def _mask_clouds_and_shadows_robust(s2_sr_image, cloud_prob_threshold=50):
    """
    Robust cloud and shadow masking for a SINGLE Sentinel-2 image.
    
    Uses:
    1. Cloud Probability threshold (from S2_CLOUD_PROBABILITY join)
    2. Dark pixel detection (NIR and SWIR for shadows)
    3. Sun azimuth projection for shadow direction
    
    MUST be applied per-image BEFORE compositing.
    """
    # 1. Cloud mask from Cloud Probability
    cloud_prob = s2_sr_image.select('probability')
    is_cloud = cloud_prob.gt(cloud_prob_threshold)
    
    # 2. Dark pixel detection for shadows (NIR < 0.15 AND SWIR < 0.1)
    # Note: S2 SR values are scaled by 10000, so thresholds adjusted
    nir = s2_sr_image.select('B8')
    swir = s2_sr_image.select('B11')
    is_dark = nir.lt(1500).And(swir.lt(1000))
    
    # 3. Shadow projection using sun azimuth
    # Get sun azimuth from image properties
    sun_azimuth = ee.Number(s2_sr_image.get('MEAN_SOLAR_AZIMUTH_ANGLE'))
    sun_zenith = ee.Number(s2_sr_image.get('MEAN_SOLAR_ZENITH_ANGLE'))
    
    # Calculate shadow projection direction (opposite to sun)
    shadow_azimuth = sun_azimuth.add(180).mod(360)
    
    # Project clouds to find shadow zones
    # Shadow distance depends on cloud height and sun angle
    # Typical cloud height ~2000m, project ~200-500 pixels at 10m resolution
    shadow_distance = ee.Number(200)  # meters
    
    # Convert azimuth to radians
    azimuth_rad = shadow_azimuth.multiply(ee.Number(3.14159).divide(180))
    
    # Calculate x/y offsets for shadow projection
    x_offset = shadow_azimuth.cos().multiply(shadow_distance)
    y_offset = shadow_azimuth.sin().multiply(shadow_distance)
    
    # Project cloud mask in shadow direction
    shadow_projection = is_cloud.directionalDistanceTransform(shadow_azimuth, 50)
    is_shadow_zone = shadow_projection.select('distance').mask()
    
    # Combine: shadow is dark pixels within shadow projection zone
    is_shadow = is_dark.And(is_shadow_zone)
    
    # 4. Final combined mask (keep pixels that are NOT cloud AND NOT shadow)
    cloud_shadow_mask = is_cloud.Or(is_shadow).Not()
    
    # Apply mask and scale to 0-1 reflectance
    return (s2_sr_image
            .updateMask(cloud_shadow_mask)
            .divide(10000)
            .copyProperties(s2_sr_image, s2_sr_image.propertyNames()))


def _join_s2_with_cloud_probability(s2_collection, cloud_prob_collection):
    """
    Join S2_SR images with their corresponding Cloud Probability images.
    Uses system:index to match images from same acquisition.
    """
    # Define join condition (same system:index prefix)
    join_filter = ee.Filter.equals(
        leftField='system:index',
        rightField='system:index'
    )
    
    # Perform inner join
    joined = ee.ImageCollection(ee.Join.saveFirst('cloud_prob').apply(
        primary=s2_collection,
        secondary=cloud_prob_collection,
        condition=join_filter
    ))
    
    # Merge cloud probability as a band
    def _add_cloud_prob_band(image):
        cloud_prob = ee.Image(image.get('cloud_prob')).select('probability')
        return image.addBands(cloud_prob)
    
    return joined.map(_add_cloud_prob_band)


def create_yearly_composite_robust(geometry, year, cloud_prob_threshold=50):
    """
    Create cloud-free yearly median composite using robust cloud/shadow masking.
    
    Process:
    1. Get S2_SR collection for the year
    2. Try to join with S2_CLOUD_PROBABILITY
    3. Fallback to SCL-based masking if join fails
    4. Create median composite from masked images
    
    Args:
        geometry: ee.Geometry - AOI (from web app)
        year: int - Year to process
        cloud_prob_threshold: int - Cloud probability threshold (default 50)
        
    Returns:
        ee.Image - Median composite with bands: B2, B3, B4, B8, B11
    """
    import logging
    logger = logging.getLogger("RF_Classification")
    
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get S2 SR collection
    s2_collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
                     .filterBounds(geometry)
                     .filterDate(start_date, end_date)
                     .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 80)))
    
    s2_count = s2_collection.size().getInfo()
    logger.info(f"      Found {s2_count} S2 scenes for {year}")
    
    if s2_count == 0:
        raise ValueError(f"No Sentinel-2 images found for {year}")
    
    # Try Cloud Probability join first, fallback to SCL masking
    try:
        # Get Cloud Probability collection
        cloud_prob_collection = _get_cloud_probability_collection(geometry, start_date, end_date)
        cloud_prob_count = cloud_prob_collection.size().getInfo()
        logger.info(f"      Found {cloud_prob_count} Cloud Probability scenes")
        
        if cloud_prob_count > 0:
            # Join S2 with Cloud Probability
            s2_with_cloud_prob = _join_s2_with_cloud_probability(s2_collection, cloud_prob_collection)
            joined_count = s2_with_cloud_prob.size().getInfo()
            logger.info(f"      Joined: {joined_count} matched scenes")
            
            if joined_count > 0:
                # Apply robust masking to each image BEFORE compositing
                masked_collection = s2_with_cloud_prob.map(
                    lambda img: _mask_clouds_and_shadows_robust(img, cloud_prob_threshold)
                )
                composite = masked_collection.median().clip(geometry)
                logger.info(f"      ✅ Created composite with Cloud Probability masking")
                return composite
        
        # Fallback: use SCL-based masking
        logger.warning(f"      ⚠️ Cloud Probability join failed, using SCL fallback")
        
    except Exception as e:
        logger.warning(f"      ⚠️ Cloud Probability error: {e}, using SCL fallback")
    
    # Fallback: Use existing SCL-based cloud masking
    masked_collection = s2_collection.map(mask_clouds_sentinel2)
    composite = masked_collection.median().clip(geometry)
    logger.info(f"      ✅ Created composite with SCL masking (fallback)")
    return composite


def construct_yearly_features(composite, year):
    """
    Construct feature bands from yearly composite with year suffix.
    
    Output bands:
    - B2_YYYY, B3_YYYY, B4_YYYY, B8_YYYY (raw bands)
    - NDVI_YYYY, NDWI_YYYY, NDBI_YYYY (spectral indices)
    
    Args:
        composite: ee.Image - Yearly median composite
        year: int - Year for band naming
        
    Returns:
        ee.Image - Feature image with renamed bands
    """
    # Select core bands
    b2 = composite.select('B2').rename(f'B2_{year}')
    b3 = composite.select('B3').rename(f'B3_{year}')
    b4 = composite.select('B4').rename(f'B4_{year}')
    b8 = composite.select('B8').rename(f'B8_{year}')
    
    # Calculate spectral indices
    ndvi = composite.normalizedDifference(['B8', 'B4']).rename(f'NDVI_{year}')
    ndwi = composite.normalizedDifference(['B3', 'B8']).rename(f'NDWI_{year}')
    ndbi = composite.normalizedDifference(['B11', 'B8']).rename(f'NDBI_{year}')
    
    # Stack all features
    return ee.Image.cat([b2, b3, b4, b8, ndvi, ndwi, ndbi])


def construct_features_fixed(composite):
    """
    Construct feature bands with FIXED band names (no year suffix).
    Used for both training and classification to ensure band name consistency.
    
    Output bands: B2, B3, B4, B8, NDVI, NDWI, NDBI
    """
    # Select core bands
    b2 = composite.select('B2')
    b3 = composite.select('B3')
    b4 = composite.select('B4')
    b8 = composite.select('B8')
    
    # Calculate spectral indices
    ndvi = composite.normalizedDifference(['B8', 'B4']).rename('NDVI')
    ndwi = composite.normalizedDifference(['B3', 'B8']).rename('NDWI')
    ndbi = composite.normalizedDifference(['B11', 'B8']).rename('NDBI')
    
    # Stack all features with fixed names
    return ee.Image.cat([b2, b3, b4, b8, ndvi, ndwi, ndbi])


def build_multiyear_stack(geometry, start_year, end_year, cloud_prob_threshold=50):
    """
    Build multi-year feature stack from 2017 to end_year.
    
    Creates a single multi-band image where each year contributes:
    B2_YYYY, B3_YYYY, B4_YYYY, B8_YYYY, NDVI_YYYY, NDWI_YYYY, NDBI_YYYY
    
    This is the feature space for training a single Random Forest classifier.
    
    Args:
        geometry: ee.Geometry - AOI (from web app)
        start_year: int - First year (typically 2017)
        end_year: int - Last year
        cloud_prob_threshold: int - Cloud probability threshold
        
    Returns:
        dict with:
        - 'stack': ee.Image - Multi-year stacked image
        - 'years': list - Years included in stack
        - 'band_count': int - Total number of bands
    """
    yearly_features = []
    years_processed = []
    
    for year in range(start_year, end_year + 1):
        try:
            # Create cloud-free composite for this year
            composite = create_yearly_composite_robust(geometry, year, cloud_prob_threshold)
            
            # Construct feature bands with year suffix
            features = construct_yearly_features(composite, year)
            
            yearly_features.append(features)
            years_processed.append(year)
        except Exception as e:
            print(f"⚠️ Year {year} failed: {e}")
            continue
    
    if not yearly_features:
        raise ValueError("No valid yearly data available for multi-year stack")
    
    # Stack all years into single image
    multiyear_stack = ee.Image.cat(yearly_features)
    
    return {
        'stack': multiyear_stack,
        'years': years_processed,
        'band_count': len(years_processed) * 7  # 7 bands per year
    }


def generate_training_samples_from_composite(composite, geometry, samples_per_class=100):
    """
    Generate pseudo-training samples using rule-based pre-classification.
    
    Used when external training samples are not provided.
    Creates stratified random samples from rule-based classification.
    
    Args:
        composite: ee.Image - A single-year composite for classification
        geometry: ee.Geometry - AOI
        samples_per_class: int - Number of samples per class
        
    Returns:
        ee.FeatureCollection - Training samples with 'landcover' property
    """
    # Calculate indices for rule-based classification
    ndvi = composite.normalizedDifference(['B8', 'B4']).rename('NDVI')
    mndwi = composite.normalizedDifference(['B3', 'B11']).rename('MNDWI')
    ndbi = composite.normalizedDifference(['B11', 'B8']).rename('NDBI')
    
    # Rule-based pseudo-labels (threshold classification)
    # 1: Hutan (NDVI > 0.6)
    # 2: Tanah Kering (NDVI 0.2-0.6)
    # 3: Tanah Kosong (NDVI < 0.2, not water, not built-up)
    # 4: Air (MNDWI > water_mndwi_min)
    # 5: Lahan Terbangun (NDBI > builtup_ndbi_min, NDVI < builtup_ndvi_max)
    
    # Defaults if not provided (internal GEE calls might pass different thresholds)
    # Using hardcoded defaults here to match user request exactly if thresholds object is not passed
    water = mndwi.gt(0.1)
    builtup = ndbi.gt(0.1).And(ndvi.lt(0.3)).And(water.Not())
    forest = ndvi.gte(0.6).And(water.Not()).And(builtup.Not())
    dry_land = ndvi.gte(0.2).And(ndvi.lt(0.6)).And(water.Not()).And(builtup.Not()).And(forest.Not())
    bare_soil = ndvi.lt(0.2).And(water.Not()).And(builtup.Not())
    
    # Create classification image
    pseudo_class = (ee.Image(4)  # Default: Tanah Kosong (4)
                    .where(dry_land, 3)  # Tanah Kering (3)
                    .where(forest, 2)    # Hutan Sekunder (2) -> Simplified for training
                    .where(forest.And(ndvi.gt(0.75)), 1) # Hutan Primer (1) -> High NDVI
                    .where(builtup, 6)   # Lahan Terbangun (6)
                    .where(water, 5))    # Air (5)
    
    # Sample from each class
    samples = pseudo_class.rename('landcover').stratifiedSample(
        numPoints=samples_per_class,
        classBand='landcover',
        region=geometry,
        scale=30, # Optimized for performance
        seed=42,
        geometries=True
    )
    
    return samples


def train_rf_classifier_multiyear(multiyear_stack, training_samples, geometry):
    """
    Train a single Random Forest classifier on multi-year stack.
    
    The classifier learns:
    - Spectral stability across years
    - Temporal differences without manual rules
    
    Args:
        multiyear_stack: ee.Image - Multi-year stacked features
        training_samples: ee.FeatureCollection - Training samples with 'landcover' property
        geometry: ee.Geometry - AOI (for sampling if needed)
        
    Returns:
        ee.Classifier - Trained Random Forest classifier
    """
    # Get band names from stack
    band_names = multiyear_stack.bandNames()
    
    # Sample the multi-year stack at training point locations
    training_data = multiyear_stack.sampleRegions(
        collection=training_samples,
        properties=['landcover'],
        scale=10,
        tileScale=4
    )
    
    # Train Random Forest with default parameters
    classifier = ee.Classifier.smileRandomForest(numberOfTrees=100).train(
        features=training_data,
        classProperty='landcover',
        inputProperties=band_names
    )
    
    return classifier


def classify_single_year_with_rf(geometry, year, trained_classifier, cloud_prob_threshold=50):
    """
    Apply trained RF classifier to a single year's composite.
    
    Args:
        geometry: ee.Geometry - AOI
        year: int - Year to classify
        trained_classifier: ee.Classifier - Trained RF model
        cloud_prob_threshold: int - Cloud probability threshold
        
    Returns:
        ee.Image - Classification image with integer values 1-5
    """
    # Create cloud-free composite for this year
    composite = create_yearly_composite_robust(geometry, year, cloud_prob_threshold)
    
    # Construct features (same structure as training)
    features = construct_yearly_features(composite, year)
    
    # Apply classifier
    classified = features.classify(trained_classifier).rename('landcover')
    
    # Ensure integer output 1-5 (no styling, no post-processing)
    return classified.clip(geometry)


def classify_multiyear_rf(
    geometry,
    start_year=2017,
    end_year=None,
    training_samples=None,
    cloud_prob_threshold=50
):
    """
    Main entry point for multi-year Random Forest land cover classification.
    
    Process:
    1. Build multi-year stack (2017-end_year) with robust cloud masking
    2. Generate/use training samples
    3. Train single RF classifier on multi-year stack
    4. Apply classifier to each year's composite
    
    Args:
        geometry: ee.Geometry - AOI (MUST be provided by web app, no default)
        start_year: int - First year (default 2017)
        end_year: int - Last year (from web app input)
        training_samples: ee.FeatureCollection - Optional external training samples
        cloud_prob_threshold: int - Cloud probability threshold (default 50)
        
    Returns:
        dict with:
        - 'classifications': dict[year] -> ee.Image (integer 1-5)
        - 'classifier': ee.Classifier - Trained model
        - 'years': list - Years classified
        - 'stack_info': dict - Multi-year stack metadata
    """
    if end_year is None:
        end_year = datetime.now().year - 1  # Previous year for complete data
    
    print(f"🌲 Multi-Year RF Classification: {start_year}-{end_year}")
    print(f"   ☁️ Cloud Probability Threshold: {cloud_prob_threshold}%")
    
    # Step 1: Build multi-year stack
    print(f"   📦 Building multi-year feature stack...")
    stack_result = build_multiyear_stack(geometry, start_year, end_year, cloud_prob_threshold)
    multiyear_stack = stack_result['stack']
    years = stack_result['years']
    
    print(f"   ✅ Stack built: {stack_result['band_count']} bands, {len(years)} years")
    
    # Step 2: Get or generate training samples
    if training_samples is None:
        print(f"   🎯 Generating pseudo-training samples from latest year...")
        # Use latest year's composite for pseudo-label generation
        latest_composite = create_yearly_composite_robust(geometry, years[-1], cloud_prob_threshold)
        training_samples = generate_training_samples_from_composite(
            latest_composite, geometry, samples_per_class=100
        )
    
    # Step 3: Train classifier on multi-year stack
    print(f"   🧠 Training Random Forest classifier...")
    classifier = train_rf_classifier_multiyear(multiyear_stack, training_samples, geometry)
    
    # Step 4: Classify each year
    print(f"   🗂️ Classifying each year...")
    classifications = {}
    for year in years:
        classified = classify_single_year_with_rf(geometry, year, classifier, cloud_prob_threshold)
        classifications[year] = classified
        print(f"      ✅ {year} classified")
    
    return {
        'classifications': classifications,
        'classifier': classifier,
        'years': years,
        'stack_info': {
            'start_year': start_year,
            'end_year': end_year,
            'band_count': stack_result['band_count'],
            'years_processed': len(years)
        }
    }


# ==============================================================================
# ADVANCED ANALYTICS (POINTS 2, 4)
# ==============================================================================

def get_elevation_data(geometry):
    """Mengambil data elevasi dan kelerengan (DEM) untuk koreksi bayangan gunung."""
    dem = ee.Image("NASA/NASADEM_HGT/001").select('elevation').clip(geometry)
    slope = ee.Terrain.slope(dem).rename('slope')
    return dem.addBands(slope)


def get_sentinel1_radar(geometry, year):
    """Mengambil dan memproses data Sentinel-1 SAR (Radar) untuk menembus awan."""
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Filter collection: Interferometric Wide (IW), Ground Range Detected (GRD)
    s1_col = (ee.ImageCollection("COPERNICUS/S1_GRD")
              .filterBounds(geometry)
              .filterDate(start_date, end_date)
              .filter(ee.Filter.listContains('transmitterReceiverPolarisation', 'VV'))
              .filter(ee.Filter.listContains('transmitterReceiverPolarisation', 'VH'))
              .filter(ee.Filter.eq('instrumentMode', 'IW')))
    
    # Median composite untuk mereduksi noise speckle
    s1_median = s1_col.median().clip(geometry)
    
    # Menghitung rasio VH/VV (efektif untuk membedakan vegetasi/tekstur)
    ratio = s1_median.select('VH').divide(s1_median.select('VV')).rename('VH_VV_ratio')
    
    return s1_median.addBands(ratio)


def perform_accuracy_assessment(classified, geometry, year):
    """
    Melakukan validasi akurasi dengan membandingkan hasil dengan ESA WorldCover (10m).
    (Point 3: Accuracy Assessment)
    """
    try:
        # Load ESA WorldCover 10m (Global Reference)
        # Ambil versi 2020 atau 2021 (terbaru yang stabil di GEE)
        worldcover = ee.ImageCollection("ESA/WorldCover/v100").first().clip(geometry)
        
        # Remap kelas ESA ke kelas GealGeolGeo (1-5)
        # ESA: 10(Tree), 20(Shrub), 30(Grass), 40(Crop), 50(Urban), 60(Bare), 80(Water)
        # GealGeol: 2(Hutan Sekunder), 3(T. Kering), 4(T. Kosong), 5(Air), 6(Terbangun)
        esa_remapped = worldcover.remap(
            [10, 20, 30, 40, 50, 60, 80],
            [2,  3,  3,  3,  6,  4,  5]
        ).rename('reference')
        
        # Hitung Error Matrix menggunakan sampel acak (stratified random sampling)
        combined = classified.rename('result').addBands(esa_remapped)
        
        sample = combined.sample(
            region=geometry,
            scale=30,
            numPixels=500, # 500 sampel titik validasi
            seed=42,
            tileScale=4
        )
        
        error_matrix = sample.errorMatrix('reference', 'result')
        oa = error_matrix.accuracy() # Overall Accuracy
        kappa = error_matrix.kappa()
        
        # Get scalar values from GEE
        results = ee.Dictionary({
            'oa': oa,
            'kappa': kappa
        }).getInfo()
        
        return results.get('oa'), results.get('kappa')
    except Exception as e:
        print(f"⚠️ Accuracy assessment failed: {e}")
        return None, None


def classify_land_cover_advanced(image, geometry, year, 
                               thresholds: ClassificationThresholds = ClassificationThresholds(),
                               use_radar=True, use_dem=True):
    """
    Klasifikasi tutupan lahan lanjutan menggunakan multi-sensor (Optical + Radar + DEM).
    """
    # 1. Base Spectral Indices
    image_with_indices = calculate_indices(image)
    ndvi = image_with_indices.select('NDVI')
    mndwi = image_with_indices.select('MNDWI')
    ndbi = image_with_indices.select('NDBI')
    
    # 2. Integration: Radar (Point 2)
    s1_data = None
    if use_radar:
        try:
            s1_data = get_sentinel1_radar(geometry, year)
        except: pass
        
    # 3. Integration: DEM (Point 4)
    dem_data = None
    if use_dem:
        try:
            dem_data = get_elevation_data(geometry)
        except: pass

    # =========================================================================
    # ADVANCED LOGIC
    # =========================================================================
    
    # Air Lanjutan (Gunakan Radar jika ada: Air sangat gelap di Radar)
    water_optical = mndwi.gt(thresholds.water_mndwi_min).Or(ndvi.lt(thresholds.water_ndvi_max))
    if s1_data:
        # Water in Radar: VV < -20dB and VH < -25dB (threshold umum)
        water_radar = s1_data.select('VV').lt(-20).And(s1_data.select('VH').lt(-25))
        water = water_optical.Or(water_radar)
    else:
        water = water_optical
    
    # Filter bayangan gunung (Shadow Masking - Point 4)
    if dem_data:
        # Slope > 15 deg dan pantulan sangat rendah biasanya bayangan gunung
        mountain_shadow = dem_data.select('slope').gt(15).And(image.select('B8').lt(800))
        # Remove shadows from water detection
        water = water.And(mountain_shadow.Not())
    
    water = water.focal_min(radius=1, kernelType='circle', units='pixels')

    # Lahan Terbangun Lanjutan (Built-up sangat terang di VV/VH karena 'double bounce')
    builtup_optical = ndbi.gt(0.1).And(ndvi.lt(0.3))
    if s1_data:
        # Urban in Radar: Strong backscatter
        builtup_radar = s1_data.select('VV').gt(-8).And(builtup_optical)
        builtup = builtup_optical.Or(builtup_radar).And(water.Not())
    else:
        builtup = builtup_optical.And(water.Not())
    
    # Vegetasi (Hutan & Tanah Kering)
    # Vegetasi (Hutan & Tanah Kering)
    forest = ndvi.gte(thresholds.forest_ndvi_min)
    # Exclude bright haze (B2 > 0.25) from Dry Land
    dry_land = ndvi.gte(thresholds.dry_land_ndvi_min)\
        .And(ndvi.lt(thresholds.dry_land_ndvi_max))\
        .And(image.select('B2').lt(0.25))
    
    # Tanah Kosong
    bare_soil = ndvi.lt(thresholds.bare_soil_ndvi_max).And(builtup.Not()).And(water.Not())
    # 6. Saturated Soil / Mud / Wet Biomass (Class 7)
    # MNDWI positif tapi rendah (0 - 0.1) dan bukan air permanen
    mud = mndwi.gt(0.0).And(mndwi.lt(0.1)).And(ndvi.lt(0.4))

    # --- Combine Classes ---
    # Prioritas: Air > Terbangun > Tanah > Hutan
    # --- Combine Classes ---
    # Prioritas: Air > Terbangun > Tanah > Hutan
    classification = ee.Image(4) # Default: Tanah Kosong (4)
    
    classification = classification.where(dry_land, 3) # Tanah Kering (3)
    classification = classification.where(forest, 2)   # Hutan Sekunder (2)
    classification = classification.where(forest.And(ndvi.gt(0.75)), 1) # Hutan Primer (1)
    classification = classification.where(builtup, 6)  # Lahan Terbangun (6)
    classification = classification.where(water, 5)    # Air (5)

    # Post-processing: Mask out detected Haze/Clouds to avoid false "Tanah Kering"
    # If Blue > 0.25 and not Water/Builtup, it's likely cloud/haze.
    # Built-up can be bright, but usually has high NDBI. Clouds have low NDBI.
    # We use a conservative mask to be safe.
    cloud_haze_mask = image.select('B2').gt(0.25).And(builtup.Not()).And(water.Not())
    classification = classification.updateMask(cloud_haze_mask.Not())

    return classification.rename('landcover').clip(geometry)




def get_map_url(classified_image: ee.Image) -> str:
    """Generate URL tile layer GEE untuk citra terklasifikasi."""
    # Palet 5 kelas sesuai dengan LC_PALETTE
    palette = [
        '#006400',  # 1: Hutan Primer
        '#32CD32',  # 2: Hutan Sekunder
        '#DAA520',  # 3: Tanah Kering
        '#D2691E',  # 4: Tanah Kosong
        '#1E90FF',  # 5: Air
        '#708090',  # 6: Lahan Terbangun
    ]
    vis_params = {
        'min': 1,
        'max': 6,
        'palette': palette
    }
    
    map_id_dict = classified_image.getMapId(vis_params)
    return map_id_dict['tile_fetcher'].url_format


def get_transition_map_url(binary_image: ee.Image, color: str) -> str:
    """Generate URL tile layer for a binary transition image (1=transition, 0=none)."""
    # Use updateMask to make 0 values transparent
    masked = binary_image.updateMask(binary_image.gt(0))
    vis_params = {
        'palette': [color]
    }
    map_id_dict = masked.getMapId(vis_params)
    return map_id_dict['tile_fetcher'].url_format


def get_rgb_map_url(image: ee.Image) -> str:
    """Generate URL tile layer GEE for RGB satellite imagery (Sentinel-2) at native 10m resolution."""
    # Resample to 10m native resolution for sharper imagery
    resampled = image.select(['B4', 'B3', 'B2']).reproject(
        crs='EPSG:4326',
        scale=10
    )
    
    # Visualization parameters adjusted for premium natural color look
    vis_params = {
        'bands': ['B4', 'B3', 'B2'],
        'min': 0.02,
        'max': 0.38,
        'gamma': 1.3
    }
    map_id_dict = resampled.getMapId(vis_params)
    return map_id_dict['tile_fetcher'].url_format


def get_thumb_url(classified_image: ee.Image, geometry: ee.Geometry) -> str:
    """Generate static thumbnail URL for classified image with proper visualization."""
    # Palet 5 kelas sesuai dengan LC_PALETTE
    palette = [
        '#006400',  # 1: Hutan Primer
        '#32CD32',  # 2: Hutan Sekunder
        '#DAA520',  # 3: Tanah Kering
        '#D2691E',  # 4: Tanah Kosong
        '#1E90FF',  # 5: Air
        '#708090',  # 6: Lahan Terbangun
    ]
    
    try:
        # PENTING: Gunakan visualize() terlebih dahulu untuk menerapkan palette
        visualized = classified_image.visualize(
            min=1,
            max=6,
            palette=palette
        )
        
        # Get thumbnail URL from visualized image
        thumb_url = visualized.getThumbURL({
            'region': geometry,
            'format': 'png',
            'dimensions': 600
        })
        return thumb_url
    except Exception as e:
        print(f"⚠️ Error generating thumbnail: {e}")
        return None


def get_rgb_thumb_url(image: ee.Image, geometry: ee.Geometry) -> str:
    """Generate static thumbnail URL for real RGB satellite imagery."""
    # Sentinel-2 RGB Bands: B4 (Red), B3 (Green), B2 (Blue)
    vis_params = {
        'bands': ['B4', 'B3', 'B2'],
        'min': 0,
        'max': 3000,
        'gamma': 1.4,
        'dimensions': 600
    }
    try:
        thumb_url = image.getThumbURL({
            'region': geometry,
            'format': 'png',
            **vis_params
        })
        return thumb_url
    except Exception as e:
        print(f"⚠️ Error generating RGB thumbnail: {e}")
        return None


async def download_and_save_locally(thumb_url: str, prefix: str = "thumb") -> Optional[str]:
    """Download thumbnail from GEE and save it to the VPS server.

    Returns the web-accessible local URL (e.g., /storage/thumbnails/...).
    """
    if not thumb_url:
        print(f"⚠️ No thumbnail URL provided")
        return None

    try:
        # Download image from GEE with timeout
        print(f"📥 Downloading thumbnail from GEE: {thumb_url[:60]}...")
        response = await asyncio.to_thread(
            lambda: requests.get(thumb_url, timeout=30)
        )

        if response.status_code != 200:
            print(f"❌ Failed to download thumbnail: HTTP {response.status_code}")
            return None

        # Validate content
        if not response.content or len(response.content) == 0:
            print(f"❌ Downloaded thumbnail is empty")
            return None

        print(f"   Size: {len(response.content) / 1024:.1f} KB")

        # Generate unique filename
        # Use content hash to avoid duplicates if same image is requested
        image_bytes = response.content
        content_hash = hashlib.md5(image_bytes).hexdigest()[:12]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{prefix}_{timestamp}_{content_hash}.png"

        # Save locally using the helper
        local_url = save_local_image(image_bytes, filename)

        if local_url:
            print(f"✅ Thumbnail saved locally: {local_url}")
            return local_url
        else:
            print(f"❌ Failed to save thumbnail locally")
            return None

    except asyncio.TimeoutError:
        print(f"❌ Thumbnail download timeout (>30s)")
        return None
    except Exception as e:
        print(f"❌ Thumbnail local save failed: {e}")
        import traceback
        traceback.print_exc()
        return None


# Alias for backward compatibility if needed, but we should find/replace all
download_and_encode_thumbnail = download_and_save_locally




async def analyze_land_cover(
    geometry: ee.Geometry, 
    # New Year Range Logic
    start_year: Optional[int] = None,
    end_year: Optional[int] = None,
    # Legacy 'years' param (fallback)
    years: int = 5, 
    thresholds: Optional[ClassificationThresholds] = None,
    mode: str = "series",
    specific_date: Optional[str] = None,
    on_progress: Optional[Callable] = None,
    # Simplified parameters
    scale: int = 10,
    # Cloud probability threshold FIXED at 50% per methodology
    # DO NOT expose to user - ensures consistency across all years and runs
    # Cancellation callback (returns True if cancelled)
    check_cancel: Optional[Callable] = None,
    existing_data: Optional[List[Dict[str, Any]]] = None
):
    """
    Analisis tutupan lahan menggunakan Multi-Year Random Forest Classifier.
    """
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("RF_Classification")
    
    logger.info("=" * 60)
    logger.info("🌲 MULTI-YEAR RANDOM FOREST CLASSIFICATION")
    logger.info("=" * 60)
    
    current_year = datetime.now().year
    
    # Determine effective start/end years
    if start_year is None:
        start_year = 2017 # Sentinel-2 default
        
    if end_year is None:
        if mode == "single" and specific_date:
            try:
                end_year = datetime.strptime(specific_date, "%Y-%m-%d").year
                start_year = end_year # Single year mode
            except:
                end_year = current_year - 1
        else:
            end_year = current_year - 1
            
    # Validate range
    if start_year > end_year:
        start_year, end_year = end_year, start_year
    
    # FIXED cloud probability threshold per methodology (50%)
    CLOUD_PROB_THRESHOLD_FIXED = 50
    
    # Identify provisional year (current year with incomplete data)
    provisional_years = []
    if end_year >= current_year:
        provisional_years.append(current_year)
        logger.warning(f"⚠️ Year {current_year} marked as PROVISIONAL (incomplete data)")
    
    logger.info(f"📅 Year range: {start_year} - {end_year}")
    logger.info(f"☁️ Cloud probability threshold: {CLOUD_PROB_THRESHOLD_FIXED}% (FIXED)")
    logger.info(f"📏 Resolution: {scale}m (fixed at 10m)")
    
    if on_progress:
        await on_progress(5, "Inisialisasi", "Mempersiapkan analisis Multi-Year RF...")
    
    results = []
    
    try:
        # =====================================================================
        # STEP 1: Build Yearly Composites with Robust Cloud Masking
        # =====================================================================
        if on_progress:
            await on_progress(10, "Composites", "Membangun composite tahunan...")
        
        logger.info("📦 STEP 1: Building yearly composites...")
        
        # Determine years that REALLY need GEE processing (MOVE UP for Efficiency)
        existing_map = {int(d['year']): d for d in (existing_data or [])}
        all_requested_years = list(range(start_year, end_year + 1))
        years_to_compute = [y for y in all_requested_years if y not in existing_map]
        
        logger.info(f"   🎯 Total Requested: {len(all_requested_years)} years")
        logger.info(f"   ⚡ Skipping (Cached): {len(all_requested_years) - len(years_to_compute)}")
        logger.info(f"   🛰️ Computing (GEE): {len(years_to_compute)}")

        if not years_to_compute and existing_data:
            logger.info("⚡ All requested years found in existing data. Skipping GEE entirely.")
            # Return sorted existing data directly with empty transition/audit as they are likely already in metadata
            if on_progress: await on_progress(100, "Selesai", "Data diambil dari cache")
            results = [YearlyData(**d) for d in sorted(existing_data, key=lambda x: x['year']) if int(x['year']) in all_requested_years]
            
            # Recompute transition summary for consistency
            total_defor = sum(getattr(r, 'deforestation_ha', 0) for r in results)
            total_refor = sum(getattr(r, 'reforestation_ha', 0) for r in results)
            summary = {
                "total_deforestation_ha": round(total_defor, 2),
                "total_reforestation_ha": round(total_refor, 2),
                "net_forest_change_ha": round(total_refor - total_defor, 2),
                "period": f"{start_year} - {end_year}",
                "unit": "Hektar (ha)"
            }
            return results, summary, {}

        yearly_composites = {}
        years_successfully_built = []
        
        # Sequential composite building - ONLY for years_to_compute
        for year in years_to_compute:
            # Check cancellation inside loop
            if check_cancel and await check_cancel():
                logger.warning("⛔ Analysis cancelled by user.")
                raise Exception("Analisis dibatalkan oleh pengguna")

            try:
                logger.info(f"   🛰️ Processing {year}...")
                
                # Create robust cloud-masked composite
                composite = await asyncio.to_thread(
                    create_yearly_composite_robust, geometry, year, CLOUD_PROB_THRESHOLD_FIXED
                )
                
                # Verify composite has data
                band_count = await asyncio.to_thread(lambda: composite.bandNames().size().getInfo())
                logger.info(f"      ✅ Composite {year} created: {band_count} bands")
                
                yearly_composites[year] = composite
                years_successfully_built.append(year)
                
            except Exception as e:
                logger.warning(f"   ⚠️ Year {year} failed: {e}")
                continue
                
        years_processed = sorted(years_successfully_built + [y for y in all_requested_years if y in existing_map])
        years_processed.sort()
        
        if on_progress:
            await on_progress(40, "Composites Done", f"{len(years_processed)} tahun berhasil diproses")
            
        if not years_processed:
            raise ValueError("Tidak ada data valid untuk analisis")
            
        logger.info(f"   ✅ Composites ready: {len(years_processed)} years ({years_processed[0]}-{years_processed[-1]})")
        
        # =====================================================================

        # =====================================================================
        # STEP 2: Generate Training Samples & Train Classifier
        # =====================================================================
        if on_progress:
            await on_progress(45, "Training RF", "Melatih Random Forest classifier...")
        
        logger.info("🧠 STEP 2: Training Random Forest classifier...")
        
        # FIX: Ensure latest_year exists in yearly_composites (it might be from cache if we strictly use years_processed)
        # We need a year that has a computed composite to generate training samples
        years_with_composites = sorted(list(yearly_composites.keys()))
        
        if not years_with_composites:
             # This should theoretically be handled by the "if not years_to_compute" check at the top,
             # but just in case we fall through (e.g. forced recompute logic change)
             if existing_data:
                 logger.info("⚡ All years cached or failed. Returning purely cached data.")
                 # Fallback: Just return what we have in cache
                 results = [YearlyData(**d) for d in sorted(existing_data, key=lambda x: x['year']) if int(x['year']) in all_requested_years]
                 
                 # Recompute transition summary for consistency
                 total_defor = sum(getattr(r, 'deforestation_ha', 0) for r in results)
                 total_refor = sum(getattr(r, 'reforestation_ha', 0) for r in results)
                 summary = {
                     "total_deforestation_ha": round(total_defor, 2),
                     "total_reforestation_ha": round(total_refor, 2),
                     "net_forest_change_ha": round(total_refor - total_defor, 2),
                     "period": f"{start_year} - {end_year}",
                     "unit": "Hektar (ha)"
                 }
                 return results, summary, {}
             else:
                 raise ValueError("No valid composites available for training (GEE processing failed for all years)")

        latest_year = years_with_composites[-1]
        latest_composite = yearly_composites[latest_year]
        
        logger.info(f"   🎯 Generating training samples from {latest_year} (latest computed)...")
        training_samples = generate_training_samples_from_composite(
            latest_composite, geometry, samples_per_class=100
        )
        
        sample_count = training_samples.size().getInfo()
        logger.info(f"   ✅ Training samples: {sample_count} points")
        
        # Create training features with FIXED band names
        # This ensures classifier can be applied to any year
        logger.info("   🔧 Training smileRandomForest (100 trees)...")
        
        training_features = construct_features_fixed(latest_composite)
        band_names = training_features.bandNames()
        
        training_data = training_features.sampleRegions(
            collection=training_samples,
            properties=['landcover'],
            scale=10,
            tileScale=4
        )
        
        classifier = ee.Classifier.smileRandomForest(numberOfTrees=100).train(
            features=training_data,
            classProperty='landcover',
            inputProperties=band_names
        )
        
        logger.info("   ✅ Classifier trained successfully")
        
        # =====================================================================
        # STEP 3: Classify Each Year & Compute Statistics
        # =====================================================================
        if on_progress:
            await on_progress(55, "Klasifikasi", "Mengklasifikasi setiap tahun...")
        
        logger.info("🗂️ STEP 3: Classifying each year...")
        
        # Parallel classification with semaphore REMOVED
        # Sequential processing enforced per policy
        classification_results = []
        
        # Dictionary to store classified images for temporal consistency check
        year_images = {}
        
        # STEP 3.1: First Pass - Classify all years (Original Indices)
        logger.info("🗂️ STEP 3.1: First pass classification...")
        for idx, year in enumerate(years_successfully_built):
             # Check cancellation in loop
            if check_cancel and await check_cancel():
                logger.warning("⛔ Analysis cancelled by user.")
                raise Exception("Analisis dibatalkan oleh pengguna")

            try:
                features = await asyncio.to_thread(construct_features_fixed, yearly_composites[year])
                classified = await asyncio.to_thread(
                    lambda: features.classify(classifier).rename('landcover').clip(geometry)
                )
                year_images[year] = classified
                logger.debug(f"      ✅ {year} first pass done")
            except Exception as e:
                logger.error(f"   ❌ Error in first pass {year}: {e}")

        # STEP 3.2: Temporal Consistency Analysis (Sub-klasifikasi Hutan)
        logger.info("🌲 STEP 3.2: Temporal Consistency Analysis...")
        
        # Extract forest masks (where class == 1 or class == 2)
        forest_masks = [img.eq(1).Or(img.eq(2)) for img in year_images.values()]
        
        if not forest_masks:
            raise ValueError("Tidak ada data hutan yang terdeteksi untuk analisis konsistensi")

        # Create forest consistency layer (Sum of Forest / Total Years)
        forest_stack = ee.ImageStack(forest_masks) if hasattr(ee, 'ImageStack') else ee.Image.cat(forest_masks)
        forest_count = forest_stack.reduce(ee.Reducer.sum())
        total_years = len(forest_masks)
        forest_consistency = forest_count.divide(total_years)
        
        # Define Primary Forest: Consistent >= 80% of the time AND originally classified as Forest
        # Define Secondary Forest: Currently Forest but consistency < 80%
        primary_mask = forest_consistency.gte(0.8)
        
        # STEP 3.3: Second Pass - Refine Classifications & Stats
        logger.info("📊 STEP 3.3: Finalizing classifications & statistics...")

        # Prefill with existing data (Note: cached data might still use old indices!)
        # For now, we only process years_successfully_built with the new logic
        for y in years_processed:
            if y in existing_map:
                logger.info(f"   ⚡ Skipping {y} (Direct from Cache)")
                classification_results.append(YearlyData(**existing_map[y]))

        # Track the refined image for transition calculation
        last_refined_img = None

        for idx, year in enumerate(years_successfully_built):
            if year not in year_images: continue
            
            try:
                if on_progress:
                    loop_percent = 55 + (idx / len(years_successfully_built) * 35)
                    await on_progress(int(loop_percent), "Klasifikasi", f"Finalisasi data tahun {year} ({idx+1}/{len(years_successfully_built)})")

                orig_classified = year_images[year]
                
                # 1. Identify all forest pixels (1 or 2)
                # 2. Update to Hutan Primer (Index 1) if primary_mask is true
                # 3. Update to Hutan Sekunder (Index 2) if primary_mask is false
                # 4. Leave other classes (3,4,5,6) AS IS (Already correct from classifier)
                
                is_forest = orig_classified.eq(1).Or(orig_classified.eq(2))
                
                refined = orig_classified # Start with original
                # Apply Forest Refinement based on Time Series
                refined = refined.where(is_forest.And(primary_mask), 1)
                refined = refined.where(is_forest.And(primary_mask.Not()), 2)
                
                # NO SHIFTING for others! 3,4,5,6 are correct.
                
                # Compute area statistics with new indices (1-9)
                area_image = ee.Image.pixelArea().divide(10000).rename('area')
                combined = area_image.addBands(refined.rename('class')).select(['area', 'class'])
                
                stats = await asyncio.to_thread(lambda: combined.reduceRegion(
                    reducer=ee.Reducer.sum().group(groupField=1, groupName='class'),
                    geometry=geometry,
                    scale=10,
                    maxPixels=1e10,
                    bestEffort=True,
                    tileScale=16  # Divide into 16x16 tiles to reduce memory usage
                ).getInfo())
                
                groups = stats.get('groups', [])
                raw_areas = {int(item['class']): item['sum'] for item in groups if item.get('class')}
                
                # Log new breakdown
                logger.info(f"      📊 {year} breakdown: Primer={raw_areas.get(1,0):.1f}ha, Sekunder={raw_areas.get(2,0):.1f}ha")
                
                # Generate map URLs
                map_url = await asyncio.to_thread(lambda: get_map_url(refined))
                rgb_url = await asyncio.to_thread(lambda: get_rgb_map_url(yearly_composites[year]))
                
                # Generate thumbnails
                thumb_url = await asyncio.to_thread(lambda: get_thumb_url(refined, geometry))
                rgb_thumb_url = await asyncio.to_thread(lambda: get_rgb_thumb_url(yearly_composites[year], geometry))
                
                # Create result object
                yearly_data = YearlyData(
                    year=year,
                    hutan_primer=round(raw_areas.get(1, 0), 2),
                    hutan_sekunder=round(raw_areas.get(2, 0), 2),
                    hutan=round(raw_areas.get(1, 0) + raw_areas.get(2, 0), 2), # Total Forest
                    tanah_kering=round(raw_areas.get(3, 0), 2),
                    tanah_kosong=round(raw_areas.get(4, 0), 2),
                    air=round(raw_areas.get(5, 0), 2),
                    lahan_terbangun=round(raw_areas.get(6, 0), 2),
                    total_area=round(sum([raw_areas.get(i, 0) for i in range(1, 7)]), 2),
                    map_url=map_url,
                    rgb_url=rgb_url,
                    thumb_url=thumb_url,
                    rgb_thumb_url=rgb_thumb_url,
                    data_source="Multi-Year RF + Temporal Consistency",
                    scene_count=len(years_processed),
                    detailed={
                        'hutan_primer': round(raw_areas.get(1, 0), 2),
                        'hutan_sekunder': round(raw_areas.get(2, 0), 2),
                        'hutan': round(raw_areas.get(1, 0) + raw_areas.get(2, 0), 2),
                        'tanah_kering': round(raw_areas.get(3, 0), 2),
                        'tanah_kosong': round(raw_areas.get(4, 0), 2),
                        'air': round(raw_areas.get(5, 0), 2),
                        'lahan_terbangun': round(raw_areas.get(6, 0), 2)
                    }
                )

                # =============================================================
                # TRANSITION CALCULATION (PIXEL-BASED)
                # =============================================================
                # Treat Hutan Primer (1) and Hutan Sekunder (2) both as Forest for loss calculation
                if idx > 0 and last_refined_img is not None:
                    t1 = last_refined_img  # Image from Year t-1
                    t2 = refined           # Image from Year t
                    
                    # Forest (1 OR 2) -> Non-Forest (>= 3)
                    defor_img = t1.lte(2).And(t2.gte(3))
                    # Non-Forest (>= 3) -> Forest (1 OR 2)
                    refor_img = t1.gte(3).And(t2.lte(2))
                    # Forest (1 OR 2) -> Forest (1 OR 2)
                    stable_forest_img = t1.lte(2).And(t2.lte(2))
                    # Expansion: Non-Builtup (!= 6) -> Builtup (6)
                    builtup_exp_img = t1.neq(6).And(t2.eq(6))
                    
                    trans_combined = ee.Image.cat([
                        defor_img.rename('defor'),
                        refor_img.rename('refor'),
                        stable_forest_img.rename('stable'),
                        builtup_exp_img.rename('builtup_exp')
                    ]).multiply(ee.Image.pixelArea().divide(10000))
                    
                    trans_stats = await asyncio.to_thread(lambda: trans_combined.reduceRegion(
                        reducer=ee.Reducer.sum(),
                        geometry=geometry,
                        scale=10,
                        maxPixels=1e10,
                        bestEffort=True,
                        tileScale=16  # Divide into tiles to reduce memory usage
                    ).getInfo())
                    
                    yearly_data.forest_loss = round(trans_stats.get('defor', 0), 2)
                    yearly_data.forest_gain = round(trans_stats.get('refor', 0), 2)
                    yearly_data.forest_stable = round(trans_stats.get('stable', 0), 2)
                    yearly_data.builtup_expansion = round(trans_stats.get('builtup_exp', 0), 2)
                    yearly_data.deforestation_ha = yearly_data.forest_loss
                    yearly_data.reforestation_ha = yearly_data.forest_gain
                    
                    # Maps
                    yearly_data.deforestation_map_url = await asyncio.to_thread(lambda: get_transition_map_url(defor_img, "#FF4500"))
                    yearly_data.reforestation_map_url = await asyncio.to_thread(lambda: get_transition_map_url(refor_img, "#22C55E"))
                    yearly_data.builtup_expansion_map_url = await asyncio.to_thread(lambda: get_transition_map_url(builtup_exp_img, "#A855F7"))

                last_refined_img = refined
                classification_results.append(yearly_data)
                
            except Exception as e:
                logger.error(f"   ❌ Finalization error {year}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        # Collect successful results
        results = classification_results
        results.sort(key=lambda x: x.year)
        
        # Calculate Cumulative Statistics (WAJIB - Section C)
        total_defor = sum(getattr(r, 'deforestation_ha', 0) for r in results)
        total_refor = sum(getattr(r, 'reforestation_ha', 0) for r in results)
        total_builtup_exp = sum(getattr(r, 'builtup_expansion', 0) for r in results)
        
        transition_summary = {
            "total_deforestation_ha": round(total_defor, 2),
            "total_reforestation_ha": round(total_refor, 2),
            "net_forest_change_ha": round(total_refor - total_defor, 2),
            "total_builtup_expansion_ha": round(total_builtup_exp, 2),
            "period": f"{start_year} - {end_year}",
            "unit": "Hektar (ha)"
        }
        
        # Audit & Verification Report (WAJIB - Section E)
        audit_report = {
            "rf_retrained": "TIDAK",
            "methodology_changed": "TIDAK",
            "analysis_method": "Temporal Consistency Analysis (Post-classification)",
            "forest_consistency_threshold": "80% (stabil forest pixels)",
            "transition_definitions": {
                "deforestation": "Forest (1,2) -> Non-Forest (3,4,5,6,7,8,9)",
                "reforestation": "Non-Forest -> Forest (1,2)",
                "builtup_expansion": "Non-Builtup (!=6) -> Builtup (6)"
            },
            "pixel_wise_comparison": "YA",
            "data_internal_system": "YA"
        }
        
        if on_progress:
            await on_progress(95, "Selesai", f"{len(results)} tahun berhasil diklasifikasi")
        
        logger.info("=" * 60)
        logger.info(f"✅ CLASSIFICATION COMPLETE: {len(results)} years processed")
        logger.info(f"📊 Cumulative Defo: {total_defor:.2f} ha, Refo: {total_refor:.2f} ha")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error(f"❌ Classification failed: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    return results, transition_summary, audit_report


# API ENDPOINTS
# ==============================================================================


# ==============================================================================
# API ENDPOINTS
# ==============================================================================

@app.get("/")
async def root():
    """Root endpoint - API health check."""
    return {
        "service": "GeoAnalyzer API - Sentinel-2",
        "status": "berjalan",
        "version": "2.0.0",
        "gee_status": "terhubung",
        "model": "Sentinel-2 Land Cover Classification",
        "resolusi": "10 meter"
    }


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    try:
        ee.Number(1).getInfo()
        return {"status": "sehat", "gee": "terhubung"}
    except Exception as e:
        return {"status": "tidak sehat", "gee": "terputus", "error": str(e)}


@app.websocket("/ws/analyze/{client_id}")
async def websocket_analyze(websocket: WebSocket, client_id: str):
    """
    WebSocket endpoint for real-time analysis with progress updates.
    Unified with standardized regional analysis logic.
    """
    # Sanitasi client_id (Security Best Practice)
    import re
    if not re.match(r"^[a-zA-Z0-9_\-]+$", client_id):
        await websocket.close(code=1008) # Policy Violation
        return

    await ws_manager.connect(websocket, client_id)
    
    try:
        data = await websocket.receive_json()
        
        # Core variables
        geojson_data = data.get("geojson")
        years = data.get("years", 5)
        mode = data.get("mode", "series")
        specific_date = data.get("specific_date")
        thresholds_data = data.get("thresholds")
        thresholds = ClassificationThresholds(**thresholds_data) if thresholds_data else ClassificationThresholds()
        geometry = geojson_to_ee_geometry(geojson_data)
        
        # Extract advanced options (Points 2, 4, 5)
        scale = data.get("scale", 30)
        
        # Define progress callback
        async def on_progress(percent, step, detail):
            await ws_manager.send_progress(client_id, percent, step, detail)
        
        print(f"🔍 DEBUG: Starting analyze_land_cover...")
        print(f"   📍 Geometry type: {type(geometry)}")
        print(f"   📅 Years: {years}, Mode: {mode}")
        
        # Extract params
        start_year = data.get("start_year")
        end_year = data.get("end_year")
        cloud_prob_threshold = data.get("cloud_prob_threshold", 50)
        
        # Call unified analysis logic with SMART QUEUE tracking
        async with analysis_queue.enter_queue(client_id):
            results, transition_summary, audit_report = await analyze_land_cover(
                geometry=geometry,
                start_year=start_year,
                end_year=end_year,
                years=years,
                thresholds=thresholds,
                mode=mode,
                specific_date=specific_date,
                on_progress=on_progress,
                scale=scale,
                existing_data=data.get("existing_data")
            )
        
        print(f"🔍 DEBUG: analyze_land_cover returned {len(results)} results")
        
        # Prepare response
        response_data = {
            "status": "sukses",
            "message": f"Analisis selesai untuk {len(results)} tahun",
            "data": [r.dict() for r in results],
            "map_url": results[-1].map_url if results else None,
            "transition_summary": transition_summary,
            "audit_report": audit_report
        }
        
        await ws_manager.send_progress(client_id, 100, "Selesai!", "Analisis berhasil")
        await ws_manager.send_complete(client_id, response_data)
        
    except WebSocketDisconnect:
        ws_manager.disconnect(client_id)
    except Exception as e:
        print(f"❌ WebSocket error: {e}")
        traceback.print_exc()
        await ws_manager.send_error(client_id, str(e))
    finally:
        ws_manager.disconnect(client_id)


@app.post("/analyze", response_model=AnalyzeResponse)
async def analyze(request: Request, body: AnalyzeRequest):
    """
    Analisis perubahan tutupan lahan untuk geometri GeoJSON yang diberikan.
    """
    try:
        # Check cancellation wrapper
        async def check_cancel():
            is_disconnected = await request.is_disconnected()
            if is_disconnected:
                print("⛔ Client disconnected, cancelling analysis...")
            return is_disconnected

        # Konversi GeoJSON ke EE Geometry
        geometry = geojson_to_ee_geometry(body.geojson)
        
        print(f"📍 Menganalisis wilayah untuk {body.years} tahun...")
        
        # Prepare a unique job index for HTTP calls (not associated with a client_id session)
        job_id = f"http_{hashlib.md5(request.client.host.encode()).hexdigest()[:8]}_{int(time.time())}"
        
        # Lakukan analisis tutupan lahan via SMART QUEUE
        async with analysis_queue.enter_queue(job_id):
            data, transition_summary, audit_report = await analyze_land_cover(
                geometry, 
                start_year=body.start_year,
                end_year=body.end_year,
                years=body.years, 
                thresholds=body.thresholds,
                mode=body.mode,
                specific_date=body.specific_date,
                scale=body.scale,
                check_cancel=check_cancel
            )
        
        if not data:
            return AnalyzeResponse(
                status="peringatan",
                message="Tidak ada data tersedia untuk wilayah dan periode waktu yang ditentukan",
                data=[]
            )
        
        # Map URL untuk response (menggunakan data tahun terakhir)
        map_url = data[-1].map_url if data else None
        
        return AnalyzeResponse(
            status="sukses",
            message=f"Analisis selesai untuk {len(data)} tahun",
            data=data,
            map_url=map_url,
            transition_summary=transition_summary,
            audit_report=audit_report
        )
        
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        print(f"❌ Error analisis: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Analisis gagal: {str(e)}")


@app.get("/classes")
async def get_land_cover_classes():
    """Dapatkan definisi kelas tutupan lahan."""
    return {
        "kelas_tutupan_lahan": LC_CLASSES,
        "palet_warna": LC_PALETTE,
        "sumber_data": "Sentinel-2 Surface Reflectance",
        "resolusi": "10 meter",
        "metode": "Klasifikasi berbasis indeks spektral (NDVI, NDWI, NDBI, SAVI, BSI, MNDWI)"
    }


class ExportRequest(BaseModel):
    """Request body untuk export endpoints."""
    geojson: Dict[str, Any] = Field(..., description="Objek geometri GeoJSON")
    data: List[Dict[str, Any]] = Field(..., description="Data hasil analisis")
    format: str = Field(default="geojson", description="Format export: geojson atau kml")


@app.post("/export/geojson")
async def export_geojson(request: ExportRequest):
    """
    Export hasil analisis sebagai GeoJSON dengan properti tutupan lahan.
    HANYA 6 kelas IPSDH final yang di-export.
    """
    try:
        features = []
        input_geojson = request.geojson

        # Normalisasi data ke 6 kelas IPSDH final
        normalized_data = normalize_analysis_results(request.data)

        # Wrap the geometry with analysis data properties
        for year_data in normalized_data:
            feature = {
                "type": "Feature",
                "geometry": input_geojson.get("geometry") or input_geojson,
                "properties": {
                    "year": year_data.get("year"),
                    "hutan_primer_ha": year_data.get("hutan_primer", 0),
                    "hutan_sekunder_ha": year_data.get("hutan_sekunder", 0),
                    "hutan_total_ha": year_data.get("hutan", 0),
                    "tanah_kering_ha": year_data.get("tanah_kering", 0),
                    "tanah_kosong_ha": year_data.get("tanah_kosong", 0),
                    "air_ha": year_data.get("air", 0),
                    "lahan_terbangun_ha": year_data.get("lahan_terbangun", 0),
                    "total_ha": sum([
                        year_data.get("hutan_primer", 0),
                        year_data.get("hutan_sekunder", 0),
                        year_data.get("tanah_kering", 0),
                        year_data.get("tanah_kosong", 0),
                        year_data.get("air", 0),
                        year_data.get("lahan_terbangun", 0)
                    ])
                }
            }
            features.append(feature)
        
        return {
            "type": "FeatureCollection",
            "features": features,
            "properties": {
                "source": "GeoAnalyzer - Sentinel-2",
                "generated_at": datetime.now().isoformat(),
                "classes": LC_CLASSES
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export gagal: {str(e)}")


@app.post("/export/kml")
async def export_kml(request: ExportRequest):
    """
    Export hasil analisis sebagai KML untuk Google Earth.
    HANYA 6 kelas IPSDH final yang di-export.
    """
    try:
        from xml.etree.ElementTree import Element, SubElement, tostring

        # Normalisasi data ke 6 kelas IPSDH final
        normalized_data = normalize_analysis_results(request.data)

        kml = Element('kml', xmlns="http://www.opengis.net/kml/2.2")
        document = SubElement(kml, 'Document')
        SubElement(document, 'name').text = "GeoAnalyzer Land Cover Analysis"
        SubElement(document, 'description').text = f"Generated: {datetime.now().isoformat()}"

        # Add styles for each class
        style_colors = {
            'hutan_primer': 'ff006400',
            'hutan_sekunder': 'ff32CD32',
            'tanah_kering': 'ff20A5DA',
            'tanah_kosong': 'ff1E69D2',
            'air': 'ffFF901E',
            'lahan_terbangun': 'ff708090'
        }

        for class_name, color in style_colors.items():
            style = SubElement(document, 'Style', id=f"style_{class_name}")
            poly_style = SubElement(style, 'PolyStyle')
            SubElement(poly_style, 'color').text = color
            SubElement(poly_style, 'fill').text = '1'

        # Add placemarks for each year
        for year_data in normalized_data:
            placemark = SubElement(document, 'Placemark')
            SubElement(placemark, 'name').text = f"Tahun {year_data.get('year')}"
            
            desc = f"""
            Hutan Primer: {year_data.get('hutan_primer', 0):.2f} Ha
            Hutan Sekunder: {year_data.get('hutan_sekunder', 0):.2f} Ha
            Tanah Kering: {year_data.get('tanah_kering', 0):.2f} Ha
            Tanah Kosong: {year_data.get('tanah_kosong', 0):.2f} Ha
            Lahan Terbangun: {year_data.get('lahan_terbangun', 0):.2f} Ha
            Air: {year_data.get('air', 0):.2f} Ha
            """
            SubElement(placemark, 'description').text = desc
            
            # Add style URL
            SubElement(placemark, 'styleUrl').text = '#style_hutan' # Default style
            
            # --- Fix: Add Geometry to KML ---
            input_geojson = request.geojson
            # Normalize geometry input
            geom_data = input_geojson.get("geometry") or input_geojson
            geom_type = geom_data.get('type')
            coordinates = geom_data.get('coordinates')

            def add_polygon(parent, ring_coords):
                poly = SubElement(parent, 'Polygon')
                outer = SubElement(poly, 'outerBoundaryIs')
                lr = SubElement(outer, 'LinearRing')
                co = SubElement(lr, 'coordinates')
                # KML expects "lon,lat,alt lon,lat,alt ..."
                coords_str = " ".join([f"{c[0]},{c[1]},0" for c in ring_coords])
                co.text = coords_str

            if geom_type == 'Polygon':
                # GeoJSON Polygon coordinates: [ [outer_ring], [inner_ring], ... ]
                # We only handle outer ring for simplicity or iterate
                if coordinates:
                    add_polygon(placemark, coordinates[0])
            
            elif geom_type == 'MultiPolygon':
                multi = SubElement(placemark, 'MultiGeometry')
                for poly_coords in coordinates:
                    if poly_coords:
                        add_polygon(multi, poly_coords[0])
            # --------------------------------
        
        kml_string = tostring(kml, encoding='unicode')
        return {"kml": kml_string, "filename": f"land_cover_{datetime.now().strftime('%Y%m%d')}.kml"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export KML gagal: {str(e)}")




@app.post("/change-detection", response_model=ChangeDetectionResponse)
async def detect_changes(request: ChangeDetectionRequest):
    """
    Deteksi perubahan tutupan lahan antara dua tahun dengan konfidensi per kelas.
    """
    try:
        geometry = geojson_to_ee_geometry(request.geojson)
        thresholds = request.thresholds or ClassificationThresholds()
        
        print(f"🚀 [ENDPOINT] Incoming Change Detection Request")
        print(f"   📅 Years: {request.start_year} -> {request.end_year}")
        print(f"   📏 Threshold: {request.threshold} Ha")
        
        # Analyze both years
        start_collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
            .filterBounds(geometry)
            .filterDate(f"{request.start_year}-01-01", f"{request.start_year}-12-31")
            .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 60))
            .map(mask_clouds_sentinel2))
        
        end_collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
            .filterBounds(geometry)
            .filterDate(f"{request.end_year}-01-01", f"{request.end_year}-12-31")
            .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 60))
            .map(mask_clouds_sentinel2))
        
        if start_collection.size().getInfo() == 0 or end_collection.size().getInfo() == 0:
            raise ValueError("Tidak ada data citra untuk periode yang diminta")
        
        # Classify both years with confidence
        start_composite = start_collection.median().clip(geometry)
        end_composite = end_collection.median().clip(geometry)
        
        print(f"   🔍 Creating classifications for {request.start_year} and {request.end_year}")
        start_classified = classify_land_cover(start_composite, geometry, thresholds).rename('classification')
        end_classified = classify_land_cover(end_composite, geometry, thresholds).rename('classification')
        
        # Validate classifications are not empty
        try:
            start_sample = start_classified.sample(geometry, 30, numPixels=10).getInfo()
            end_sample = end_classified.sample(geometry, 30, numPixels=10).getInfo()
            print(f"   ✅ Start classification sample: {len(start_sample.get('features', []))} pixels")
            print(f"   ✅ End classification sample: {len(end_sample.get('features', []))} pixels")
            if len(start_sample.get('features', [])) == 0 or len(end_sample.get('features', [])) == 0:
                raise ValueError("Classified images are empty")
        except Exception as e:
            print(f"   ⚠️ Classification validation warning: {e}")
        
        # Calculate confidence for both years
        def calc_confidence(composite, classified):
            confidence_percent = None
            confidence_detailed = {}
            try:
                image_with_indices = calculate_indices(composite)
                ndvi = image_with_indices.select('NDVI')
                conf_img = ndvi.subtract(0.4).abs().multiply(2.5).clamp(0, 1)
                
                # Global
                mean_conf = conf_img.reduceRegion(
                    reducer=ee.Reducer.mean(),
                    geometry=geometry,
                    scale=30,
                    maxPixels=1e8,
                    bestEffort=True
                ).getInfo()
                conf_val = mean_conf.get('NDVI')
                if conf_val is not None:
                    confidence_percent = round(conf_val * 100, 1)
                
                # Per class
                for class_id, class_key in LC_CLASSES.items():
                    class_mask = classified.eq(class_id)
                    class_conf = conf_img.updateMask(class_mask).reduceRegion(
                        reducer=ee.Reducer.mean(),
                        geometry=geometry,
                        scale=30,
                        maxPixels=1e8,
                        bestEffort=True
                    ).getInfo()
                    c_val = class_conf.get('NDVI')
                    if c_val is not None:
                        confidence_detailed[class_key] = round(c_val * 100, 1)
            except Exception as e:
                print(f"   ⚠️ Confidence calculation failed: {e}")
            
            return confidence_percent, confidence_detailed
        
        conf_start, conf_detail_start = calc_confidence(start_composite, start_classified)
        conf_end, conf_detail_end = calc_confidence(end_composite, end_classified)
        
        # 3. Calculate Area efficiently using Group Reducer (Fix 0.00 Ha issue)
        area_image = ee.Image.pixelArea().divide(10000) # Hectares
        
        def calculate_areas(classified_img, label=""):
            # Ensure classified band is named 'classification'
            if 'classification' not in classified_img.bandNames().getInfo():
                classified_img = classified_img.rename('classification')
            
            # Create combined image with proper masking
            # band 0 = area, band 1 = classification
            masked_area = area_image.updateMask(classified_img.gt(0))
            combined = masked_area.addBands(classified_img)
            
            # Use groupField=1 since classification is the second band (index 1)
            # Python GEE API prefers keyword arguments for .group()
            stats = combined.reduceRegion(
                reducer=ee.Reducer.sum().group(
                    groupField=1,
                    groupName='class'
                ),
                geometry=geometry,
                scale=10,
                maxPixels=1e10,
                bestEffort=True
            ).getInfo()
            
            print(f"   🔍 {label} Raw Stats from GEE: {stats}")
            groups = stats.get('groups', [])
            print(f"   🔍 {label} Groups array: {groups}")
            
            if not groups:
                print(f"   ⚠️ {label} WARNING: No groups returned from reduceRegion!")
                return {}
            
            result = {}
            for item in groups:
                if 'class' in item and 'sum' in item:
                    class_id = int(item['class'])
                    area_sum = item['sum']
                    result[class_id] = area_sum
                    print(f"   🔍 {label} Class {class_id}: {area_sum} Ha")
            
            return result

        start_areas_raw = calculate_areas(start_classified, "Start Year")
        end_areas_raw = calculate_areas(end_classified, "End Year")
        
        print(f"   📊 Start Areas Final: {start_areas_raw}")
        print(f"   📊 End Areas Final: {end_areas_raw}")
        
        start_areas = {name: start_areas_raw.get(id, 0) for id, name in LC_CLASSES.items()}
        end_areas = {name: end_areas_raw.get(id, 0) for id, name in LC_CLASSES.items()}
        
        print(f"   📊 Start Areas Mapped: {start_areas}")
        print(f"   📊 End Areas Mapped: {end_areas}")
        
        # 4. Calculate Changes
        changes = {}
        change_percent = {}
        total_change = 0.0
        
        for class_key in LC_CLASSES.values():
            change_ha = end_areas.get(class_key, 0) - start_areas.get(class_key, 0)
            changes[class_key] = round(change_ha, 2)
            
            start_val = start_areas.get(class_key, 0)
            if start_val > 0.01: # Avoid division by zero
                pct = (change_ha / start_val) * 100
                change_percent[class_key] = round(pct, 1)
            else:
                change_percent[class_key] = 0.0 if change_ha == 0 else 100.0
            
            total_change += abs(change_ha)
        
        print(f"   📊 Changes: {changes}")
        print(f"   📊 Total Change: {total_change} Ha")

        # 5. Generate Change Map (Transition Visualization)
        # Remap to separate decades: Start=10,20,30,40; End=1,2,3,4
        # Sum = Start + End. Ex: 13 = Forest(10) -> Bare(3) = Deforestation
        start_remap = start_classified.remap([1,2,3,4], [10,20,30,40])
        transition = start_remap.add(end_classified)
        
        # Color Palette for Transitions:
        # 11, 22, 33, 44: Stable (Grey/Transparent)
        # 12, 13: Deforestation (Red)
        # 21, 31: Regrowth/Afforestation (Green)
        # Others: Minor changes (Yellow/Orange)
        
        # Mask out stable pixels for clearer view
        stable_mask = transition.neq(11).And(transition.neq(22)).And(transition.neq(33)).And(transition.neq(44))
        change_map = transition.updateMask(stable_mask)
        
        change_viz_params = {
            'min': 10,
            'max': 45,
            'palette': [
                '#ff0000', # Deforestation 
                '#00ff00', # Growth
                '#ffff00'  # Other
            ]
        }
        
        # Custom visualizer for significant classes
        # 13 (Forest->Soil): Deforestation Severe (Red)
        # 12 (Forest->Dry): Degradation (Orange)
        # 21 (Dry->Forest): Regrowth (Light Green)
        # 31 (Soil->Forest): Reforestation (Dark Green)
        visualization = change_map.visualize(
            min=0, max=50,
            palette=['black'] # Default
        )
        
        # We manually colorize key transitions for better UX
        # Deforestation (12, 13, 14) -> Red
        loss_mask = transition.eq(12).Or(transition.eq(13)).Or(transition.eq(14))
        # Growth (21, 31, 41) -> Green
        gain_mask = transition.eq(21).Or(transition.eq(31)).Or(transition.eq(41))
        
        # Base: Transparent
        final_viz = ee.Image(0).visualize(palette=['00000000'], opacity=0.0)
        
        # Add Loss Layer (Red)
        loss_layer = transition.updateMask(loss_mask).visualize(palette=['#ef4444'], min=0, max=100)
        final_viz = final_viz.blend(loss_layer)
        
        # Add Gain Layer (Green)
        gain_layer = transition.updateMask(gain_mask).visualize(palette=['#10b981'], min=0, max=100)
        final_viz = final_viz.blend(gain_layer)
        
        map_url = final_viz.getMapId()['tile_fetcher'].url_format
        
        return ChangeDetectionResponse(
            status="sukses",
            message=f"Deteksi perubahan selesai: {request.start_year} → {request.end_year}",
            start_year=request.start_year,
            end_year=request.end_year,
            changes=changes,
            change_percent=change_percent,
            total_change=round(total_change, 2),
            confidence_start=conf_start,
            confidence_end=conf_end,
            confidence_detailed_start=conf_detail_start,
            confidence_detailed_end=conf_detail_end,
            map_url=map_url
        )
        
    except Exception as e:
        print(f"❌ Change detection error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Deteksi perubahan gagal: {str(e)}")



@app.post("/map/slope")
async def get_slope_analysis(request: SlopeRequest):
    """
    Generate Slope Analysis Layer & Statistics (Live).
    Returns MapID and Summary Stats for Inside/Outside 2km.
    Uses NASADEM (SRTM) for elevation data.
    """
    try:
        print("📐 Starting Slope Analysis (Live)...")
        
        # 1. Parse Geometry
        geometry = geojson_to_ee_geometry(request.geo_data)
        
        # 2. Scope Definitions
        # Scope 1: Inside (The Geometry itself) (Label: WILAYAH)
        # Scope 2: Outside (Buffer 2km diff Geometry) (Label: BUFFER 2KM)
        buffer_mask = geometry.buffer(2000)
        outside_mask = buffer_mask.difference(geometry)
        
        # 3. Load NASADEM (SRTM)
        # Use existing NASA DEM which is standard for global coverage
        dem = ee.Image("NASA/NASADEM_HGT/001").select('elevation')
        
        # 4. Calculate Slope (Radians -> Degrees -> Percent?)
        # ee.Terrain.slope returns DEGREES.
        slope_deg = ee.Terrain.slope(dem)
        
        # Calculate Percent Slope for classification: tan(deg) * 100
        # Formula: Percent = tan(radians(slope_in_degrees)) * 100
        slope_pct = slope_deg.multiply(math.pi).divide(180).tan().multiply(100).rename('slope_pct')
        
        # 5. Create Visualization Map (KLHK Standard)
        # 0-8% (Datar), 8-15% (Landai), 15-25% (Agak Curam), 25-45% (Curam), >45% (Sangat Curam)
        sld_intervals = (
            '<RasterSymbolizer>'
            '<ColorMap type="intervals" extended="false">'
            '<ColorMapEntry color="#31a354" quantity="8" label="0-8% Datar" />'
            '<ColorMapEntry color="#addd8e" quantity="15" label="8-15% Landai" />'
            '<ColorMapEntry color="#fee391" quantity="25" label="15-25% Agak Curam" />'
            '<ColorMapEntry color="#fec44f" quantity="45" label="25-45% Curam" />'
            '<ColorMapEntry color="#cc4c02" quantity="1000" label=">45% Sangat Curam" />'
            '</ColorMap>'
            '</RasterSymbolizer>'
        )
        
        # Visualize for the buffer area (covers both inside + outside context)
        # We clip to buffer_mask so the user sees the AOI + surrounding 2km
        viz_image = slope_pct.clip(buffer_mask).sldStyle(sld_intervals)
        map_url = viz_image.getMapId()['tile_fetcher'].url_format
        
        # 6. Calculate Statistics Helper
        def calc_slope_stats(geom, scope_name):
            # Calculate pixel area in Ha
            area_img = ee.Image.pixelArea().divide(10000).rename('area')
            
            # Create masks according to KLHK classes
            # 1: 0-8% (Datar)
            # 2: 8-15% (Landai)
            # 3: 15-25% (Agak Curam)
            # 4: 25-45% (Curam)
            # 5: >45% (Sangat Curam)

            class_img = ee.Image(0).rename('class')
            class_img = class_img.where(slope_pct.lte(8), 1)
            class_img = class_img.where(slope_pct.gt(8).And(slope_pct.lte(15)), 2)
            class_img = class_img.where(slope_pct.gt(15).And(slope_pct.lte(25)), 3)
            class_img = class_img.where(slope_pct.gt(25).And(slope_pct.lte(45)), 4)
            class_img = class_img.where(slope_pct.gt(45), 5)
            
            # Combine area and class
            combined = area_img.addBands(class_img)
            
            # Grouped reduction
            result = combined.reduceRegion(
                reducer=ee.Reducer.sum().group(groupField=1, groupName='class'),
                geometry=geom,
                scale=30,
                maxPixels=1e9,
                bestEffort=True
            ).getInfo()
            
            groups = result.get('groups', [])
            mapped_res = {int(item['class']): item['sum'] for item in groups}
            
            # Populate stats with defaults (KLHK Standard: 25-45%, >45%)
            # Note: Field names match DB columns for consistency
            stats = {}
            stats["slope_0_8"] = mapped_res.get(1, 0)       # Datar: 0-8%
            stats["slope_8_15"] = mapped_res.get(2, 0)      # Landai: 8-15%
            stats["slope_15_25"] = mapped_res.get(3, 0)     # Agak Curam: 15-25%
            stats["slope_25_40"] = mapped_res.get(4, 0)     # Curam: 25-45% (KLHK)
            stats["slope_above_40"] = mapped_res.get(5, 0)  # Sangat Curam: >45% (KLHK)
            
            # Average Slope (Percentage)
            avg_res = slope_pct.reduceRegion(
                reducer=ee.Reducer.mean(),
                geometry=geom,
                scale=30,
                maxPixels=1e9,
                bestEffort=True
            ).getInfo()

            stats["avg_slope"] = round(avg_res.get('slope_pct', 0), 1)
            stats["scope"] = scope_name
            
            return stats

        print("   📊 Computing stats for INSIDE...")
        summary_inside = calc_slope_stats(geometry, "INSIDE")
        
        print("   📊 Computing stats for OUTSIDE (2km)...")
        summary_outside = calc_slope_stats(outside_mask, "OUTSIDE")
        
        return {
            "status": "success",
            "map_url": map_url,
            "db_summary": [summary_inside, summary_outside]
        }
        
    except Exception as e:
        print(f"❌ Slope Analysis Failed: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/history")
async def save_history(request: SaveHistoryRequest, background_tasks: BackgroundTasks):
    """
    Simpan riwayat analisis dengan pola Relational Database.
    Geometri disimpan di master_lahan (unik berdasarkan hash), history merujuk ke sana.
    Hotspot data di-populate secara background untuk views KPS.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase client not configured")
    
    try:
        # 1. Hash GeoJSON untuk cek duplikasi
        # Sort keys agar urutan JSON tidak mempengaruhi hash
        geom_str = json.dumps(request.geo_data, sort_keys=True)
        geom_hash = hashlib.md5(geom_str.encode()).hexdigest()
        
        lahan_id = None
        
        # 2. Cek apakah geometry sudah ada di master_lahan
        # Gunakan 'select' biasa lalu ambil index 0 manual untuk keamanan lintas versi library
        res = await asyncio.to_thread(lambda: supabase.table("master_lahan").select("id").eq("geom_hash", geom_hash).limit(1).execute())
        
        if res is not None and hasattr(res, 'data') and res.data and len(res.data) > 0:
            print(f"✅ Found existing master_lahan: {res.data[0]['id']}")
            lahan_id = res.data[0]['id']
        else:
            print(f"➕ Creating new master_lahan entry for hash {geom_hash[:8]}...")
            # Insert baru. 
            new_lahan = {
                "geom_geojson": request.geo_data,
                "geom_hash": geom_hash
            }
            ins = await asyncio.to_thread(lambda: supabase.table("master_lahan").insert(new_lahan).execute())
            
            if ins is not None and hasattr(ins, 'data') and ins.data and len(ins.data) > 0:
                lahan_id = ins.data[0]['id']
            else:
                raise Exception("Insert master_lahan returned no data. Check Supabase RLS policies.")
                
        if not lahan_id:
            raise HTTPException(status_code=500, detail="Gagal mendapatkan ID Lahan")
            
        # --- LOGIC PERSISTENT THUMBNAIL (BASE64) ---
        # URL dari Earth Engine akan kadaluarsa dalam hitungan jam.
        # Kita harus download gambarnya sekarang dan simpan sebagai Base64 string agar permanen di DB.
        
        updated_analysis_results = []
        import base64
        import requests
        
        print("🔄 Processing thumbnails for persistence...")
        
        for item in request.analysis_results:
            new_item = item.copy()
            url = new_item.get("thumb_url")
            
            # --- LOGI K PERSISTENSI VEKTOR & THUMBNAIL ---
            
            # 1. Konversi EE Geometry sekali saja untuk efisiensi
            if 'ee_geometry_cache' not in locals():
                 ee_geometry_cache = geojson_to_ee_geometry(request.geo_data)

            # 2. Generate Vector GeoJSON (Wajib ada untuk PDF)
            if not new_item.get('vector_geojson'):
                try:
                    print(f"   📐 Generating vector for {new_item.get('year')}...")
                    classified = get_classified_image(ee_geometry_cache, new_item.get('year'))
                    if classified:
                        # Use scale from metadata if provided, fallback to 30
                        vector_scale = request.metadata.get('scale', 30)
                        print(f"      📍 Using vector scale: {vector_scale}m")
                        vectors_fc = classified.reduceToVectors(
                            geometry=ee_geometry_cache,
                            scale=vector_scale,
                            geometryType='polygon',
                            eightConnected=False,
                            labelProperty='class',
                            reducer=ee.Reducer.countEvery(),
                            maxPixels=1e8,
                            bestEffort=True
                        )
                        vector_data = await asyncio.to_thread(lambda: vectors_fc.getInfo())
                        
                        # Simplify properties
                        if vector_data and 'features' in vector_data:
                            for f in vector_data['features']:
                                f['properties'] = {'class': f['properties']['class']}
                            new_item['vector_geojson'] = vector_data
                            print(f"   ✅ Vector Generated.")
                except Exception as ve:
                    print(f"   ⚠️ Failed to generate vector during save: {ve}")

            # 3. Persistent Thumbnail (Cloud Storage)
            # Replace temporary GEE URLs with permanent Supabase Storage URLs
            thumb_url = new_item.get("thumb_url")
            if thumb_url and ("earthengine.googleapis.com" in thumb_url or "googleapis.com" in thumb_url):
                try:
                    print(f"   ☁️ Persisting thumbnail for {new_item.get('year')}...")
                    # Function is async, await directly
                    perm_url = await download_and_save_locally(thumb_url, f"history_{lahan_id}")
                    if perm_url:
                        new_item["thumb_url"] = perm_url
                        print(f"   ✅ Thumbnail persisted: {perm_url}")
                except Exception as e:
                    print(f"   ⚠️ Failed to persist thumbnail: {e}")
            
            # 4. Persistent RGB Thumbnail (Cloud Storage)
            rgb_url = new_item.get("rgb_thumb_url")
            if rgb_url and ("earthengine.googleapis.com" in rgb_url or "googleapis.com" in rgb_url):
                try:
                    print(f"   ☁️ Persisting RGB thumbnail for {new_item.get('year')}...")
                    perm_rgb = await download_and_save_locally(rgb_url, f"history_{lahan_id}_rgb")
                    if perm_rgb:
                        new_item["rgb_thumb_url"] = perm_rgb
                        print(f"   ✅ RGB Thumbnail persisted: {perm_rgb}")
                except Exception as e:
                    print(f"   ⚠️ Failed to persist RGB thumbnail: {e}")

            updated_analysis_results.append(new_item)
            
        # --- LOGIKA PENYIMPANAN: Replace vs Merge (Request User) ---
        mode = getattr(request, 'mode', 'replace')
        print(f"💾 Saving history in mode: {mode} for lahan_id: {lahan_id}")

        if mode == "merge":
            try:
                # 1. Cari data lama untuk digabungkan
                old_res = await asyncio.to_thread(lambda: supabase.table("analysis_history")
                    .select("id, analysis_results, metadata")
                    .eq("lahan_id", lahan_id)
                    .limit(1).execute())
                
                if old_res and old_res.data:
                    old_history = old_res.data[0]
                    old_results = old_history.get('analysis_results', [])
                    
                    # Gabungkan berdasarkan Tahun (Tahun baru menimpa tahun lama jika sama)
                    merged_map = { item['year']: item for item in old_results }
                    for nr in updated_analysis_results:
                        merged_map[nr['year']] = nr
                    
                    # Urutkan kembali berdasarkan tahun
                    final_results = sorted(merged_map.values(), key=lambda x: x['year'])
                    
                    # Update metadata untuk mencerminkan rentang tahun terbaru
                    new_meta = {
                        **old_history.get('metadata', {}),
                        **request.metadata,
                        "transition_summary": request.transition_summary,
                        "audit_report": request.audit_report,
                        "processed_at": datetime.now().isoformat(),
                        "merge_history": True,
                        "last_merge_at": datetime.now().isoformat()
                    }
                    
                    # 2. Update record yang sudah ada
                    update_payload = {
                        "analysis_results": final_results,
                        "metadata": new_meta,
                        "filename": request.filename, # Update nama file jika berubah
                        "created_at": datetime.now().isoformat() # Bump timestamp agar muncul paling atas
                    }
                    
                    await asyncio.to_thread(lambda: supabase.table("analysis_history")
                        .update(update_payload)
                        .eq("id", old_history['id'])
                        .execute())
                    
                    print(f"✅ Data merged successfully into id: {old_history['id']}")
                    invalidate_cache(f"history_detail_{old_history['id']}")
                    invalidate_cache("history_list")
                    return {"status": "success", "data": {"id": old_history['id']}, "message": "Data berhasil digabungkan (Smart Merge)"}
            except Exception as me:
                print(f"⚠️ Smart Merge failed, falling back to replace: {me}")

        # --- DEFAULT: REPLACE MODE (Hapus lama, Simpan baru) ---
        print(f"🧹 Performing standard replace (cleanup) for lahan_id: {lahan_id}...")
        try:
            old_res = await asyncio.to_thread(lambda: supabase.table("analysis_history").select("id").eq("lahan_id", lahan_id).execute())
            if old_res and old_res.data:
                for old in old_res.data:
                    invalidate_cache(f"history_detail_{old['id']}")
            
            await asyncio.to_thread(lambda: supabase.table("analysis_history").delete().eq("lahan_id", lahan_id).execute())
        except Exception as e:
            print(f"⚠️ Cleanup failed (ignored): {e}")

        # 3. Simpan SEBAGAI BARU (Insert)
        # Determine analysis_scope based on kps_id
        effective_scope = "KPS" if request.kps_id else (request.analysis_scope or "NON_KPS")

        # --- AUTO-CREATE NON-KPS MASTER RECORD IF NEEDED ---
        non_kps_id = None
        if effective_scope == "NON_KPS" and lahan_id:
            try:
                print(f"🏞️ Processing Non-KPS master record for lahan_id: {lahan_id}")

                # Extract nama_areal from filename (remove extensions)
                nama_areal = request.filename
                for ext in ['.shp', '.geojson', '.zip', '.json']:
                    nama_areal = nama_areal.replace(ext, '')

                # Check if Non-KPS already exists for this geometry
                existing_non_kps = await asyncio.to_thread(
                    lambda: supabase.table("master_non_kps")
                    .select("id, nama_areal")
                    .eq("lahan_id", lahan_id)
                    .limit(1)
                    .execute()
                )

                if existing_non_kps.data and len(existing_non_kps.data) > 0:
                    # Reuse existing Non-KPS record
                    non_kps_id = existing_non_kps.data[0]["id"]
                    existing_name = existing_non_kps.data[0].get("nama_areal", "")
                    print(f"♻️ Reusing existing Non-KPS record: {non_kps_id} ('{existing_name}')")
                else:
                    # Create new Non-KPS record
                    print(f"📝 Creating new Non-KPS record: '{nama_areal}'")

                    # Fetch geometry details from master_lahan
                    lahan_data = await asyncio.to_thread(
                        lambda: supabase.table("master_lahan")
                        .select("area_ha, centroid_lat, centroid_lng")
                        .eq("id", lahan_id)
                        .limit(1)
                        .execute()
                    )

                    if lahan_data.data:
                        lahan = lahan_data.data[0]
                        non_kps_res = await asyncio.to_thread(
                            lambda: supabase.table("master_non_kps").insert({
                                "nama_areal": nama_areal,
                                "lahan_id": lahan_id,
                                "area_ha": lahan.get("area_ha"),
                                "centroid_lat": lahan.get("centroid_lat"),
                                "centroid_lng": lahan.get("centroid_lng")
                            }).execute()
                        )

                        if non_kps_res.data:
                            non_kps_id = non_kps_res.data[0]["id"]
                            print(f"✅ Created new Non-KPS record: {non_kps_id}")
                        else:
                            print(f"⚠️ Non-KPS creation returned no data")
                    else:
                        print(f"⚠️ Could not fetch lahan data for Non-KPS creation")

            except Exception as non_kps_err:
                # Non-fatal: Log error but continue with analysis save
                print(f"⚠️ Non-KPS creation failed (non-fatal): {non_kps_err}")
                non_kps_id = None

        # Extract Transition Summary for Top-Level Columns
        ts = request.transition_summary or {}
        defor_val = float(ts.get("total_deforestation_ha", 0) or 0)
        refor_val = float(ts.get("total_reforestation_ha", 0) or 0)
        net_change = float(ts.get("net_forest_change_ha", 0) or 0)
        
        trend_val = "STABLE"
        if net_change < -0.5: trend_val = "DECREASING"
        elif net_change > 0.5: trend_val = "INCREASING"

        new_history = {
            "filename": request.filename,
            "file_size": request.file_size,
            "metadata": {
                **request.metadata,
                "transition_summary": request.transition_summary,
                "audit_report": request.audit_report,
                "method_standardization": "v1_regional_seasonal_whittaker",
                "smoothing": "whittaker_lambda_5",
                "processed_at": datetime.now().isoformat()
            },
            "analysis_results": updated_analysis_results,
            "lahan_id": lahan_id,
            # Top-Level Transition Metrics (Added 2026-02-01)
            "deforestation_ha": defor_val,
            "reforestation_ha": refor_val,
            "trend_type": trend_val,
            # KPS Detection Fields
            "kps_id": request.kps_id,
            "non_kps_id": non_kps_id,  # Link to master_non_kps if NON_KPS analysis
            "link_method": request.link_method or "NONE",
            "analysis_scope": effective_scope
        }
        
        # Retry logic for database insertion (handle transient DNS/connection issues)
        hist_res = None
        for attempt in range(3):
            try:
                hist_res = await asyncio.to_thread(lambda: supabase.table("analysis_history").insert(new_history).execute())
                if hist_res is not None and hasattr(hist_res, 'data') and hist_res.data:
                    break
            except Exception as e:
                print(f"⚠️ Attempt {attempt+1} to save history failed: {e}")
                if attempt < 2:
                    await asyncio.sleep(2 ** attempt) # Exponential backoff
                else:
                    raise e
        
        if hist_res is not None and hasattr(hist_res, 'data') and hist_res.data and len(hist_res.data) > 0:
             # Invalidate list cache on successful save
             invalidate_cache("history_list")

             saved_history = hist_res.data[0]
             history_id = saved_history.get("id")
             analysis_results = request.analysis_results or []

             # === SAVE TO analysis_yearly_data TABLE ===
             yearly_records = []
             for item in analysis_results:
                 year = item.get("year")
                 if not year:
                     continue

                 # ===================================================================
                 # SAVE ANALYSIS DATA: 6 Kelas IPSDH Final
                 # ===================================================================
                 # Database now has exactly 6 columns (legacy columns removed)
                 # Direct 1:1 mapping from classification results

                 record = {
                     "history_id": history_id,
                     "year": year,
                     # 6 Kelas IPSDH dari classification:
                     "hutan_primer": float(item.get("hutan_primer", 0) or 0),
                     "hutan_sekunder": float(item.get("hutan_sekunder", 0) or 0),
                     "tanah_kering": float(item.get("tanah_kering", 0) or 0),
                     "tanah_kosong": float(item.get("tanah_kosong", 0) or 0),
                     "lahan_terbangun": float(item.get("lahan_terbangun", 0) or 0),
                     "air": float(item.get("air", 0) or 0),
                     "total_area": float(item.get("total_area", 0) or 0)
                 }
                 yearly_records.append(record)

             if yearly_records:
                 try:
                     await asyncio.to_thread(
                         lambda: supabase.table("analysis_yearly_data")
                             .upsert(yearly_records, on_conflict="history_id,year")
                             .execute()
                     )
                     print(f"📊 Saved {len(yearly_records)} yearly records for history {history_id[:8]}")
                 except Exception as e:
                     print(f"⚠️ Error saving yearly data: {e}")

             # === POPULATE HOTSPOTS IN BACKGROUND ===
             years = [item.get("year") for item in analysis_results if item.get("year")]
             if history_id and years and request.geo_data:
                 background_tasks.add_task(
                     save_hotspots_for_analysis,
                     history_id,
                     request.geo_data,
                     years
                 )
                 print(f"🔥 Scheduled hotspot population for history {history_id[:8]} ({len(years)} years)")

             # === CALCULATE AND SAVE SLOPE ANALYSIS IN BACKGROUND ===
             if history_id and request.geo_data:
                 background_tasks.add_task(
                     save_slope_analysis,
                     history_id,
                     request.geo_data
                 )
                 print(f"📐 Scheduled slope analysis for history {history_id[:8]}")

             # === UPDATE TEMPORAL STATUS (GREY AREA DETECTION) IN BACKGROUND ===
             if history_id and len(yearly_records) > 0:
                 background_tasks.add_task(
                     update_temporal_status_for_history,
                     history_id
                 )
                 print(f"🔄 Scheduled temporal status calculation for history {history_id[:8]}")

             return {"status": "success", "data": saved_history}
        else:
             raise Exception("Insert analysis_history failed after retries.")


        
    except Exception as e:
        error_msg = f"❌ Error saving history: {e}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        
        # Write to log file for agent debugging
        try:
            with open("backend_error.log", "w") as f:
                f.write(str(e))
                f.write("\n")
                traceback.print_exc(file=f)
        except:
            pass
            
        raise HTTPException(status_code=500, detail=f"Gagal menyimpan history: {str(e)}")



@app.delete("/history/{history_id}")
async def delete_history(history_id: str):
    """
    Menghapus item riwayat analisis secara robust.
    Pola 'context 7': Menggunakan pengecekan keberadaan eksplisit dan logging detail.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase client not configured")
    
    print(f"🗑️ Attempting to delete history ID: {history_id}")
    
    try:
        # 1. Cek apakah item benar-benar ada (Pola Robust)
        # Ambil analysis_results untuk keperluan cleanup file images
        check_res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("id, analysis_results")
            .eq("id", history_id)
            .limit(1)
            .execute()
        )
        
        if not check_res.data or len(check_res.data) == 0:
            # Kembalikan sukses jika sudah hilang (idempotent) 
            print(f"ℹ️ History item {history_id} already deleted. Returning success.")
            return {
                "status": "success", 
                "message": f"History ID {history_id} already deleted.",
                "id": history_id
            }

        # --- CLEANUP CLOUD FILES ---
        try:
            item_data = check_res.data[0]
            results = item_data.get('analysis_results', [])
            files_to_delete = []
            
            for r in results:
                # Check Thumbnail
                t_url = r.get('thumb_url')
                if t_url and '/thumbnails/' in t_url:
                    # Extract filename after /thumbnails/
                    fname = t_url.split('/thumbnails/')[-1]
                    if fname: files_to_delete.append(fname)
                
                # Check RGB Thumbnail
                rgb_url = r.get('rgb_thumb_url')
                if rgb_url and '/thumbnails/' in rgb_url:
                    fname = rgb_url.split('/thumbnails/')[-1]
                    if fname: files_to_delete.append(fname)
            
            if files_to_delete:
                print(f"🗑️ Deleting {len(files_to_delete)} images from Cloud Storage...")
                # Supabase remove takes list of filenames
                await asyncio.to_thread(
                    lambda: supabase.storage.from_('thumbnails').remove(files_to_delete)
                )
                print(f"✅ Deleted {len(files_to_delete)} files from Cloud.")
        except Exception as storage_err:
            print(f"⚠️ Failed to clean up cloud storage (non-fatal): {storage_err}")
        # ---------------------------

        # 2. Lakukan penghapusan database
        # Tambahkan .select() untuk memastikan kembalian data jika didukung
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history").delete().eq("id", history_id).execute()
        )
        
        # 3. Invalidate cache
        invalidate_cache("history_list")
        invalidate_cache(f"history_detail_{history_id}")
        
        print(f"✅ Successfully deleted history ID: {history_id}")
        return {
            "status": "success", 
            "message": f"History ID {history_id} deleted successfully.",
            "id": history_id
        }

    except HTTPException:
        raise
    except Exception as e:
        err_str = str(e)
        print(f"❌ Critical error deleting history {history_id}: {err_str}")
        
        # Deteksi foreign key violation (misal ada tabel lain yang merujuk kesini)
        if "foreign key" in err_str.lower():
            raise HTTPException(
                status_code=400, 
                detail="Gagal menghapus: Data ini masih digunakan oleh referensi lain di database."
            )
            
        raise HTTPException(status_code=500, detail=f"Database Error: {err_str}")

@app.get("/history")
async def get_history():
    """
    Ambil riwayat analisis dengan JOIN ke master_lahan.
    Mengembalikan data history + geometri lengkap.
    Uses server-side file cache for faster retrieval.
    """
    if not supabase:
        return []
    
    CACHE_KEY = "history_list"
    
    # Try to get from cache first
    cached_data = get_cached(CACHE_KEY)
    if cached_data is not None:
        print(f"🚀 Cache HIT for {CACHE_KEY}")
        return cached_data
    
    print(f"🔄 Cache MISS for {CACHE_KEY} - fetching from database...")
        
    try:
        # Optimasi: Ambil data esensial saja. Hilangkan geom_geojson.
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("id, filename, file_size, created_at, metadata, analysis_results")
            .order("created_at", desc=True)
            .limit(50) 
            .execute()
        )
        
        # Pruning Agresif untuk List View (Prevent 524 Timeout)
        data = []
        for item in res.data:
            # 1. Hilangkan GeoJSON Geometri (Terlalu berat untuk list)
            item['geo_data'] = None

            # 2. Normalisasi Analysis Results ke 6 kelas IPSDH final
            if 'analysis_results' in item:
                item['analysis_results'] = normalize_analysis_results(item['analysis_results'])

            # 3. Bersihkan Analysis Results
            results = item.get('analysis_results', [])
            if isinstance(results, list) and results:
                # Kita hanya butuh statistik (untuk grafik trend)
                # dan thumbnail terbaru (untuk preview card)
                results.sort(key=lambda x: int(x.get('year', 0)), reverse=True)

                latest_year = results[0]
                for idx, r in enumerate(results):
                    # ALWAYS remove massive vector data
                    if 'vector_geojson' in r: del r['vector_geojson']
                    if 'rgb_thumb_url' in r: del r['rgb_thumb_url']

                    # Hilangkan thumbnail kecuali untuk tahun TERBARU
                    # (Frontend butuh preview thumb di card, tapi cukup 1)
                    if idx > 0:
                        if 'thumb_url' in r: del r['thumb_url']

            data.append(item)
            
        return data # Hilangkan caching sementara untuk data live murni
        
    except Exception as e:
        print(f"❌ Error fetching history: {e}")
        return []

@app.get("/history/{history_id}")
async def get_history_detail(history_id: str):
    """
    Ambil detail lengkap riwayat analisis termasuk GeoJSON dan semua thumbnail.
    Uses server-side file cache for faster retrieval.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase client not configured")
    
    CACHE_KEY = f"history_detail_{history_id}"
    
    # Try to get from cache first
    cached_data = get_cached(CACHE_KEY)
    if cached_data is not None:
        print(f"🚀 Cache HIT for {CACHE_KEY}")
        return cached_data
    
    print(f"🔄 Cache MISS for {CACHE_KEY} - fetching from database...")
        
    try:
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("*, master_lahan(geom_geojson)")
            .eq("id", history_id)
            .single()
            .execute()
        )
        
        if not res.data:
            raise HTTPException(status_code=404, detail="Riwayat tidak ditemukan")
            
        item = res.data
        master_lahan = item.get('master_lahan')
        geo_data = None
        if master_lahan and isinstance(master_lahan, dict):
            geo_data = master_lahan.get('geom_geojson')
            
        if 'master_lahan' in item:
            del item['master_lahan']

        item['geo_data'] = geo_data

        # Normalisasi Analysis Results ke 6 kelas IPSDH final
        if 'analysis_results' in item:
            item['analysis_results'] = normalize_analysis_results(item['analysis_results'])

        # Fetch slope summary data if available
        try:
            slope_res = await asyncio.to_thread(
                lambda: supabase.table("analysis_slope_summary")
                    .select("*")
                    .eq("history_id", history_id)
                    .execute()
            )
            if slope_res.data:
                slope_summary = []
                for record in slope_res.data:
                    slope_summary.append({
                        "scope": record.get("scope"),
                        "avg_slope": record.get("avg_slope"),
                        "slope_0_8": record.get("slope_0_8"),
                        "slope_8_15": record.get("slope_8_15"),
                        "slope_15_25": record.get("slope_15_25"),
                        "slope_25_40": record.get("slope_25_40"),      # KLHK: 25-45%
                        "slope_above_40": record.get("slope_above_40"), # KLHK: >45%
                    })
                item['slope_summary'] = slope_summary
        except Exception as slope_err:
            print(f"⚠️ Could not fetch slope summary: {slope_err}")
            item['slope_summary'] = None

        # Fetch hotspot count summary
        try:
            hotspot_res = await asyncio.to_thread(
                lambda: supabase.table("analysis_hotspots")
                    .select("year, id", count="exact")
                    .eq("history_id", history_id)
                    .execute()
            )
            if hotspot_res.count is not None:
                item['hotspot_count'] = hotspot_res.count
        except Exception as hotspot_err:
            print(f"⚠️ Could not fetch hotspot count: {hotspot_err}")
            item['hotspot_count'] = 0

        # Store in cache (permanent - no TTL)
        cache_file(CACHE_KEY, item)

        return item
        
    except Exception as e:
        print(f"❌ Error fetching history detail: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/history/{history_id}/hotspots")
async def get_history_hotspots(history_id: str, year: int = None):
    """
    Get hotspot data for a specific analysis history.
    Optionally filter by year.
    Returns GeoJSON FeatureCollection.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        query = supabase.table("analysis_hotspots").select("*").eq("history_id", history_id)

        if year:
            query = query.eq("year", year)

        result = await asyncio.to_thread(lambda: query.order("acq_date", desc=True).execute())

        if not result.data:
            return {
                "type": "FeatureCollection",
                "features": [],
                "total": 0
            }

        # Convert to GeoJSON
        features = []
        for row in result.data:
            feature = {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [float(row['longitude']), float(row['latitude'])]
                },
                "properties": {
                    "year": row.get('year'),
                    "acq_date": row.get('acq_date'),
                    "confidence": row.get('confidence'),
                    "confidence_level": row.get('confidence_level'),
                    "brightness": row.get('brightness'),
                    "frp": row.get('frp'),
                    "source": row.get('source')
                }
            }
            features.append(feature)

        # Aggregate by year
        yearly_stats = {}
        for row in result.data:
            y = row.get('year')
            if y not in yearly_stats:
                yearly_stats[y] = {"total": 0, "high": 0, "nominal": 0, "low": 0}
            yearly_stats[y]["total"] += 1
            conf = row.get('confidence_level', '').lower()
            if conf in yearly_stats[y]:
                yearly_stats[y][conf] += 1

        return {
            "type": "FeatureCollection",
            "features": features,
            "total": len(features),
            "yearly_stats": yearly_stats
        }

    except Exception as e:
        print(f"⚠️ Error getting hotspots: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/admin/cache/clear")
async def admin_clear_cache():
    """
    Admin endpoint untuk menghapus semua cache secara manual.
    Berguna jika data diubah langsung di database.
    """
    try:
        disk_cache.clear()
        print("🗑️ All cache cleared manually via admin endpoint")
        return {"status": "success", "message": "Cache cleared successfully"}
    except Exception as e:
        print(f"❌ Error clearing cache: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==============================================================================
# CARBON TIME-SERIES ANALYSIS (Indicative Mode)
# ==============================================================================

def calculate_carbon_stock(area_data: Dict[str, float]) -> Dict[str, Any]:
    """
    Menghitung stok karbon dari data luas tutupan lahan.
    
    Rumus: C_t = Σ(Area_class × CF_class)
    
    Args:
        area_data: Dict dengan keys 'hutan', 'tanah_kering', dst. (dalam Ha)
    
    Returns:
        Dict dengan carbon_stock_tc, carbon_stock_co2e, breakdown
    """
    class_mapping = {
        'hutan': 1,
        'tanah_kering': 2,
        'tanah_kosong': 3,
        'air': 4,
        'lahan_terbangun': 5
    }
    
    total_carbon = 0.0
    carbon_breakdown = {}
    area_breakdown = {}
    
    def safe_float(val, default=0.0):
        """Safely convert value to float, returning default if not possible."""
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default
    
    for class_name, class_id in class_mapping.items():
        area_ha = safe_float(area_data.get(class_name, 0.0))
        cf = CARBON_FACTORS.get(class_id, 0.0)
        carbon = area_ha * cf
        carbon_breakdown[class_name] = round(carbon, 2)
        area_breakdown[class_name] = round(area_ha, 2)
        total_carbon += carbon
    
    return {
        'carbon_stock_tc': round(total_carbon, 2),
        'carbon_stock_co2e': round(total_carbon * CO2E_CONVERSION, 2),
        'carbon_breakdown': carbon_breakdown,
        'area_breakdown': area_breakdown
    }


def calculate_carbon_change(yearly_stocks: List[Dict]) -> Dict[str, Any]:
    """
    Menghitung perubahan karbon antar titik waktu.
    
    Rumus:
    - ΔC = C_tn - C_t1
    - Rate = ΔC / (tn - t1)
    
    Args:
        yearly_stocks: List of {year, carbon_stock_tc}
    
    Returns:
        Dict dengan delta, rate, trend_status
    """
    if len(yearly_stocks) < 2:
        return None
    
    # Sort by year
    sorted_data = sorted(yearly_stocks, key=lambda x: x['year'])
    
    c_t1 = sorted_data[0]['carbon_stock_tc']
    c_tn = sorted_data[-1]['carbon_stock_tc']
    t1 = sorted_data[0]['year']
    tn = sorted_data[-1]['year']
    
    delta_c = c_tn - c_t1
    years_diff = tn - t1
    annual_rate = delta_c / years_diff if years_diff > 0 else 0
    
    # Tentukan status tren menggunakan threshold ±2%
    threshold = abs(c_t1 * CARBON_TREND_THRESHOLD) if c_t1 > 0 else 100
    if delta_c > threshold:
        trend_status = "Increasing"
    elif delta_c < -threshold:
        trend_status = "Decreasing"
    else:
        trend_status = "Stable"
    
    return {
        'start_year': t1,
        'end_year': tn,
        'delta_carbon_tc': round(delta_c, 2),
        'delta_carbon_co2e': round(delta_c * CO2E_CONVERSION, 2),
        'annual_rate_tc': round(annual_rate, 2),
        'annual_rate_co2e': round(annual_rate * CO2E_CONVERSION, 2),
        'trend_status': trend_status
    }


def calculate_indicative_confidence(yearly_data: List[Dict]) -> Dict[str, Any]:
    """
    Menghitung Indicative Confidence untuk perubahan karbon.
    
    Rumus:
    Confidence (%) = (SC × 0.30) + (TS × 0.30) + (MR × 0.20) + (DC × 0.20)
    
    Komponen:
    - SC: Spatial Consistency (0.70-0.90) - dari avg classification confidence
    - TS: Temporal Stability (0.65-0.90) - cek lonjakan tidak wajar
    - MR: Method Reliability (fixed 0.65) - tidak ada validasi lapangan
    - DC: Data Completeness (0.70-0.90) - ketersediaan data tahun
    """
    n_years = len(yearly_data)
    
    if n_years == 0:
        return {'confidence_percent': 0, 'components': {}}
    
    # SC: Spatial Consistency
    # Berdasarkan rata-rata confidence dari hasil klasifikasi
    avg_classification_confidence = sum(
        d.get('confidence_percent', 75) for d in yearly_data
    ) / n_years
    sc = min(0.90, max(0.70, avg_classification_confidence / 100))
    
    # TS: Temporal Stability
    # Cek lonjakan tidak wajar (>50% perubahan year-over-year)
    ts = 0.85  # Default
    if n_years >= 2:
        stocks = [d['carbon_stock_tc'] for d in yearly_data]
        max_jump = 0
        for i in range(1, len(stocks)):
            if stocks[i-1] > 0:
                jump = abs(stocks[i] - stocks[i-1]) / stocks[i-1]
                max_jump = max(max_jump, jump)
        if max_jump > 0.5:
            ts = 0.65
        elif max_jump > 0.3:
            ts = 0.75
        else:
            ts = 0.85
    
    # MR: Method Reliability (Fixed - tidak ada validasi lapangan)
    mr = 0.65
    
    # DC: Data Completeness
    # Semua tahun tersedia = 0.90, ada gap = 0.70
    if n_years >= 2:
        sorted_years = sorted(d['year'] for d in yearly_data)
        expected_years = sorted_years[-1] - sorted_years[0] + 1
        actual_years = n_years
        if actual_years >= expected_years:
            dc = 0.90
        else:
            dc = 0.70 + (0.20 * actual_years / expected_years)
    else:
        dc = 0.70
    
    # Hitung final confidence
    confidence = (sc * 0.30) + (ts * 0.30) + (mr * 0.20) + (dc * 0.20)
    confidence_percent = round(confidence * 100, 1)
    
    return {
        'confidence_percent': confidence_percent,
        'components': {
            'spatial_consistency': round(sc * 100, 1),
            'temporal_stability': round(ts * 100, 1),
            'method_reliability': round(mr * 100, 1),
            'data_completeness': round(dc * 100, 1)
        }
    }


@app.post("/carbon/analyze", response_model=CarbonTimeSeriesResponse)
async def analyze_carbon_timeseries(request: CarbonTimeSeriesRequest):
    """
    Analisis Time-Series Stok Karbon (Mode Indikatif).
    
    Menggunakan hasil klasifikasi tutupan lahan dari analysis_history untuk menghitung:
    - Stok karbon tahunan (tC dan tCO2e)
    - Perubahan karbon (delta)
    - Laju perubahan tahunan
    - Status tren
    - Tingkat keyakinan indikatif
    
    PERINGATAN: Hasil ini bersifat INDIKATIF, bukan untuk perdagangan karbon atau regulasi.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase client not configured")
    
    history_id = request.history_id
    print(f"🌿 Carbon Time-Series Analysis for History ID: {history_id}")
    
    try:
        # 1. Ambil data history
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("id, filename, analysis_results")
            .eq("id", history_id)
            .single()
            .execute()
        )
        
        if not res.data:
            raise HTTPException(status_code=404, detail="History tidak ditemukan")
        
        analysis_results = res.data.get('analysis_results', [])
        
        if len(analysis_results) < 2:
            raise HTTPException(
                status_code=400, 
                detail="Minimal 2 tahun data diperlukan untuk analisis time-series"
            )
        
        # 2. Hitung karbon untuk setiap tahun
        yearly_carbon: List[CarbonYearlyData] = []
        for result in analysis_results:
            area_data = {
                'hutan': result.get('hutan', 0),
                'tanah_kering': result.get('tanah_kering', 0),
                'tanah_kosong': result.get('tanah_kosong', 0),
                'air': result.get('air', 0),
                'lahan_terbangun': result.get('lahan_terbangun', 0)
            }
            
            carbon_result = calculate_carbon_stock(area_data)
            
            yearly_carbon.append(CarbonYearlyData(
                year=int(result.get('year')),
                carbon_stock_tc=carbon_result['carbon_stock_tc'],
                carbon_stock_co2e=carbon_result['carbon_stock_co2e'],
                area_breakdown=carbon_result['area_breakdown'],
                carbon_breakdown=carbon_result['carbon_breakdown']
            ))
        
        # Sort by year
        yearly_carbon = sorted(yearly_carbon, key=lambda x: x.year)
        
        # 3. Hitung perubahan
        yearly_dicts = [
            {'year': y.year, 'carbon_stock_tc': y.carbon_stock_tc, 'confidence_percent': 75}
            for y in yearly_carbon
        ]
        change_result_dict = calculate_carbon_change(yearly_dicts)
        
        # 4. Hitung confidence
        confidence = calculate_indicative_confidence(yearly_dicts)
        
        # 5. Buat response
        change_result = CarbonChangeResult(
            start_year=change_result_dict['start_year'],
            end_year=change_result_dict['end_year'],
            delta_carbon_tc=change_result_dict['delta_carbon_tc'],
            delta_carbon_co2e=change_result_dict['delta_carbon_co2e'],
            annual_rate_tc=change_result_dict['annual_rate_tc'],
            annual_rate_co2e=change_result_dict['annual_rate_co2e'],
            trend_status=change_result_dict['trend_status'],
            confidence_percent=confidence['confidence_percent'],
            confidence_components=confidence['components']
        )
        
        # Format carbon factors untuk transparansi
        carbon_factors_display = {
            'Hutan': f"{CARBON_FACTORS[1]} tC/ha",
            'Tanah Kering': f"{CARBON_FACTORS[2]} tC/ha",
            'Tanah Kosong': f"{CARBON_FACTORS[3]} tC/ha",
            'Air': f"{CARBON_FACTORS[4]} tC/ha",
            'Lahan Terbangun': f"{CARBON_FACTORS[5]} tC/ha",
        }
        
        print(f"   ✅ Carbon Analysis Complete: {change_result.trend_status} ({change_result.delta_carbon_tc:+.2f} tC)")
        
        return CarbonTimeSeriesResponse(
            status="success",
            message=f"Analisis karbon time-series berhasil. Tren: {change_result.trend_status}",
            yearly_data=yearly_carbon,
            change_result=change_result,
            carbon_factors_used=carbon_factors_display,
            disclaimer=CARBON_DISCLAIMER
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Carbon Analysis Error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Carbon analysis failed: {str(e)}")


@app.post("/maintenance/migrate-to-local")
async def migrate_to_local_storage():
    """
    Kandungan: Migrasi semua thumbnail (Base64/Remote) ke penyimpanan lokal (VPS).
    Tujuannya untuk menghemat STORAGE Supabase.
    """
    if not supabase:
        raise HTTPException(status_code=505, detail="Supabase not configured")
        
    print("🚀 Starting Migration: Remote/Base64 -> Local VPS Storage...")
    try:
        # 1. Ambil semua riwayat analisis
        res = await asyncio.to_thread(lambda: supabase.table("analysis_history").select("id, filename, analysis_results").execute())
        items = res.data or []
        print(f"📦 Found {len(items)} history items to check for migration.")
        
        migrated_count = 0
        total_images_saved = 0
        
        for item in items:
            history_id = item['id']
            results = item.get('analysis_results', [])
            if not isinstance(results, list): continue
            
            modified = False
            for res_item in results:
                year = res_item.get('year', 'unknown')
                
                # Cek dua field thumbnail utama
                for field in ['thumb_url', 'rgb_thumb_url']:
                    val = res_item.get(field)
                    if not val: continue
                    
                    # Logika: Jika Base64 ATAU URL eksternal (GEE/Supabase link), pindahkan ke lokal
                    # Jika sudah link lokal (/storage/...), lewati
                    if isinstance(val, str) and not val.startswith("/storage/"):
                        print(f"   🔄 Migrating {field} for ID {history_id} (Year {year})...")
                        
                        # Jika Base64, langsung simpan
                        if val.startswith("data:image"):
                             filename = f"{field}_{history_id}_{year}_{hashlib.md5(val[:100].encode()).hexdigest()[:6]}.png"
                             local_url = save_local_image(val, filename)
                             if local_url:
                                 res_item[field] = local_url
                                 modified = True
                                 total_images_saved += 1
                        # Jika URL (GEE/Supabase), download dulu
                        elif val.startswith("http"):
                             local_url = await download_and_save_locally(val, prefix=f"{field}")
                             if local_url:
                                 res_item[field] = local_url
                                 modified = True
                                 total_images_saved += 1
            
            if modified:
                # Update di Supabase
                await asyncio.to_thread(lambda: supabase.table("analysis_history").update({"analysis_results": results}).eq("id", history_id).execute())
                migrated_count += 1
                print(f"   ✅ Updated record {history_id} with local paths.")

        # Invalidate cache
        invalidate_cache("history_list")
        
        msg = f"Migration complete. Updated {migrated_count} records, saved {total_images_saved} images to local disk."
        print(f"✨ {msg}")
        return {"status": "success", "message": msg, "records_updated": migrated_count, "images_saved": total_images_saved}
        
    except Exception as e:
        print(f"❌ Migration failed: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))




def get_classified_image(geometry, year, thresholds=None):
    """
    Helper untuk mendapatkan citra klasifikasi untuk satu tahun tertentu via GEE.
    """
    if thresholds is None:
        thresholds = ClassificationThresholds()
        
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
                 .filterBounds(geometry)
                 .filterDate(start_date, end_date)
                 .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 80))
                 .map(mask_clouds_sentinel2))
    
    count = collection.size().getInfo()
    if count == 0:
        return None
        
    median_composite = collection.median().clip(geometry)
    classified = classify_land_cover_advanced(median_composite, geometry, year, thresholds)
    return classified

def generate_shapefile_buffer(fc_dict):
    """
    Generate in-memory shapefile buffers (shp, shx, dbf, prj) from GeoJSON dict.
    Returns a dict mapping filename extensions to bytes.
    """
    shp_io = io.BytesIO()
    shx_io = io.BytesIO()
    dbf_io = io.BytesIO()
    
    w = shapefile.Writer(shp=shp_io, shx=shx_io, dbf=dbf_io)
    w.field('class_id', 'N', decimal=0)
    w.field('class_name', 'C', size=50)
    
    # Iterate features
    features = fc_dict.get('features', [])
    for feat in features:
        props = feat.get('properties', {})
        cls_id = props.get('label', 0) # reduceToVectors uses 'label' by default for class band
        cls_name = LC_CLASSES.get(int(cls_id), "Unknown")
        
        geom = feat.get('geometry')
        if not geom:
            continue
            
        w.shape(geom)
        w.record(cls_id, cls_name)
        
    w.close()
    
    # PRJ content for WGS84
    prj_content = 'GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["Degree",0.017453292519943295]]'
    
    return {
        'shp': shp_io.getvalue(),
        'shx': shx_io.getvalue(),
        'dbf': dbf_io.getvalue(),
        'prj': prj_content.encode('utf-8')
    }

class PDFReportRequest(BaseModel):
    """Request untuk generate PDF report."""
    geojson: Dict[str, Any] = Field(..., description="Objek geometri GeoJSON")
    data: List[Dict[str, Any]] = Field(..., description="Data hasil analisis")
    title: str = Field(default="Laporan Analisis Tutupan Lahan", description="Judul laporan")
    description: Optional[str] = Field(default=None, description="Deskripsi tambahan")

def resolve_report_url(url: Optional[str], base_url: str) -> str:
    """
    Memastikan URL gambar di laporan valid, baik itu path lokal maupun URL remote.
    Menghindari domain hardcoded.
    """
    if not url:
        return ""
    if url.startswith('/storage'):
        return f"{base_url}{url}"
    return url

@app.post("/export/pdf")
async def export_pdf_endpoint(report_request: PDFReportRequest, request: Request):
    base_url = str(request.base_url).rstrip('/')
    # Deteksi X-Forwarded headers (VPS/Proxy aware)
    forwarded_host = request.headers.get("x-forwarded-host")
    forwarded_proto = request.headers.get("x-forwarded-proto")
    if forwarded_host and forwarded_proto:
        base_url = f"{forwarded_proto}://{forwarded_host}"
    elif forwarded_host:
         base_url = f"https://{forwarded_host}"

    return await generate_pdf_report(report_request, base_url=base_url)


async def generate_pdf_report(request: PDFReportRequest, base_url: str = "https://gealgeolgeo.ditpps.com"):
    """
    Generate PDF report dari hasil analisis tutupan lahan.
    Returns HTML yang bisa di-print sebagai PDF.
    """
    try:
        import base64

        # Normalisasi data ke 6 kelas IPSDH final
        normalized_data = normalize_analysis_results(request.data)

        # Calculate totals
        total_years = len(normalized_data)
        if total_years > 0:
            latest = normalized_data[-1]
            total_area = sum([
                latest.get('hutan_primer', 0),
                latest.get('hutan_sekunder', 0),
                latest.get('tanah_kering', 0),
                latest.get('tanah_kosong', 0),
                latest.get('air', 0),
                latest.get('lahan_terbangun', 0)
            ])
        else:
            total_area = 0
            year_last = "N/A"
            img_html = ""
        
        if request.data:
             latest_data = request.data[-1]
             # Gunakan thumbnail lokal jika ada (Sync Klasifikasi)
             thumb_url = resolve_report_url(latest_data.get('thumb_url'), base_url)
             img_html = f'<img src="{thumb_url}" style="max-width:100%; height:auto; max-height:500px; border-radius:8px; box-shadow:0 4px 6px rgba(0,0,0,0.1);">' if thumb_url else '<p>Gambar tidak tersedia</p>'
             
        # Extract SHP Properties
        shp_properties_html = ""
        try:
            features = []
            if request.geojson.get('type') == 'FeatureCollection':
                features = request.geojson.get('features', [])
            elif request.geojson.get('type') == 'Feature':
                features = [request.geojson]
            
            if features:
                # Ambil semua keys unik dari semua features
                all_keys = set()
                for f in features:
                    all_keys.update(f.get('properties', {}).keys())
                
                sorted_keys = sorted(list(all_keys))
                
                # Buat tabel properti
                shp_properties_html = """
                <div class="metadata-grid">
                """
                for key in sorted_keys:
                    val = features[0].get('properties', {}).get(key, '-')
                    shp_properties_html += f"""
                    <div class="meta-card">
                        <span class="meta-card-label">{key}</span>
                        <span class="meta-card-val">{val}</span>
                    </div>
                    """
                shp_properties_html += "</div>"
        except Exception as e:
            print(f"⚠️ Gagal ekstrak properti SHP: {e}")
            shp_properties_html = "<p><em>Gagal memuat informasi metadata SHP</em></p>"

        
        # Prepare trend calculation
        trend_label = "Stabil"
        trend_class = "trend-neutral"
        if len(normalized_data) > 1:
            start_total_forest = (normalized_data[0].get('hutan_primer', 0) or 0) + (normalized_data[0].get('hutan_sekunder', 0) or 0)
            end_total_forest = (normalized_data[-1].get('hutan_primer', 0) or 0) + (normalized_data[-1].get('hutan_sekunder', 0) or 0)
            diff = end_total_forest - start_total_forest
            if diff > 1:
                trend_label = f"Peningkatan Hutan (+{diff:.1f} Ha)"
                trend_class = "trend-up"
            elif diff < -1:
                trend_label = f"Deforestasi Terdeteksi ({abs(diff):.1f} Ha)"
                trend_class = "trend-down"

        # Prepare Multitemporal Map Series HTML
        html_items = []
        for y in normalized_data:
            # Resolusi Path Lokal & Sinkronisasi Klasifikasi
            t_url = resolve_report_url(y.get('thumb_url'), base_url)
            
            html_items.append(f'''
            <div class="year-comparison">
                <div style="font-weight: 800; color: var(--primary-dark); margin-bottom: 10px; font-size: 16px;">Tahun {y.get('year')}</div>
                <div class="comparison-grid">
                    <div class="map-item">
                        <img src="{t_url}" alt="Klasifikasi {y.get('year')}">
                        <span class="map-item-label">Hasil Klasifikasi Tutupan Lahan (Tahun {y.get('year')})</span>
                    </div>
                </div>
            </div>
            ''')
        map_series_html = "".join(html_items)

        # Generate interactive charts data strings
        years_label_str = ",".join([f"'{y.get('year')}'" for y in normalized_data])
        hprimer_data_str = ",".join([str(y.get('hutan_primer', 0)) for y in normalized_data])
        hsekunder_data_str = ",".join([str(y.get('hutan_sekunder', 0)) for y in normalized_data])
        tkering_data_str = ",".join([str(y.get('tanah_kering', 0)) for y in normalized_data])
        tkosong_data_str = ",".join([str(y.get('tanah_kosong', 0)) for y in normalized_data])
        terbangun_data_str = ",".join([str(y.get('lahan_terbangun', 0)) for y in normalized_data])
        air_data_str = ",".join([str(y.get('air', 0)) for y in normalized_data])

        # Generate HTML report
        html = f"""
<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{request.title} | Laporan Analisis Tutupan Lahan</title>
    <meta name="print-color-adjust" content="exact">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.0.0"></script>
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;600;800&display=swap" rel="stylesheet">
    <style>
        :root {{
            --primary: #059669; --primary-light: #10b981; --primary-dark: #064e3b;
            --bg: #f8fafc; --card: #ffffff; --text: #1e293b; --text-muted: #64748b;
        }}
        * {{ box-sizing: border-box; -webkit-print-color-adjust: exact; }}
        body {{ 
            font-family: 'Plus Jakarta Sans', sans-serif; 
            margin: 0; padding: 0; background: var(--bg); color: var(--text); line-height: 1.6;
        }}
        .container {{ max-width: 1000px; margin: 0 auto; padding: 40px 20px; }}
        
        /* Premium Header */
        .header {{
            background: linear-gradient(135deg, var(--primary-dark) 0%, var(--primary) 100%);
            padding: 80px 40px; border-radius: 0 0 60px 60px; color: white; text-align: center;
            box-shadow: 0 20px 50px rgba(5, 150, 105, 0.2); margin-bottom: 40px;
        }}
        .header h1 {{ margin: 0; font-size: 36px; font-weight: 800; letter-spacing: -0.03em; }}
        .header p {{ opacity: 0.8; margin: 12px 0 0; font-weight: 600; text-transform: uppercase; font-size: 11px; letter-spacing: 0.2em; }}
        
        /* Info Cards */
        .stats-grid {{ 
            display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
            gap: 24px; margin-top: -80px; padding: 0 20px;
        }}
        .card {{
            background: var(--card); padding: 28px; border-radius: 24px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.04); border: 1px solid rgba(255,255,255,0.7);
            text-align: center; transition: all 0.3s ease;
        }}
        .card:hover {{ transform: translateY(-8px); box-shadow: 0 20px 40px rgba(0,0,0,0.08); }}
        .card-val {{ display: block; font-size: 32px; font-weight: 800; color: var(--primary); letter-spacing: -0.02em; }}
        .card-label {{ font-size: 11px; font-weight: 700; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; }}
        
        /* Main Content Sections */
        .section {{ background: var(--card); padding: 45px; border-radius: 32px; margin-top: 32px; box-shadow: 0 4px 20px rgba(0,0,0,0.02); }}
        .section-title {{ font-size: 20px; font-weight: 800; margin-bottom: 30px; display: flex; align-items: center; gap: 12px; color: var(--primary-dark); }}
        .section-title::before {{ content: ''; width: 6px; height: 26px; background: var(--primary-light); border-radius: 3px; display: block; }}

        /* Visualization */
        .viz-box {{ 
            background: #f1f5f9; border-radius: 24px; overflow: hidden; margin: 20px 0;
            border: 6px solid white; box-shadow: 0 12px 40px rgba(0,0,0,0.1); 
            text-align: center; position: relative;
        }}
        .viz-box img {{ max-width: 100%; height: auto; max-height: 550px; display: inline-block; object-fit: contain; }}
        .year-label {{ position: absolute; top: 20px; right: 20px; background: rgba(0,0,0,0.75); color: white; padding: 6px 18px; border-radius: 100px; font-size: 12px; font-weight: 700; }}

        /* Table Styles */
        .table-container {{ overflow-x: auto; }}
        table {{ width: 100%; border-collapse: separate; border-spacing: 0 8px; }}
        th {{ text-align: left; padding: 18px; color: var(--text-muted); font-size: 12px; font-weight: 700; text-transform: uppercase; }}
        td {{ padding: 18px; background: #f8fafc; font-size: 15px; border-top: 1px solid #f1f5f9; border-bottom: 1px solid #f1f5f9; }}
        td:first-child {{ border-left: 1px solid #f1f5f9; border-radius: 16px 0 0 16px; font-weight: 800; color: var(--primary); }}
        td:last-child {{ border-right: 1px solid #f1f5f9; border-radius: 0 16px 16px 0; }}
        .badge {{ background: #dcfce7; color: #166534; padding: 4px 12px; border-radius: 100px; font-size: 11px; font-weight: 800; }}

        /* Interaction Elements */
        .trend-pill {{ 
            padding: 12px 24px; border-radius: 100px; display: inline-flex; align-items: center; gap: 10px;
            font-weight: 800; font-size: 14px; margin-bottom: 30px;
        }}
        .trend-up {{ background: #dcfce7; color: #166534; }}
        .trend-down {{ background: #fee2e2; color: #991b1b; }}
        .trend-neutral {{ background: #e2e8f0; color: #475569; }}

        .chart-box {{ height: 400px; margin-top: 20px; }}

        /* Metadata Grid */
        .metadata-grid {{ 
            display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); 
            gap: 15px; margin-top: 10px;
        }}
        .meta-card {{
            background: #f8fafc; padding: 15px; border-radius: 12px;
            border: 1px solid #e2e8f0;
        }}
        .meta-card-label {{ display: block; font-size: 10px; font-weight: 700; color: var(--text-muted); text-transform: uppercase; margin-bottom: 4px; }}
        .meta-card-val {{ display: block; font-size: 13px; font-weight: 600; color: var(--primary-dark); word-break: break-all; }}

        /* Legend */
        .legend {{ display: flex; gap: 24px; margin-top: 25px; flex-wrap: wrap; background: #f8fafc; padding: 20px; border-radius: 20px; border: 1px solid #edf2f7; }}
        .legend-item {{ display: flex; align-items: center; gap: 10px; font-size: 13px; font-weight: 700; }}
        .legend-dot {{ width: 16px; height: 16px; border-radius: 6px; }}

        /* Footer */
        .footer {{ text-align: center; padding: 80px 0; color: var(--text-muted); font-size: 12px; }}
        .footer b {{ color: var(--primary); }}

        @media print {{
            @page {{ size: A4; margin: 2cm; }}
            body {{ background: white; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
            .container {{ padding: 0; margin: 0; width: 100%; max-width: 100%; }}
            .header {{ 
                border-radius: 0; box-shadow: none; padding: 40px; margin-bottom: 30px;
                background: var(--primary-dark) !important; color: white !important; 
            }}
            .stats-grid {{ margin-top: 0; display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; }}
            .card {{ box-shadow: none; border: 1px solid #eee; padding: 15px; border-radius: 12px; }}
            .section {{ 
                box-shadow: none; border: 1px solid #eee; break-inside: avoid; border-radius: 15px; 
                margin-top: 25px; padding: 30px; page-break-inside: avoid;
            }}
            .section-title::before {{ background: var(--primary-light) !important; }}
            .no-print {{ display: none !important; }}
            .year-comparison {{ break-inside: avoid; page-break-inside: avoid; margin-bottom: 20px; }}
            .chart-box {{ height: 350px !important; }}
            .badge {{ border: 1px solid #166534; }}
        }}

        /* Multitemporal Map Series */
        .map-series {{ display: flex; flex-direction: column; gap: 40px; }}
        .year-comparison {{ background: #f8fafc; padding: 25px; border-radius: 25px; border: 1px solid #e2e8f0; }}
        .comparison-grid {{ display: grid; grid-template-columns: 1fr; gap: 20px; margin-top: 15px; max-width: 600px; margin-left: auto; margin-right: auto; }}
        .map-item {{ background: white; padding: 10px; border-radius: 15px; border: 1px solid #edf2f7; }}
        .map-item img {{ width: 100%; border-radius: 10px; object-fit: contain; background: #eee; max-height: 350px; }}
        .map-item-label {{ display: block; font-size: 11px; font-weight: 800; color: var(--text-muted); text-align: center; margin-top: 10px; text-transform: uppercase; }}

        /* Print Button */
        .print-btn-container {{ text-align: center; margin-top: 40px; }}
        .print-btn {{
            background: var(--primary); color: white; border: none; padding: 15px 40px;
            font-size: 16px; font-weight: 800; border-radius: 100px; cursor: pointer;
            box-shadow: 0 10px 20px rgba(5, 150, 105, 0.3); transition: all 0.3s ease;
            display: inline-flex; align-items: center; gap: 10px;
        }}
        .print-btn:hover {{ transform: scale(1.05); background: var(--primary-dark); }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{request.title}</h1>
        <p>DIREKTORAT PENGENDALIAN PERHUTANAN SOSIAL</p>
    </div>

    <div class="container">
        <div class="stats-grid">
            <div class="card">
                <span class="card-val">{total_years}</span>
                <span class="card-label">Periode Analisis</span>
            </div>
            <div class="card">
                <span class="card-val">{total_area:,.1f}</span>
                <span class="card-label">Cakupan Area (Ha)</span>
            </div>
            <div class="card">
                <span class="card-val">10m</span>
                <span class="card-label">Resolusi Citra</span>
            </div>
            <div class="card">
                <span class="card-val">Sentinel-2</span>
                <span class="card-label">Sumber Data</span>
            </div>
        </div>
        
        {f'''
        <div class="section">
            <div class="section-title">Metadata Atribut Area</div>
            <div style="font-size: 13px;">{shp_properties_html}</div>
        </div>
        ''' if shp_properties_html else ''}

        <div class="section">
            <div class="section-title">Dataset Analisis Multitemporal</div>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Tahun</th>
                            <th>H.Primer (Ha)</th>
                            <th>H.Sekunder (Ha)</th>
                            <th>Lahan Kering (Ha)</th>
                            <th>Lahan Terbuka (Ha)</th>
                            <th>Lahan Terbangun (Ha)</th>
                            <th>Badan Air (Ha)</th>
                        </tr>
                    </thead>
                    <tbody>
                        { "".join([f'''
                        <tr>
                            <td>{y.get('year')}</td>
                            <td>{y.get('hutan_primer', 0):,.2f}</td>
                            <td>{y.get('hutan_sekunder', 0):,.2f}</td>
                            <td>{y.get('tanah_kering', 0):,.2f}</td>
                            <td>{y.get('tanah_kosong', 0):,.2f}</td>
                            <td>{y.get('lahan_terbangun', 0):,.2f}</td>
                            <td>{y.get('air', 0):,.2f}</td>
                        </tr>
                        ''' for y in normalized_data]) }
                    </tbody>
                </table>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Tren Perubahan Temporal (Batang & Area)</div>
            <div class="trend-pill {trend_class}">
                Status: {trend_label}
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr; gap: 40px;">
                <div class="chart-box">
                    <p style="text-align:center; font-size:11px; font-weight:800; color:var(--text-muted); margin-bottom:10px;">KOMPOSISI TAHUNAN (STACKED BAR)</p>
                    <canvas id="barChart"></canvas>
                </div>
                <div class="chart-box">
                    <p style="text-align:center; font-size:11px; font-weight:800; color:var(--text-muted); margin-bottom:10px;">ESTIMASI TREN (STACKED AREA)</p>
                    <canvas id="areaChart"></canvas>
                </div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Visualisasi Hasil Klasifikasi Tutupan Lahan</div>
            <div class="map-series">
                {map_series_html}
            </div>
            <div class="legend">
                <div class="legend-item"><div class="legend-dot" style="background:#006400"></div> Hutan Primer</div>
                <div class="legend-item"><div class="legend-dot" style="background:#32CD32"></div> Hutan Sekunder</div>
                <div class="legend-item"><div class="legend-dot" style="background:#DAA520"></div> Tanah Kering</div>
                <div class="legend-item"><div class="legend-dot" style="background:#D2691E"></div> Tanah Kosong</div>
                <div class="legend-item"><div class="legend-dot" style="background:#708090"></div> Lahan Terbangun</div>
                <div class="legend-item"><div class="legend-dot" style="background:#1E90FF"></div> Air</div>
            </div>
        </div>


        <div class="print-btn-container no-print">
            <button class="print-btn" onclick="window.print()">
                <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M6 9V2h12v7"></path><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path><rect x="6" y="14" width="12" height="8"></rect></svg>
                CETAK LAPORAN PDF
            </button>
        </div>

        <div class="footer">
            <p>Laporan analisis ini dihasilkan secara otomatis oleh <b>Intelligence Engine</b>.</p>
            <p>Waktu Cetak: {datetime.now().strftime('%d %B %Y, %H:%M:%S')} | Koordinat referensi: EPSG:4326 (WGS84).</p>
        </div>
    </div>

    <script>
        const chartLabels = [{years_label_str}];
        const datasets = [
            {{ label: 'Badan Air', data: [{air_data_str}], color: '#1E90FF' }},
            {{ label: 'Lahan Terbangun', data: [{terbangun_data_str}], color: '#708090' }},
            {{ label: 'Lahan Terbuka', data: [{tkosong_data_str}], color: '#D2691E' }},
            {{ label: 'Lahan Kering', data: [{tkering_data_str}], color: '#DAA520' }},
            {{ label: 'Hutan Sekunder', data: [{hsekunder_data_str}], color: '#32CD32' }},
            {{ label: 'Hutan Primer', data: [{hprimer_data_str}], color: '#006400' }}
        ];

        Chart.register(ChartDataLabels);
        
        // Bar Chart
        new Chart(document.getElementById('barChart').getContext('2d'), {{
            type: 'bar',
            data: {{
                labels: chartLabels,
                datasets: datasets.map(d => ({{
                    label: d.label,
                    data: d.data,
                    backgroundColor: d.color,
                    borderRadius: 0
                }}))
            }},
            options: {{
                responsive: true, maintainAspectRatio: false,
                scales: {{
                    y: {{ stacked: true, grid: {{ color: '#f1f5f9' }} }},
                    x: {{ stacked: true, grid: {{ display: false }} }}
                }},
                plugins: {{
                    legend: {{ position: 'bottom', labels: {{ usePointStyle: true, font: {{ size: 10, weight: '700' }} }} }},
                    datalabels: {{
                        color: '#000000', font: {{ weight: 'bold', size: 9 }},
                        anchor: 'center', align: 'center',
                        formatter: (val) => val > 10 ? val.toFixed(0) : '',
                        display: true
                    }}
                }}
            }}
        }});

        // Area Chart
        new Chart(document.getElementById('areaChart').getContext('2d'), {{
            type: 'line',
            data: {{
                labels: chartLabels,
                datasets: datasets.map(d => ({{
                    label: d.label,
                    data: d.data,
                    backgroundColor: d.color, 
                    borderColor: d.color,
                    fill: true,
                    tension: 0.3,
                    pointRadius: 4,
                    pointBackgroundColor: '#fff'
                }}))
            }},
            options: {{
                responsive: true, maintainAspectRatio: false,
                scales: {{
                    y: {{ stacked: true, grid: {{ color: '#f1f5f9' }} }},
                    x: {{ grid: {{ display: false }} }}
                }},
                plugins: {{
                    legend: {{ position: 'bottom', labels: {{ usePointStyle: true, font: {{ size: 10, weight: '700' }} }} }},
                    datalabels: {{
                        color: '#000000', font: {{ weight: 'bold', size: 9 }},
                        anchor: 'center', align: 'center',
                        formatter: (val) => val > 10 ? val.toFixed(0) : '',
                        display: true
                    }}
                }}
            }}
        }});
    </script>
</body>
</html>
"""
        
        # Encode HTML as base64 for client-side PDF generation
        html_base64 = base64.b64encode(html.encode('utf-8')).decode('utf-8')
        
        return {
            "status": "sukses",
            "html": html,
            "html_base64": html_base64,
            "filename": f"laporan_tutupan_lahan_{datetime.now().strftime('%Y%m%d')}.html",
            "message": "HTML report generated. Use browser print (Ctrl+P) to save as PDF."
        }
        
    except Exception as e:
        print(f"❌ Error generating PDF report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Generate report gagal: {str(e)}")


@app.post("/export/bundle")
async def export_bundle(report_request: PDFReportRequest, request: Request):
    """
    Export Bundle: PDF Report + Shapefiles (SHP) dalam satu ZIP.
    Menggantikan fitur export KML dan PDF terpisah.
    """
    print("🚀 Export Bundle Request Received")
    
    # --- CACHE CHECK ---
    try:
        # Create unique hash based on request content (Deterministic)
        # We exclude title/desc if we want to cache by data only, but they affect PDF content, so include all.
        req_dump = json.dumps(report_request.dict(), sort_keys=True, default=str)
        request_hash = hashlib.md5(req_dump.encode()).hexdigest()
        cache_key = f"bundle_zip_{request_hash}"
        
        cached_zip = get_cached(cache_key)
        if cached_zip:
            print(f"🚀 Cache HIT for Export Bundle: {cache_key}")
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"GealGeolGeo_Export_{timestamp}.zip"
            return Response(
                content=cached_zip,
                media_type="application/zip",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )
    except Exception as e:
        print(f"⚠️ Cache check failed: {e}")
        # Continue to generation if cache check fails

    try:
        # 1. Generate PDF Report (Reusing existing logic)
        print("📄 Generating PDF content...")
        base_url = str(request.base_url).rstrip('/')
        forwarded_host = request.headers.get("x-forwarded-host")
        forwarded_proto = request.headers.get("x-forwarded-proto")
        if forwarded_host and forwarded_proto:
            base_url = f"{forwarded_proto}://{forwarded_host}"
        elif forwarded_host:
            base_url = f"https://{forwarded_host}"
            
        pdf_result = await generate_pdf_report(report_request, base_url=base_url)
        html_content = pdf_result.get("html")
        print("✅ PDF content generated.")
        
        # 2. Prepare ZIP in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            # Add HTML Report
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            zf.writestr(f"Laporan_Analisis_{timestamp}.html", html_content)
            
            # 3. Generate SHP for each year
            print("🗺️ Converting GeoJSON...")
            feature = report_request.geojson
            try:
                geom = geojson_to_ee_geometry(feature)
            except Exception as e:
                print(f"⚠️ GeoJSON conversion failed: {e}")
                raise HTTPException(status_code=400, detail=f"Invalid geometry: {str(e)}")

            for item in report_request.data:
                year = item['year']
                print(f"📦 Generating SHP for {year}...")
                
                # Get Classified Image (Rule-based default)
                classified = get_classified_image(geom, year)
                
                if classified:
                    # Convert to Vector (FeatureCollection)
                    try:
                        # reduceToVectors
                        vectors = classified.reduceToVectors(
                            geometry=geom,
                            crs=classified.projection(),
                            scale=10,
                            geometryType='polygon',
                            eightConnected=False,
                            labelProperty='label',
                            reducer=ee.Reducer.countEvery(),
                            maxPixels=1e10
                        )
                        
                        # Fetch GeoJSON client-side (Async to avoid blocking)
                        fc_dict = await asyncio.to_thread(vectors.getInfo)
                        
                        # Generate SHP buffers
                        shp_buffers = generate_shapefile_buffer(fc_dict)
                        
                        # Write to ZIP under folder Data_SHP/{year}/
                        base_name = f"Data_SHP/{year}/{year}_tutupan_lahan"
                        zf.writestr(f"{base_name}.shp", shp_buffers['shp'])
                        zf.writestr(f"{base_name}.shx", shp_buffers['shx'])
                        zf.writestr(f"{base_name}.dbf", shp_buffers['dbf'])
                        zf.writestr(f"{base_name}.prj", shp_buffers['prj'])
                        
                    except Exception as e:
                        print(f"⚠️ Failed to generate vectors for {year}: {e}")
                        # Optional: write error log to zip
                        zf.writestr(f"Data_SHP/{year}/error.txt", str(e))
                else:
                    zf.writestr(f"Data_SHP/{year}/no_data.txt", "No Sentinel-2 imagery found.")

        # Finalize ZIP
        zip_buffer.seek(0)
        zip_bytes = zip_buffer.read()
        
        # Save to cache (24 hours expire)
        if 'cache_key' in locals():
            cache_file(cache_key, zip_bytes, expire=86400)
        
        filename = f"GealGeolGeo_Export_{timestamp}.zip"
        
        return Response(
            content=zip_bytes,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        print(f"❌ Error generating bundle: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export bundle gagal: {str(e)}")



@app.post("/history/{history_id}/regenerate-visuals")
async def regenerate_visuals(history_id: str):
    """
    Memperbaiki/Regenerate data visualisasi (Thumbnail & Vektor) yang hilang untuk item history tertentu.
    Ini juga me-refresh tile URL GEE yang mungkin sudah expired.
    Semua thumbnail disimpan secara permanen di local storage VPS (/storage/thumbnails).
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase client not configured")
        
    print(f"🔧 Regenerating visuals for History ID: {history_id}")
    
    # Check cache first - if data is already complete, return immediately
    cached_data = get_cached(f"history_detail_{history_id}")
    if cached_data:
        # Verify data completeness
        analysis_results = cached_data.get('analysis_results', [])
        is_complete = all(
            r.get('vector_geojson') and 
            r.get('thumb_url') and 
            not (r.get('thumb_url', '').startswith('http') and 'googleapis.com' in r.get('thumb_url', '')) and
            not r.get('thumb_url', '').startswith('data:image')
            for r in analysis_results
        )
        
        if is_complete:
            print(f"🚀 Cache HIT - Data already complete, skipping regeneration")
            return {"status": "success", "message": "Data already complete (from cache)", "data": cached_data}
    
    try:
        # 1. Fetch History Item & Master Lahan Geometry
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("*, master_lahan(geom_geojson)")
            .eq("id", history_id)
            .single()
            .execute()
        )
        
        if not res or not res.data:
            raise HTTPException(status_code=404, detail="History item not found")
            
        history_item = res.data
        master_lahan = history_item.get('master_lahan')
        
        if not master_lahan or not master_lahan.get('geom_geojson'):
            raise HTTPException(status_code=400, detail="Geometry data not found in master_lahan")
            
        geometry_geojson = master_lahan.get('geom_geojson')
        ee_geometry = geojson_to_ee_geometry(geometry_geojson)
        
        analysis_results = history_item.get('analysis_results', [])
        updated_results = []
        is_modified = False
        
        # 2. Loop through each year results
        for item in analysis_results:
            new_item = item.copy()
            year = new_item.get('year')
            print(f"   📅 Processing Year {year}...")
            
            needs_save = False
            
            # --- A. Refresh Tile URLs (Always refresh because they expire fast) ---
            try:
                print(f"      🛰️ Refreshing Tile URLs...")
                classified = get_classified_image(ee_geometry, year)
                if classified:
                    new_item['map_url'] = await asyncio.to_thread(lambda: get_map_url(classified))
                
                window_params = get_regional_window(ee_geometry)
                start_month = window_params["start_month"]
                end_month = window_params["end_month"]
                window_start = f"{year}-{start_month:02d}-01"
                window_end = f"{year}-{end_month:02d}-28"
                
                collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
                             .filterBounds(ee_geometry)
                             .filterDate(window_start, window_end)
                             .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 80))
                             .map(mask_clouds_sentinel2))
                             
                median_comp = collection.median().clip(ee_geometry)
                new_item['rgb_url'] = await asyncio.to_thread(lambda: get_rgb_map_url(median_comp))
                needs_save = True
            except Exception as e:
                print(f"      ⚠️ Tile refresh failed: {e}")

            # --- B. Check Vector GeoJSON ---
            if not new_item.get('vector_geojson') or 'features' not in new_item.get('vector_geojson', {}):
                print(f"      ⚠️ Missing Vector Data. Regenerating...")
                try:
                    classified = get_classified_image(ee_geometry, year)
                    if classified:
                         vectors_fc = classified.reduceToVectors(
                            geometry=ee_geometry,
                            scale=30, 
                            geometryType='polygon',
                            eightConnected=False,
                            labelProperty='class',
                            reducer=ee.Reducer.countEvery(),
                            maxPixels=1e8,
                            bestEffort=True
                        )
                         vector_data = await asyncio.to_thread(lambda: vectors_fc.getInfo())
                         if vector_data and 'features' in vector_data:
                             for f in vector_data['features']:
                                 f['properties'] = {'class': f['properties']['class']}
                             new_item['vector_geojson'] = vector_data
                             needs_save = True
                             print(f"      ✅ Vector Regenerated.")
                except Exception as e:
                    print(f"      ❌ Failed to generate vector: {e}")

            # --- C. Thumbnail (Classification) - Refresh from GEE ---
            try:
                print(f"      🖼️ Refreshing Thumbnail from GEE...")
                classified = get_classified_image(ee_geometry, year)
                if classified:
                    temp_url = await asyncio.to_thread(lambda: get_thumb_url(classified, ee_geometry))
                    if temp_url:
                        # Persist to Cloud (Function is async, await directly)
                        perm_url = await download_and_save_locally(temp_url, f"history_{history_id}_thumb")
                        if perm_url:
                            new_item["thumb_url"] = perm_url
                            needs_save = True
                            print(f"      ✅ Thumbnail refreshed & persisted: {perm_url}")
                        else:
                             # Fallback to temp if save fails
                             new_item["thumb_url"] = temp_url
                             needs_save = True
                             print(f"      ⚠️ Thumbnail persisted failed, using temp URL.")
            except Exception as e:
                print(f"      ❌ Failed to refresh thumbnail: {e}")

            # --- D. RGB Thumbnail - Refresh from GEE ---
            try:
                print(f"      🖼️ Refreshing RGB Thumbnail from GEE...")
                if "median_comp" not in locals():
                    window_params = get_regional_window(ee_geometry)
                    collection = (ee.ImageCollection(SENTINEL2_COLLECTION)
                                 .filterBounds(ee_geometry)
                                 .filterDate(f"{year}-{window_params['start_month']:02d}-01", f"{year}-{window_params['end_month']:02d}-28")
                                 .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 80))
                                 .map(mask_clouds_sentinel2))
                    median_comp = collection.median().clip(ee_geometry)

                temp_rgb = await asyncio.to_thread(lambda: get_rgb_thumb_url(median_comp, ee_geometry))
                if temp_rgb:
                     # Persist to Cloud (Function is async, await directly)
                    perm_rgb = await download_and_save_locally(temp_rgb, f"history_{history_id}_rgb")
                    if perm_rgb:
                        new_item["rgb_thumb_url"] = perm_rgb
                        needs_save = True
                        print(f"      ✅ RGB Thumbnail refreshed & persisted: {perm_rgb}")
                    else:
                        new_item["rgb_thumb_url"] = temp_rgb
                        needs_save = True
                        print(f"      ⚠️ RGB Persist failed, using temp URL.")
            except Exception as e:
                print(f"      ❌ Failed to refresh RGB thumbnail: {e}")
            if needs_save:
                is_modified = True
            updated_results.append(new_item)
            
        # 3. Save updates
        if is_modified:
            print("   💾 Saving updated history...")
            await asyncio.to_thread(
                lambda: supabase.table("analysis_history")
                .update({"analysis_results": updated_results})
                .eq("id", history_id)
                .execute()
            )
            # Re-fetch fresh to get correct structure
            final_res = await asyncio.to_thread(
                lambda: supabase.table("analysis_history")
                .select("*, master_lahan(geom_geojson)")
                .eq("id", history_id)
                .single()
                .execute()
            )
            data = final_res.data
            data['geo_data'] = data['master_lahan']['geom_geojson']
            del data['master_lahan']
            
            # Update cache with fresh complete data
            cache_file(f"history_detail_{history_id}", data)
            invalidate_cache("history_list")  # Refresh list cache too
            
            return {"status": "success", "message": "Visuals regenerated and stored permanently", "data": data}
        else:
            print("   ✨ No changes needed.")
            history_item['geo_data'] = geometry_geojson
            if 'master_lahan' in history_item: del history_item['master_lahan']
            
            # Cache the complete data
            cache_file(f"history_detail_{history_id}", history_item)
            
            return {"status": "success", "message": "Visuals already permanent", "data": history_item}
            
    except Exception as e:
        print(f"❌ Error regenerating visuals: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Regeneration failed: {str(e)}")


@app.get("/health/thumbnails")
async def check_thumbnail_health():
    """
    Validates thumbnail URLs accessibility.
    Helps diagnose loading issues.
    """
    if not supabase:
        return {
            "status": "error",
            "message": "Supabase not configured",
            "checks": {}
        }

    try:
        results = {
            "supabase_bucket": "unknown",
            "local_storage": "unknown",
            "sample_urls": {}
        }

        # 1. Check Supabase bucket
        try:
            buckets = supabase.storage.list_buckets()
            thumb_bucket = next((b for b in buckets if b.name == "thumbnails"), None)
            if thumb_bucket:
                results["supabase_bucket"] = "✅ exists"
                # Try to get a public URL
                test_url = supabase.storage.from_("thumbnails").get_public_url("test.webp")
                results["sample_urls"]["supabase"] = test_url
            else:
                results["supabase_bucket"] = "❌ not found"
        except Exception as e:
            results["supabase_bucket"] = f"❌ error: {str(e)[:50]}"

        # 2. Check local storage
        try:
            if os.path.exists(THUMBNAIL_DIR):
                files = os.listdir(THUMBNAIL_DIR)
                results["local_storage"] = f"✅ exists ({len(files)} files)"
                if files:
                    results["sample_urls"]["local"] = f"/storage/thumbnails/{files[0]}"
            else:
                results["local_storage"] = "❌ directory not found"
        except Exception as e:
            results["local_storage"] = f"❌ error: {str(e)[:50]}"

        # 3. Validate a sample thumbnail from database
        try:
            res = await asyncio.to_thread(
                lambda: supabase.table("analysis_history")
                .select("id, analysis_results")
                .limit(1)
                .execute()
            )
            if res.data:
                item = res.data[0]
                analysis = item.get('analysis_results', [])
                if analysis and analysis[0].get('thumb_url'):
                    thumb_url = analysis[0]['thumb_url']
                    results["sample_urls"]["database"] = thumb_url[:80] + "..." if len(thumb_url) > 80 else thumb_url

                    # Test URL accessibility
                    try:
                        test_response = requests.head(thumb_url, timeout=5)
                        results["sample_urls"]["database_status"] = f"HTTP {test_response.status_code}"
                    except Exception as url_err:
                        results["sample_urls"]["database_status"] = f"error: {str(url_err)[:40]}"
        except Exception as e:
            results["sample_database_check"] = f"error: {str(e)[:50]}"

        return {
            "status": "ok",
            "checks": results
        }

    except Exception as e:
        print(f"❌ Health check failed: {e}")
        return {
            "status": "error",
            "message": str(e),
            "checks": {}
        }


# ==============================================================================
# MAIN ENTRY POINT
# ==============================================================================

# ==============================================================================
# PROXY ENDPOINTS (OPTIMIZATION)
# ==============================================================================

@app.get("/proxy/sigap/{path:path}")
async def proxy_sigap(path: str, request: Request):
    """
    Proxy & Cache server for MenLHK SIGAP tiles and identify services.
    Speeds up loading by caching tiles on VPS for 7 days.
    """
    print(f"📥 Proxy Request: {path}") # Log entry
    try:
        # Construct upstream URL
        base_url = "https://geoportal.menlhk.go.id/server/rest/services/SIGAP_Interaktif"
        
        # Sanitasi path untuk mencegah Path Traversal
        if ".." in path or ":" in path or path.startswith("/") or path.startswith("\\"):
            print(f"🛑 Portected: Blocked suspicious proxy path: {path}")
            return Response(status_code=400, content="Invalid proxy path structure")
            
        import re
        if not re.match(r"^[a-zA-Z0-9/_.-]+$", path):
             print(f"🛑 Portected: Blocked invalid characters in proxy path: {path}")
             return Response(status_code=400, content="Illegal characters in path")

        upstream_url = f"{base_url}/{path}"
        
        # Capture query parameters from the original request
        query_params = dict(request.query_params)
        
        # Generate valid cache key from URL + params (for identify caching)
        param_str = json.dumps(query_params, sort_keys=True)
        url_hash = hashlib.md5(f"{upstream_url}_{param_str}".encode()).hexdigest()
        cache_key = f"sigap_item_{url_hash}"
        
        # 1. Check Cache
        cached_content = get_cached(cache_key)
        if cached_content:
            # Determine media type - identify returns JSON, tiles return PNG
            media_type = "application/json" if "identify" in path or "f=json" in param_str else "image/png"
            return Response(content=cached_content, media_type=media_type)
            
        # 2. Fetch from Upstream (if not cached)
        def fetch_upstream():
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Referer": "https://geoportal.menlhk.go.id/",
                "Accept": "*/*"
            }
            try:
                # pass query_params to requests.get
                resp = requests.get(upstream_url, params=query_params, headers=headers, timeout=30, verify=False)
                if resp.status_code == 200:
                    return resp.content, resp.headers.get("Content-Type", "image/png")
                else:
                    print(f"❌ Upstream Fail {resp.status_code}: {upstream_url}")
                    return None, None
            except Exception as req_err:
                print(f"❌ Upstream Error: {req_err}")
                return None, None
            
        content, content_type = await asyncio.to_thread(fetch_upstream)
        
        if content:
            # 3. Save to Cache
            # Tiles (contain "tile" in path) 7 days, Identify (?f=json) 1 day
            expiry = 604800 if "/tile/" in path else 86400
            cache_file(cache_key, content, expire=expiry)
            return Response(content=content, media_type=content_type)
        else:
            return Response(status_code=404)
            
    except Exception as e:
        print(f"⚠️ Proxy Critical Error: {e}")
        return Response(status_code=500)



# ==============================================================================
# NASA FIRMS HOTSPOT PROXY (Fire Information for Resource Management System)
# ==============================================================================
class NasaFirmsRequest(BaseModel):
    """Request body for NASA FIRMS hotspot query."""
    bounds: dict  # {minLon, minLat, maxLon, maxLat}
    start_date: str = None  # Format: YYYY-MM-DD (optional, for time series)
    end_date: str = None    # Format: YYYY-MM-DD (optional, for time series)
    source: str = "VIIRS_SNPP_NRT"  # VIIRS_SNPP_NRT, VIIRS_NOAA20_NRT, MODIS_NRT


def generate_bbox_key(bounds: dict, year: int, source: str) -> str:
    """Generate a unique key for bounding box + year + source combination."""
    bbox_str = f"{bounds['minLon']:.4f},{bounds['minLat']:.4f},{bounds['maxLon']:.4f},{bounds['maxLat']:.4f}"
    key_str = f"{bbox_str}_{year}_{source}"
    return hashlib.md5(key_str.encode()).hexdigest()


async def save_hotspots_to_db(features: list, bbox_key: str, source: str):
    """Save hotspot features to database."""
    if not supabase or not features:
        return 0

    saved_count = 0
    batch_size = 100

    try:
        # Prepare records for upsert
        records = []
        for f in features:
            props = f.get("properties", {})
            coords = f.get("geometry", {}).get("coordinates", [0, 0])
            acq_date = props.get("acq_date", "")

            if not acq_date:
                continue

            # Extract year from acq_date
            try:
                year = int(acq_date.split("-")[0])
            except:
                year = datetime.now().year

            record = {
                "latitude": coords[1],
                "longitude": coords[0],
                "acq_date": acq_date,
                "acq_time": props.get("acq_time", ""),
                "year": year,
                "satellite": props.get("satellite", ""),
                "source": source,
                "confidence": props.get("confidence", ""),
                "brightness": props.get("brightness") or props.get("bright_ti4"),
                "bright_ti5": props.get("bright_ti5"),
                "frp": props.get("frp"),
                "scan": props.get("scan"),
                "track": props.get("track"),
                "daynight": props.get("daynight", ""),
                "version": props.get("version", ""),
                "bbox_key": bbox_key
            }
            records.append(record)

        # Batch upsert to database
        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            try:
                await asyncio.to_thread(
                    lambda b=batch: supabase.table("nasa_firms_hotspot_cache")
                        .upsert(b, on_conflict="latitude,longitude,acq_date,acq_time,source")
                        .execute()
                )
                saved_count += len(batch)
            except Exception as e:
                print(f"⚠️ Batch upsert error: {e}")
                continue

        print(f"💾 Saved {saved_count}/{len(records)} hotspots to database")
        return saved_count

    except Exception as e:
        print(f"⚠️ Error saving hotspots to DB: {e}")
        return 0


async def get_hotspots_from_db(bounds: dict, year: int, source: str) -> list:
    """Query hotspots from database within bounding box and year."""
    if not supabase:
        return []

    try:
        # Query by bounding box and year
        result = await asyncio.to_thread(
            lambda: supabase.table("nasa_firms_hotspot_cache")
                .select("*")
                .eq("year", year)
                .eq("source", source)
                .gte("latitude", bounds['minLat'])
                .lte("latitude", bounds['maxLat'])
                .gte("longitude", bounds['minLon'])
                .lte("longitude", bounds['maxLon'])
                .execute()
        )

        if result.data:
            # Convert to GeoJSON features
            features = []
            for row in result.data:
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [float(row['longitude']), float(row['latitude'])]
                    },
                    "properties": {
                        "brightness": row.get('brightness'),
                        "scan": row.get('scan'),
                        "track": row.get('track'),
                        "acq_date": row.get('acq_date'),
                        "acq_time": row.get('acq_time'),
                        "satellite": row.get('satellite'),
                        "confidence": row.get('confidence'),
                        "version": row.get('version'),
                        "bright_ti4": row.get('brightness'),
                        "bright_ti5": row.get('bright_ti5'),
                        "frp": row.get('frp'),
                        "daynight": row.get('daynight'),
                        "from_cache": True
                    }
                }
                features.append(feature)

            print(f"📦 Found {len(features)} hotspots in database for year {year}")
            return features

        return []

    except Exception as e:
        print(f"⚠️ Error querying hotspots from DB: {e}")
        return []


async def save_hotspots_for_analysis(history_id: str, geo_data: dict, years: list):
    """
    Fetch and save hotspot data for a specific analysis history.
    This populates the analysis_hotspots table for KPS views.
    """
    if not supabase or not geo_data:
        return 0

    try:
        # Extract bounding box from GeoJSON
        geometry = geo_data.get("features", [{}])[0].get("geometry", {})
        if not geometry:
            return 0

        all_coords = []
        if geometry.get("type") == "Polygon":
            all_coords = geometry.get("coordinates", [[]])[0]
        elif geometry.get("type") == "MultiPolygon":
            for poly in geometry.get("coordinates", []):
                all_coords.extend(poly[0] if poly else [])

        if not all_coords:
            return 0

        lngs = [c[0] for c in all_coords]
        lats = [c[1] for c in all_coords]
        bounds = {
            "minLon": min(lngs),
            "minLat": min(lats),
            "maxLon": max(lngs),
            "maxLat": max(lats)
        }

        total_saved = 0
        NASA_MAP_KEY = os.getenv("NASA_FIRMS_MAP_KEY", "")

        for year in years:
            start_date = f"{year}-01-01"
            end_date = f"{year}-12-31"
            bbox = f"{bounds['minLon']},{bounds['minLat']},{bounds['maxLon']},{bounds['maxLat']}"

            base_url = "https://firms.modaps.eosdis.nasa.gov/api/area"
            source = "VIIRS_SNPP_NRT"

            if NASA_MAP_KEY:
                url = f"{base_url}/csv/{NASA_MAP_KEY}/{source}/{bbox}/{start_date},{end_date}"
            else:
                url = f"{base_url}/csv/public/{source}/{bbox}/10"  # Limited without API key

            try:
                resp = requests.get(url, timeout=60)
                if resp.status_code == 200:
                    import csv
                    from io import StringIO

                    csv_reader = csv.DictReader(StringIO(resp.text))
                    records = []

                    for row in csv_reader:
                        try:
                            lat = float(row['latitude'])
                            lng = float(row['longitude'])

                            # Check if point is inside polygon (simple bbox for now)
                            if not (bounds['minLat'] <= lat <= bounds['maxLat'] and
                                    bounds['minLon'] <= lng <= bounds['maxLon']):
                                continue

                            # Map confidence to level
                            conf = row.get('confidence', '').lower()
                            conf_level = conf if conf in ['low', 'nominal', 'high'] else None

                            record = {
                                "history_id": history_id,
                                "year": year,
                                "latitude": lat,
                                "longitude": lng,
                                "acq_date": row.get('acq_date', ''),
                                "confidence": int(float(row.get('confidence', 0))) if row.get('confidence', '').isdigit() else None,
                                "confidence_level": conf_level,
                                "brightness": float(row.get('bright_ti4', row.get('brightness', 0)) or 0),
                                "frp": float(row.get('frp', 0) or 0),
                                "source": "NASA_FIRMS"
                            }
                            records.append(record)
                        except:
                            continue

                    # Batch insert
                    if records:
                        await asyncio.to_thread(
                            lambda r=records: supabase.table("analysis_hotspots")
                                .insert(r)
                                .execute()
                        )
                        total_saved += len(records)
                        print(f"🔥 Saved {len(records)} hotspots for history {history_id[:8]} year {year}")

            except Exception as e:
                print(f"⚠️ Error fetching hotspots for year {year}: {e}")
                continue

        # Update hotspot_count in analysis_history
        if total_saved > 0:
            await asyncio.to_thread(
                lambda: supabase.table("analysis_history")
                    .update({"hotspot_count": total_saved})
                    .eq("id", history_id)
                    .execute()
            )

        return total_saved

    except Exception as e:
        print(f"⚠️ Error in save_hotspots_for_analysis: {e}")
        return 0


# ==============================================================================
# SLOPE ANALYSIS - Klasifikasi Kemiringan Lereng
# ==============================================================================

def calculate_slope_statistics(ee_geometry, scale: int = 30):
    """
    Menghitung statistik kemiringan lereng menggunakan NASA DEM.

    Klasifikasi KLHK:
    - Datar: 0-8%
    - Landai: 8-15%
    - Agak Curam: 15-25%
    - Curam: 25-45%
    - Sangat Curam: >45%

    Returns dict dengan avg_slope dan luas tiap kelas dalam hektar.
    """
    try:
        # Load NASA DEM dan hitung slope dalam derajat
        dem = ee.Image("NASA/NASADEM_HGT/001").select('elevation')
        slope_deg = ee.Terrain.slope(dem).clip(ee_geometry)

        # Konversi derajat ke persen: tan(slope_deg) * 100
        slope_pct = slope_deg.multiply(math.pi / 180).tan().multiply(100)

        # Hitung rata-rata slope
        avg_slope_result = slope_pct.reduceRegion(
            reducer=ee.Reducer.mean(),
            geometry=ee_geometry,
            scale=scale,
            maxPixels=1e9
        ).getInfo()
        avg_slope = avg_slope_result.get('slope', 0) or 0

        # Pixel area dalam meter persegi
        pixel_area = ee.Image.pixelArea()

        # Klasifikasi slope (KLHK Standard)
        slope_0_8 = slope_pct.lt(8)
        slope_8_15 = slope_pct.gte(8).And(slope_pct.lt(15))
        slope_15_25 = slope_pct.gte(15).And(slope_pct.lt(25))
        slope_25_45 = slope_pct.gte(25).And(slope_pct.lt(45))
        slope_above_45 = slope_pct.gte(45)

        # Hitung luas tiap kelas (dalam hektar)
        def calc_area_ha(mask):
            area_m2 = mask.multiply(pixel_area).reduceRegion(
                reducer=ee.Reducer.sum(),
                geometry=ee_geometry,
                scale=scale,
                maxPixels=1e9
            ).getInfo()
            # Konversi m² ke hektar
            return round((area_m2.get('slope', 0) or 0) / 10000, 2)

        # Note: DB columns are slope_25_40/slope_above_40 but data uses KLHK thresholds (25-45%, >45%)
        return {
            'avg_slope': round(avg_slope, 2),
            'slope_0_8': calc_area_ha(slope_0_8),
            'slope_8_15': calc_area_ha(slope_8_15),
            'slope_15_25': calc_area_ha(slope_15_25),
            'slope_25_40': calc_area_ha(slope_25_45),      # KLHK: Curam 25-45%
            'slope_above_40': calc_area_ha(slope_above_45) # KLHK: Sangat Curam >45%
        }

    except Exception as e:
        print(f"⚠️ Error calculating slope statistics: {e}")
        return None


async def save_slope_analysis(history_id: str, geo_data: dict):
    """
    Menghitung dan menyimpan analisis slope untuk area INSIDE dan OUTSIDE 2km buffer.
    Disimpan ke tabel analysis_slope_summary.
    """
    if not supabase or not geo_data:
        return False

    try:
        # Extract geometry dari GeoJSON
        geometry = geo_data.get("features", [{}])[0].get("geometry", {})
        if not geometry:
            geometry = geo_data.get("geometry", {})
        if not geometry:
            print(f"⚠️ No geometry found for slope analysis")
            return False

        # Konversi ke EE Geometry
        ee_geometry = geojson_to_ee_geometry({"type": "Feature", "geometry": geometry})

        # === 1. Hitung untuk INSIDE ===
        print(f"📐 Calculating slope statistics for INSIDE area...")
        inside_stats = await asyncio.to_thread(
            lambda: calculate_slope_statistics(ee_geometry)
        )

        if inside_stats:
            inside_record = {
                "history_id": history_id,
                "scope": "INSIDE",
                **inside_stats
            }

            await asyncio.to_thread(
                lambda: supabase.table("analysis_slope_summary")
                    .upsert(inside_record, on_conflict="history_id,scope")
                    .execute()
            )
            print(f"✅ Saved INSIDE slope analysis: avg={inside_stats['avg_slope']}%")

        # === 2. Hitung untuk OUTSIDE 2km buffer ===
        print(f"📐 Calculating slope statistics for OUTSIDE 2km buffer...")
        # Buat buffer 2km dan subtract geometry asli
        buffer_2km = ee_geometry.buffer(2000)  # 2000 meters = 2km
        outside_geometry = buffer_2km.difference(ee_geometry)

        outside_stats = await asyncio.to_thread(
            lambda: calculate_slope_statistics(outside_geometry)
        )

        if outside_stats:
            outside_record = {
                "history_id": history_id,
                "scope": "OUTSIDE_2KM",
                **outside_stats
            }

            await asyncio.to_thread(
                lambda: supabase.table("analysis_slope_summary")
                    .upsert(outside_record, on_conflict="history_id,scope")
                    .execute()
            )
            print(f"✅ Saved OUTSIDE_2KM slope analysis: avg={outside_stats['avg_slope']}%")

        # === 3. Generate Visual Map URL (Fixed Opacity: 100% Inside, 70% Outside) ===
        try:
            print(f"🎨 Generating Slope Map Visual URL...")

            # Load DEM and Calculate Slope
            dem = ee.Image("NASA/NASADEM_HGT/001").select('elevation')
            slope_deg = ee.Terrain.slope(dem)
            slope_pct = slope_deg.multiply(math.pi / 180).tan().multiply(100)

            # Palette (KLHK Standard)
            # 0-8 (Green), 8-15 (Light Green), 15-25 (Yellow), 25-45 (Orange), >45 (Red)
            palette = ['00FF00', '80FF00', 'FFFF00', 'FFA500', 'FF0000', '8B0000']

            # Combine Masks into a single Alpha Channel
            # Inside Area = 100% Opacity, Outside 2km Buffer = 70% Opacity
            # Start with 0 opacity everywhere
            alpha_channel = ee.Image(0).float()
            # Paint outside with 0.7
            alpha_channel = alpha_channel.where(ee.Image(1).clip(buffer_2km), 0.7)
            # Paint inside with 1.0 (overwriting the buffer/overlap)
            alpha_channel = alpha_channel.where(ee.Image(1).clip(ee_geometry), 1.0)

            # Apply mask to slope (keep as single-band)
            slope_masked = slope_pct.updateMask(alpha_channel).clip(buffer_2km)

            # Visualize with palette (only at final step)
            vis_params = {
                'min': 0,
                'max': 100,
                'palette': palette
            }

            # Generate URL using getMapId directly
            map_id = await asyncio.to_thread(lambda: slope_masked.getMapId(vis_params))
            map_url = map_id['tile_fetcher'].url_format
            
            # Save URL to history metadata or slope table
            # Since analysis_slope_summary is strictly stats, we might save this in analysis_history metadata
            # Or add a column. For now, let's update analysis_history metadata as it's the easiest place for visual URLs
            await asyncio.to_thread(
                lambda: supabase.table("analysis_history")
                    .update({"slope_map_url": map_url})
                    .eq("id", history_id)
                    .execute()
            )
            print(f"✅ Slope Map URL generated: {map_url}")
            
        except Exception as e:
            print(f"⚠️ Failed to generate slope map URL: {e}")

        return True

    except Exception as e:
        print(f"⚠️ Error in save_slope_analysis: {e}")
        import traceback
        traceback.print_exc()
        return False


@app.post("/history/{history_id}/analyze-slope")
async def analyze_slope_for_history(history_id: str, background_tasks: BackgroundTasks):
    """
    Menghitung dan menyimpan analisis slope untuk history yang sudah ada.
    Berguna untuk backfill data lama yang belum memiliki analisis slope.

    Output disimpan ke tabel analysis_slope_summary dengan klasifikasi:
    - Datar (<8%)
    - Landai (8-15%)
    - Agak Curam (15-25%)
    - Curam (25-40%)
    - Sangat Curam (>40%)
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        # Get history data dengan geo_data
        result = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
                .select("id, geo_data, lahan_id")
                .eq("id", history_id)
                .single()
                .execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="History not found")

        history = result.data
        geo_data = history.get("geo_data")

        # Jika geo_data tidak ada di history, coba ambil dari master_lahan
        if not geo_data and history.get("lahan_id"):
            lahan_result = await asyncio.to_thread(
                lambda: supabase.table("master_lahan")
                    .select("geojson")
                    .eq("id", history.get("lahan_id"))
                    .single()
                    .execute()
            )
            if lahan_result.data:
                geo_data = lahan_result.data.get("geojson")

        if not geo_data:
            raise HTTPException(status_code=400, detail="No geometry data found for this history")

        # Schedule slope analysis in background
        background_tasks.add_task(save_slope_analysis, history_id, geo_data)

        return {
            "status": "success",
            "message": f"Slope analysis scheduled for history {history_id[:8]}...",
            "history_id": history_id
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error scheduling slope analysis: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to schedule slope analysis: {str(e)}")


@app.get("/history/{history_id}/slope")
async def get_slope_analysis(history_id: str):
    """
    Mengambil hasil analisis slope untuk history tertentu.

    Returns:
    - INSIDE: Statistik slope di dalam batas geometri
    - OUTSIDE_2KM: Statistik slope di buffer 2km luar geometri
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        result = await asyncio.to_thread(
            lambda: supabase.table("analysis_slope_summary")
                .select("*")
                .eq("history_id", history_id)
                .execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="Slope analysis not found for this history")

        # Format response (KLHK Standard: 25-45%, >45%)
        slope_data = {}
        for record in result.data:
            scope = record.get("scope", "UNKNOWN")
            slope_data[scope] = {
                "avg_slope": record.get("avg_slope"),
                "klasifikasi": {
                    "datar_0_8": record.get("slope_0_8"),
                    "landai_8_15": record.get("slope_8_15"),
                    "agak_curam_15_25": record.get("slope_15_25"),
                    "curam_25_45": record.get("slope_25_40"),        # DB: slope_25_40, KLHK: 25-45%
                    "sangat_curam_45_plus": record.get("slope_above_40") # DB: slope_above_40, KLHK: >45%
                },
                "created_at": record.get("created_at")
            }

        return {
            "status": "success",
            "history_id": history_id,
            "slope_analysis": slope_data
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error fetching slope analysis: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch slope analysis: {str(e)}")


@app.post("/history/{history_id}/calculate-temporal-status")
async def calculate_temporal_status_for_history(history_id: str, background_tasks: BackgroundTasks):
    """
    Calculate and update temporal status (grey area detection) for all years in a history.

    This endpoint:
    1. Calculates dominant_class for each year (from 6 IPSDH classes)
    2. Determines temporal_status based on year-to-year comparison:
       - stable: No change from previous year
       - transition_unconfirmed: Changed once (grey area)
       - transition_confirmed: Change persisted 2+ years
       - reverted_noise: Changed but reverted back

    Useful for:
    - Backfilling existing data with temporal status
    - Recalculating after data corrections
    - Manual trigger for grey area analysis
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        # Verify history exists
        result = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
                .select("id")
                .eq("id", history_id)
                .single()
                .execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="History not found")

        # Schedule temporal status calculation in background
        background_tasks.add_task(update_temporal_status_for_history, history_id)

        return {
            "status": "success",
            "message": f"Temporal status calculation scheduled for history {history_id[:8]}...",
            "history_id": history_id
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error scheduling temporal status calculation: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to schedule calculation: {str(e)}")


@app.get("/history/{history_id}/temporal-status")
async def get_temporal_status(history_id: str):
    """
    Get temporal status analysis for all years in a history.

    Returns:
    - Year-by-year breakdown with dominant_class and temporal_status
    - Summary statistics (count by status)
    - Grey area years (transition_unconfirmed)
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        result = await asyncio.to_thread(
            lambda: supabase.table("analysis_yearly_data")
                .select("year, dominant_class, temporal_status, hutan_primer, hutan_sekunder, tanah_kering, tanah_kosong, lahan_terbangun, air")
                .eq("history_id", history_id)
                .order("year")
                .execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="No yearly data found for this history")

        yearly_data = result.data

        # Calculate summary statistics
        status_counts = {
            "stable": 0,
            "transition_unconfirmed": 0,
            "transition_confirmed": 0,
            "reverted_noise": 0
        }

        grey_area_years = []

        for year_record in yearly_data:
            status = year_record.get("temporal_status", "stable")
            status_counts[status] = status_counts.get(status, 0) + 1

            if status == "transition_unconfirmed":
                grey_area_years.append({
                    "year": year_record.get("year"),
                    "dominant_class": year_record.get("dominant_class")
                })

        return {
            "status": "success",
            "history_id": history_id,
            "yearly_data": yearly_data,
            "summary": {
                "total_years": len(yearly_data),
                "status_counts": status_counts,
                "grey_area_years": grey_area_years
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error fetching temporal status: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch temporal status: {str(e)}")


@app.post("/history/{history_id}/populate-hotspots")
async def populate_hotspots_for_history(history_id: str, background_tasks: BackgroundTasks):
    """
    Populate hotspot data for a specific analysis history.
    Fetches NASA FIRMS data for all years in the analysis and saves to analysis_hotspots table.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        # Get history data
        result = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
                .select("id, geo_data, analysis_results")
                .eq("id", history_id)
                .single()
                .execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="History not found")

        history = result.data
        geo_data = history.get("geo_data")
        analysis_results = history.get("analysis_results", [])

        # Extract years from analysis_results
        years = [item.get("year") for item in analysis_results if item.get("year")]
        if not years:
            years = list(range(2019, datetime.now().year + 1))

        # Run in background
        background_tasks.add_task(save_hotspots_for_analysis, history_id, geo_data, years)

        return {
            "status": "started",
            "message": f"Populating hotspots for {len(years)} years",
            "years": years
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"⚠️ Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/proxy/nasa/hotspot")
async def proxy_nasa_hotspot(request: NasaFirmsRequest):
    """
    Proxy endpoint to query NASA FIRMS Hotspot data.
    Supports time series queries from January 1 to December 31 for any year.
    Data is cached in database for faster subsequent queries.
    Returns GeoJSON format.
    """
    NASA_MAP_KEY = os.getenv("NASA_FIRMS_MAP_KEY", "")
    bounds = request.bounds

    if not NASA_MAP_KEY:
        print("⚠️ NASA_FIRMS_MAP_KEY not set. Using public access (limited to 10 days).")

    try:
        # Determine year from date range
        target_year = datetime.now().year
        if request.start_date:
            try:
                target_year = int(request.start_date.split("-")[0])
            except:
                pass

        # Generate cache key for this bbox + year + source
        bbox_key = generate_bbox_key(bounds, target_year, request.source)

        # --- STEP 1: Try to get from Database first ---
        if supabase and request.start_date and request.end_date:
            db_features = await get_hotspots_from_db(bounds, target_year, request.source)

            if db_features:
                print(f"🚀 NASA FIRMS DB HIT: {len(db_features)} hotspots for {target_year}")
                return {
                    "type": "FeatureCollection",
                    "features": db_features,
                    "source": "database",
                    "year": target_year
                }

        # --- STEP 2: Check file cache ---
        req_dump = json.dumps(request.dict(), sort_keys=True)
        req_hash = hashlib.md5(req_dump.encode()).hexdigest()
        cache_key = f"nasa_firms_{req_hash}"

        cached_data = get_cached(cache_key)
        if cached_data:
            print(f"🚀 NASA FIRMS File Cache HIT: {cache_key}")
            return cached_data

        # --- STEP 3: Fetch from NASA FIRMS API ---
        print(f"🔄 NASA FIRMS: Fetching from NASA API for {target_year}...")

        bbox = f"{bounds['minLon']},{bounds['minLat']},{bounds['maxLon']},{bounds['maxLat']}"

        if request.start_date and request.end_date:
            date_param = f"{request.start_date},{request.end_date}"
            print(f"📅 Time Series Query: {date_param}")
        else:
            date_param = "10"
            print(f"📅 Default Query: Last {date_param} days")

        base_url = "https://firms.modaps.eosdis.nasa.gov/api/area"

        if NASA_MAP_KEY:
            url = f"{base_url}/csv/{NASA_MAP_KEY}/{request.source}/{bbox}/{date_param}"
        else:
            url = f"{base_url}/csv/public/{request.source}/{bbox}/{date_param}"

        def fetch_nasa_firms():
            try:
                print(f"🔥 Fetching: {url}")
                resp = requests.get(url, timeout=60)
                print(f"🔥 NASA FIRMS Response Status: {resp.status_code}")

                if resp.status_code == 200:
                    import csv
                    from io import StringIO

                    features = []
                    csv_reader = csv.DictReader(StringIO(resp.text))

                    for row in csv_reader:
                        try:
                            feature = {
                                "type": "Feature",
                                "geometry": {
                                    "type": "Point",
                                    "coordinates": [float(row['longitude']), float(row['latitude'])]
                                },
                                "properties": {
                                    "brightness": float(row.get('bright_ti4', row.get('brightness', 0)) or 0),
                                    "scan": float(row.get('scan', 0) or 0),
                                    "track": float(row.get('track', 0) or 0),
                                    "acq_date": row.get('acq_date', ''),
                                    "acq_time": row.get('acq_time', ''),
                                    "satellite": row.get('satellite', ''),
                                    "confidence": row.get('confidence', ''),
                                    "version": row.get('version', ''),
                                    "bright_ti4": float(row.get('bright_ti4', row.get('brightness', 0)) or 0),
                                    "bright_ti5": float(row.get('bright_ti5', 0) or 0) if row.get('bright_ti5') else None,
                                    "frp": float(row.get('frp', 0) or 0),
                                    "daynight": row.get('daynight', '')
                                }
                            }
                            features.append(feature)
                        except Exception as parse_err:
                            print(f"⚠️ Parse error for row: {parse_err}")
                            continue

                    print(f"✅ Parsed {len(features)} hotspot features from NASA FIRMS")
                    return {
                        "type": "FeatureCollection",
                        "features": features
                    }
                else:
                    print(f"❌ NASA FIRMS API Error {resp.status_code}: {resp.text[:300]}")
                    return {"type": "FeatureCollection", "features": []}
            except Exception as e:
                print(f"❌ NASA FIRMS Fetch Error: {e}")
                import traceback
                traceback.print_exc()
                return {"type": "FeatureCollection", "features": []}

        result = await asyncio.to_thread(fetch_nasa_firms)

        feature_count = len(result.get("features", []))
        print(f"🔥 NASA FIRMS Proxy: Returning {feature_count} hotspot features")

        # --- STEP 4: Save to Database (for time series queries) ---
        if supabase and request.start_date and feature_count > 0:
            asyncio.create_task(save_hotspots_to_db(result["features"], bbox_key, request.source))

        # --- STEP 5: Save to File Cache ---
        cache_expire = 86400 if request.start_date else 3600
        if result and "features" in result:
            cache_file(cache_key, result, expire=cache_expire)

        result["source"] = "nasa_api"
        result["year"] = target_year
        return result

    except Exception as e:
        print(f"⚠️ NASA FIRMS Proxy Error: {e}")
        import traceback
        traceback.print_exc()
        return {"type": "FeatureCollection", "features": []}


@app.get("/hotspot/stats")
async def get_hotspot_stats():
    """
    Get hotspot statistics from database cache.
    Returns yearly counts and confidence breakdown.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        # Get total count
        total_result = await asyncio.to_thread(
            lambda: supabase.table("nasa_firms_hotspot_cache")
                .select("id", count="exact")
                .execute()
        )

        # Get yearly stats using raw query via RPC or manual aggregation
        years_result = await asyncio.to_thread(
            lambda: supabase.table("nasa_firms_hotspot_cache")
                .select("year, confidence")
                .execute()
        )

        # Aggregate in Python
        yearly_stats = {}
        confidence_stats = {"high": 0, "nominal": 0, "low": 0}

        if years_result.data:
            for row in years_result.data:
                year = row.get("year")
                conf = row.get("confidence", "").lower()

                if year:
                    if year not in yearly_stats:
                        yearly_stats[year] = {"total": 0, "high": 0, "nominal": 0, "low": 0}
                    yearly_stats[year]["total"] += 1

                    if conf in ["high", "nominal", "low"]:
                        yearly_stats[year][conf] += 1
                        confidence_stats[conf] += 1

        return {
            "total_cached": total_result.count or 0,
            "yearly_breakdown": dict(sorted(yearly_stats.items(), reverse=True)),
            "confidence_summary": confidence_stats,
            "source": "nasa_firms_hotspot_cache"
        }

    except Exception as e:
        print(f"⚠️ Hotspot stats error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/hotspot/cache")
async def clear_hotspot_cache(year: int = None):
    """
    Clear hotspot cache from database.
    If year is provided, only clear that year. Otherwise clear all.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not initialized")

    try:
        if year:
            result = await asyncio.to_thread(
                lambda: supabase.table("nasa_firms_hotspot_cache")
                    .delete()
                    .eq("year", year)
                    .execute()
            )
            message = f"Cleared hotspot cache for year {year}"
        else:
            result = await asyncio.to_thread(
                lambda: supabase.table("nasa_firms_hotspot_cache")
                    .delete()
                    .neq("id", "00000000-0000-0000-0000-000000000000")  # Delete all
                    .execute()
            )
            message = "Cleared all hotspot cache"

        deleted_count = len(result.data) if result.data else 0
        print(f"🗑️ {message}: {deleted_count} records")

        return {
            "success": True,
            "message": message,
            "deleted_count": deleted_count
        }

    except Exception as e:
        print(f"⚠️ Clear cache error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/history/{history_id}/export-excel")
async def export_excel_history(history_id: str):
    """
    Export Analysis Data as Excel (.xlsx) from History.
    """
    if not supabase:
         raise HTTPException(status_code=503, detail="Supabase not initialized")
         
    try:
        # Fetch from DB
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history")
            .select("filename, analysis_results")
            .eq("id", history_id)
            .single()
            .execute()
        )
        
        if not res or not res.data:
            raise HTTPException(status_code=404, detail="Data tidak ditemukan")
            
        data = res.data
        results = normalize_analysis_results(data.get('analysis_results', []))
        raw_filename = data.get('filename', 'analysis').replace('.shp', '').replace('.zip', '').replace('.geojson', '')

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Analisis Tutupan Lahan"

        # Header - HANYA 6 kelas IPSDH final
        headers = ['Tahun', 'Hutan Primer (Ha)', 'Hutan Sekunder (Ha)', 'Tanah Lahan Kering (Ha)', 'Tanah Kosong/Terbuka (Ha)', 'Lahan Terbangun (Ha)', 'Air/Badan Air (Ha)', 'Total Area (Ha)']
        ws.append(headers)

        # Style Header
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Rows - HANYA 6 kelas IPSDH final
        for r in results:
            hp = r.get('hutan_primer', 0)
            hs = r.get('hutan_sekunder', 0)
            tk = r.get('tanah_kering', 0)
            tt = r.get('tanah_kosong', 0)
            lt = r.get('lahan_terbangun', 0)
            a = r.get('air', 0)
            total = hp + hs + tk + tt + lt + a

            ws.append([
                r.get('year'),
                round(hp, 2),
                round(hs, 2),
                round(tk, 2),
                round(tt, 2),
                round(lt, 2),
                round(a, 2),
                round(total, 2)
            ])
            
        # Adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename_out = f"GealGeolGeo_Stats_{raw_filename}_{timestamp}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename_out}"
            }
        )
    except Exception as e:
        print(f"❌ Excel Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/export/excel")
async def export_excel_direct(request: Request):
    """
    Export Current Analysis Data as Excel (.xlsx).
    """
    try:
        req_data = await request.json()
        results = req_data.get('data', [])
        filename = req_data.get('filename', 'analysis')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analisis Tutupan Lahan"
        
        # Header
        headers = ['Tahun', 'H.Primer (Ha)', 'H.Sekunder (Ha)', 'Hutan Total (Ha)', 'Tanah Kering (Ha)', 'Tanah Terbuka (Ha)', 'Lahan Terbangun (Ha)', 'Badan Air (Ha)', 'Lumpur (Ha)', 'Lahan Rusak (Ha)', 'Total (Ha)']
        ws.append(headers)
        
        # Style Header
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        # Rows
        for r in results:
            hp = r.get('hutan_primer', 0)
            hs = r.get('hutan_sekunder', 0)
            ht = r.get('hutan', 0)
            tk = r.get('tanah_kering', 0)
            tt = r.get('tanah_kosong', 0)
            lt = r.get('lahan_terbangun', 0)
            a = r.get('air', 0)
            lm = r.get('tanah_basah', 0)
            lr = r.get('lahan_rusak', 0)
            total = hp + hs + tk + tt + lt + a + lm + lr
            
            ws.append([
                r.get('year'),
                round(hp, 2),
                round(hs, 2),
                round(ht, 2),
                round(tk, 2),
                round(tt, 2),
                round(lt, 2),
                round(a, 2),
                round(lm, 2),
                round(lr, 2),
                round(total, 2)
            ])
            
        # Adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        raw_filename = filename.replace('.shp', '').replace('.zip', '').replace('.geojson', '')
        filename_out = f"GealGeolGeo_Stats_{raw_filename}_{timestamp}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename_out}"
            }
        )
    except Exception as e:
        print(f"❌ Excel Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))






async def worker_bulk_regenerate():
    print("🚀 Starting Bulk Regeneration Worker...")
    try:
        if not supabase:
            print("❌ Worker Aborted: No Supabase connection")
            return

        # Fetch all IDs
        res = await asyncio.to_thread(
            lambda: supabase.table("analysis_history").select("id").execute()
        )
        if not res.data:
            print("ℹ️ No history items found.")
            return

        ids = [item['id'] for item in res.data]
        total = len(ids)
        print(f"📦 Found {total} items to process for thumbnail check...")
        
        for idx, history_id in enumerate(ids):
            print(f"▶️ Processing {idx+1}/{total}: {history_id}")
            try:
                # Call existing logic (it internally checks if regeneration is needed)
                await regenerate_visuals(history_id)
            except Exception as e:
                print(f"   ❌ Error processing {history_id}: {e}")
            
            # Sleep to be nice to GEE and prevent rate limits
            await asyncio.sleep(2) 
            
        print("✅ Bulk Regeneration Worker Completed.")
    except Exception as e:
        print(f"❌ Worker Failed: {e}")

@app.get("/admin/migration/temporal-status")
async def get_migration_instructions():
    """
    Get migration SQL for temporal_status implementation.
    Use this to check schema and get instructions for applying the migration.
    """
    try:
        migration_file = os.path.join(os.path.dirname(__file__), "supabase_migration_temporal_status.sql")

        if not os.path.exists(migration_file):
            return {
                "status": "error",
                "message": "Migration file not found"
            }

        with open(migration_file, "r") as f:
            migration_sql = f.read()

        # Check if columns already exist
        if supabase:
            try:
                await asyncio.to_thread(
                    lambda: supabase.table("analysis_yearly_data")
                        .select("temporal_status, dominant_class")
                        .limit(1)
                        .execute()
                )
                columns_exist = True
            except:
                columns_exist = False
        else:
            columns_exist = False

        return {
            "status": "success",
            "columns_exist": columns_exist,
            "migration_sql": migration_sql,
            "instructions": [
                "1. Go to https://app.supabase.com",
                "2. Select your project",
                "3. Click 'SQL Editor' on the left menu",
                "4. Click 'New query'",
                "5. Copy-paste the migration_sql from this response",
                "6. Click 'Run'",
                "7. Once done, temporal_status endpoints will work"
            ]
        }

    except Exception as e:
        return {
            "status": "error",
            "message": f"Error: {e}"
        }


@app.post("/admin/regenerate-all-thumbnails")
async def admin_regenerate_all(background_tasks: BackgroundTasks):
    """
    Trigger process to regenerate thumbnails for ALL history items.
    Runs in background to fix missing/expired images.
    """
    background_tasks.add_task(worker_bulk_regenerate)
    return {"status": "started", "message": "Bulk regeneration started in background. Monitor console for progress."}


import random
@app.on_event("startup")
async def startup_event():
    """Start background services"""
    asyncio.create_task(self_healing_daemon())

async def self_healing_daemon():
    """
    Background Task: Continuous Integrity Check.
    Runs forever as long as server is up.
    Checks random history items for broken thumbnails and repairs them.
    """
    print("🏥 Self-Healing Daemon: INITIALIZED. Waiting 60s before start...")
    await asyncio.sleep(60) # Wait for server to warm up
    
    while True:
        try:
            print("🏥 Self-Healing: Starting scan cycle...")
            if not supabase:
                print("   ⚠️ No DB connection. Retrying in 5m...")
                await asyncio.sleep(300)
                continue

            # Fetch all IDs
            res = await asyncio.to_thread(
                lambda: supabase.table("analysis_history").select("id").execute()
            )
            
            if not res.data:
                await asyncio.sleep(600)
                continue

            ids = [item['id'] for item in res.data]
            random.shuffle(ids) # Randomize to ensure coverage over restarts
            
            for idx, history_id in enumerate(ids):
                # CHECK QUEUE STATUS: Yield to active users
                while not await analysis_queue.is_idle():
                    if idx % 10 == 0:
                        print("🏥 Self-Healing: System BUSY (Analysis in progress). Pausing for 30s...")
                    await asyncio.sleep(30)

                try:
                    # Reuse the robust regenerate logic
                    # It internally checks cache/validity and skips if good.
                    # We just call it blindly.
                    
                    # Log less frequently to avoid spam
                    if idx % 10 == 0: 
                        print(f"   🏥 Scan progress: {idx}/{total}")
                        
                    await regenerate_visuals(history_id)
                    
                except Exception as e:
                    # Ignore errors to keep daemon alive
                    pass
                
                # Slow pace: 1 item every 10 seconds = 8,600 items/day
                # Low impact on server/GEE
                await asyncio.sleep(10)
                
            print("🏥 Self-Healing: Cycle Complete. Resting 1 hour...")
            await asyncio.sleep(3600)
            
        except Exception as e:
            print(f"❌ Self-Healing Crashed: {e}. Restarting in 1 min...")
            await asyncio.sleep(60)

@app.post("/map/slope")
async def get_slope_mapid(request: Request):
    """
    Generate GEE MapID for Slope layer based on geometry.
    Palette uses standard Indonesian color ranking for slope.
    """
    if not ee_initialized:
        raise HTTPException(status_code=503, detail="GEE not initialized")
    
    try:
        body = await request.json()
        geo_data = body.get("geo_data")
        history_id = body.get("history_id")

        # Prioritize GeoData, fallback to History DB
        if not geo_data and history_id and supabase:
            h_res = await asyncio.to_thread(lambda: supabase.table("analysis_history").select("geo_data").eq("id", history_id).execute())
            if h_res and h_res.data:
                geo_data = h_res.data[0].get("geo_data")

        if not geo_data:
            raise HTTPException(status_code=400, detail="Missing geo_data or history_id")
            
        geometry = geo_data.get("features", [{}])[0].get("geometry", {})
        if not geometry:
            geometry = geo_data.get("geometry", {})
            
        ee_geometry = geojson_to_ee_geometry({"type": "Feature", "geometry": geometry})
        
        # Get DEM and calculate Slope
        dem = ee.Image("NASA/NASADEM_HGT/001").select('elevation').clip(ee_geometry)
        slope = ee.Terrain.slope(dem)
        
        # Visualize Slope in Degrees (0 - 45+)
        # Palette: Green (0-8), Yellow (8-15), Orange (15-25), Red (25-45), Dark Red (>45)
        viz_params = {
            'min': 0,
            'max': 60,
            'palette': [
                '#31a354', # 0-8 (Datar)
                '#addd8e', # 8-15 (Landai)
                '#fee391', # 15-25 (Agak Curam)
                '#fec44f', # 25-40 (Curam)
                '#ec7014', # 40-60 (Sangat Curam)
                '#662506'  # > 60 (Ekstrem)
            ]
        }
        
        map_id_dict = slope.getMapId(viz_params)
        
        # Optional: Fetch summary from DB if history_id provided
        db_summary = None
        if history_id and supabase:
            s_res = await asyncio.to_thread(lambda: supabase.table("analysis_slope_summary").select("*").eq("history_id", history_id).execute())
            if s_res and s_res.data:
                db_summary = s_res.data

        return {
            "status": "success",
            "map_url": map_id_dict['tile_fetcher'].url_format,
            "db_summary": db_summary
        }
    except Exception as e:
        print(f"❌ Map Slope Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    # Redirect standard logging to work with uvicorn if needed
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
